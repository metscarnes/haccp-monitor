"""
routes_vente.py — Catalogue de VENTE (produits finis fabriqués)

Le catalogue de vente est la source des « produits finis » des recettes
(sortie de production, étiquette). Il est indépendant du catalogue interne
(produits) et du catalogue d'achats (catalogue_fournisseur).

GET    /api/vente/catalogue            → liste des produits finis (filtres)
GET    /api/vente/catalogue/{id}       → détail
POST   /api/vente/catalogue            → créer un produit fini
PUT    /api/vente/catalogue/{id}       → modifier
DELETE /api/vente/catalogue/{id}       → désactiver (ou supprimer si permanent=true)
GET    /api/vente/catalogue/template   → télécharger template Excel
POST   /api/vente/catalogue/import     → importer depuis Excel
"""

import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from src.api.routes_auth import require_admin
from src.database import get_db

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/vente", tags=["vente"])


# ---------------------------------------------------------------------------
# Schémas Pydantic
# ---------------------------------------------------------------------------

class ProduitVenteCreate(BaseModel):
    nom: str
    code_vente: Optional[str] = None
    prix_vente_ttc: Optional[float] = None
    tva_percent: Optional[float] = 5.5
    dlc_jours: int = 3
    temperature_conservation: Optional[str] = "0°C à +4°C"
    format_etiquette: Optional[str] = "standard_60x40"
    famille: Optional[str] = None
    sous_famille: Optional[str] = None


class ProduitVenteUpdate(BaseModel):
    nom: Optional[str] = None
    code_vente: Optional[str] = None
    prix_vente_ttc: Optional[float] = None
    tva_percent: Optional[float] = None
    dlc_jours: Optional[int] = None
    temperature_conservation: Optional[str] = None
    format_etiquette: Optional[str] = None
    famille: Optional[str] = None
    sous_famille: Optional[str] = None
    actif: Optional[bool] = None


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/catalogue")
async def liste_catalogue_vente(
    q: Optional[str] = Query(None),
    actif_only: bool = Query(True),
):
    async with get_db() as db:
        sql = "SELECT * FROM catalogue_vente WHERE boutique_id = 1"
        params: list = []
        if actif_only:
            sql += " AND actif = 1"
        if q:
            sql += " AND nom LIKE ?"
            params.append(f"%{q}%")
        sql += " ORDER BY famille, sous_famille, nom"
        cur = await db.execute(sql, params)
        return [dict(r) for r in await cur.fetchall()]


@router.get("/catalogue/export")
async def export_catalogue_vente():
    """Exporte le catalogue vente en Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl requis")

    async with get_db() as db:
        cur = await db.execute(
            """SELECT * FROM catalogue_vente
               WHERE boutique_id = 1
               ORDER BY famille, sous_famille, nom"""
        )
        produits = [dict(r) for r in await cur.fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue Vente"

    cols = [
        ("nom",                      "Nom du produit fini"),
        ("code_vente",               "Code de vente"),
        ("famille",                  "Famille"),
        ("sous_famille",             "Sous-famille"),
        ("prix_vente_ttc",           "Prix vente TTC (€)"),
        ("tva_percent",              "TVA (%)"),
        ("dlc_jours",                "DLC (jours)"),
        ("temperature_conservation", "Température de conservation"),
        ("format_etiquette",         "Format étiquette"),
        ("actif",                    "Actif"),
    ]

    header_fill = PatternFill("solid", fgColor="8B1A1A")
    for col_idx, (_, label) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 24

    for row_idx, p in enumerate(produits, 2):
        for col_idx, (key, _) in enumerate(cols, 1):
            val = p.get(key)
            if key == "actif":
                val = "Oui" if val else "Non"
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")

    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(cols)).column_letter}1"
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=catalogue_vente.xlsx"},
    )


@router.get("/catalogue/template")
async def download_template_vente():
    """Télécharger le template Excel d'import catalogue vente."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.workbook.defined_name import DefinedName
    except ImportError:
        raise HTTPException(500, "openpyxl requis pour générer le template")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue vente"

    colonnes = [
        ("nom",                      "Nom du produit fini",
         "Libellé complet du produit fini. OBLIGATOIRE.",
         "Merguez de bœuf"),
        ("code_vente",               "Code de vente",
         "Référence interne (PLU, code barre…). Facultatif.",
         "MRG-001"),
        ("famille",                  "Famille",
         "Catégorie : Viande | Charcuterie | Traiteur | Aide culinaire | Hygiène et emballage. Facultatif.",
         "Charcuterie"),
        ("sous_famille",             "Sous-famille",
         "Sous-catégorie selon la famille choisie. Facultatif.",
         "Saucisse à cuire et Saucisson cuit"),
        ("prix_vente_ttc",           "Prix de vente TTC (€)",
         "Prix toutes taxes comprises. Facultatif.",
         "3.50"),
        ("tva_percent",              "TVA (%)",
         "5.5 pour l'alimentaire, 10 ou 20 sinon.",
         "5.5"),
        ("dlc_jours",                "DLC (jours)",
         "Durée de vie en jours après fabrication. OBLIGATOIRE.",
         "4"),
        ("temperature_conservation", "Température de conservation",
         "0°C à +4°C | +2°C / +4°C | -18°C | Ambiant",
         "0°C à +4°C"),
        ("format_etiquette",         "Format étiquette",
         "standard_60x40 | grand_100x60 | petit_40x30. Facultatif.",
         "standard_60x40"),
    ]

    header_fill = PatternFill("solid", fgColor="8B1A1A")
    note_fill   = PatternFill("solid", fgColor="FFF0F0")
    ex_fill     = PatternFill("solid", fgColor="FFF8E1")
    wrap_top    = Alignment(horizontal="left", vertical="top", wrap_text=True)

    FORMATS_NUM = {
        "prix_vente_ttc": "0.00",
        "tva_percent":    "0.0",
        "dlc_jours":      "0",
    }

    def _to_num(v):
        if v in (None, ""):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return v

    DERNIERE_LIGNE = 500

    for col_idx, (key, header, note, exemple) in enumerate(colonnes, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        est_num = key in FORMATS_NUM

        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(f"Colonne technique : {key}", "HACCP")

        note_cell = ws.cell(row=2, column=col_idx, value=note)
        note_cell.fill = note_fill
        note_cell.font = Font(italic=True, size=9, color="8B1A1A")
        note_cell.alignment = wrap_top

        v3 = _to_num(exemple) if est_num else exemple
        c3 = ws.cell(row=3, column=col_idx, value=v3)
        c3.fill = ex_fill
        c3.font = Font(size=10, color="8A6D3B")
        c3.alignment = wrap_top

        if est_num:
            for r in range(3, DERNIERE_LIGNE + 1):
                ws.cell(row=r, column=col_idx).number_format = FORMATS_NUM[key]

        ws.column_dimensions[letter].width = 22

    ws.cell(row=3, column=len(colonnes) + 2,
            value="◀ Exemple de produit fini").font = Font(italic=True, size=9, color="8A6D3B")
    ws.row_dimensions[2].height = 52
    ws.freeze_panes = "A4"

    # ── Onglet Listes (caché) : sous-familles par colonne + plages nommées ──
    FAMILLES_SF = {
        'Viande':               ['Boeuf', 'Veau', 'Agneau', 'Porc', 'Volaille', 'Cheval'],
        'Charcuterie':          ['Jambon', 'Pâté, Terrine et Rillette', 'Salaison et Pièce séchée',
                                 'Saucisse à cuire et Saucisson cuit', 'Spécialité charcutière'],
        'Traiteur':             ['Crudité', 'Fromage', 'Plat préparé', 'Accompagnement', 'Pané', 'Dessert'],
        'Aide culinaire':       ['Épices et Aromates', 'Boyaux et Ficellerie', 'Marinades, Sauces et huile',
                                 'Bases et Liants', 'Fruits secs et Inclusions', 'Alcools de cuisson'],
        'Hygiène et emballage': ['Hygiène', 'Emballage'],
    }
    ws_listes = wb.create_sheet("Listes")
    ws_listes.sheet_state = "hidden"
    for col_l, (fam, sfs) in enumerate(FAMILLES_SF.items(), 1):
        for row_l, sf in enumerate(sfs, 1):
            ws_listes.cell(row=row_l, column=col_l, value=sf)
        col_letter = ws_listes.cell(row=1, column=col_l).column_letter
        range_ref  = f"Listes!${col_letter}$1:${col_letter}${len(sfs)}"
        safe_name  = fam.replace(" ", "_").replace("é", "e").replace("è", "e").replace("ê", "e").replace("î", "i").replace("ô", "o").replace("â", "a").replace("û", "u")
        dn = DefinedName(safe_name, attr_text=range_ref)
        wb.defined_names.add(dn)

    # Validation famille
    fam_col    = next(i for i, (k, *_) in enumerate(colonnes, 1) if k == "famille")
    fam_letter = ws.cell(row=1, column=fam_col).column_letter
    dv_fam = DataValidation(
        type="list",
        formula1='"Viande,Charcuterie,Traiteur,Aide culinaire,Hygiène et emballage"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Famille invalide", error="Choisissez une famille dans la liste.",
    )
    ws.add_data_validation(dv_fam)
    dv_fam.add(f"{fam_letter}4:{fam_letter}500")

    # Validation sous-famille : liste dépendante de la famille via INDIRECT
    sf_col    = next(i for i, (k, *_) in enumerate(colonnes, 1) if k == "sous_famille")
    sf_letter = ws.cell(row=1, column=sf_col).column_letter
    dv_sf = DataValidation(
        type="list",
        formula1=f'INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({fam_letter}4,"é","e"),"è","e")," ","_"))',
        allow_blank=True, showErrorMessage=False,
    )
    dv_sf.promptTitle = "Sous-famille"
    dv_sf.prompt = "Choisissez d'abord une famille (colonne Famille)"
    ws.add_data_validation(dv_sf)
    dv_sf.add(f"{sf_letter}4:{sf_letter}500")

    tva_col    = next(i for i, (k, *_) in enumerate(colonnes, 1) if k == "tva_percent")
    tva_letter = ws.cell(row=1, column=tva_col).column_letter
    dv_tva = DataValidation(
        type="list", formula1='"5,5;10;20"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="TVA invalide", error="Choisissez 5,5 ; 10 ou 20.",
    )
    ws.add_data_validation(dv_tva)
    dv_tva.add(f"{tva_letter}4:{tva_letter}500")

    temp_col    = next(i for i, (k, *_) in enumerate(colonnes, 1) if k == "temperature_conservation")
    temp_letter = ws.cell(row=1, column=temp_col).column_letter
    dv_temp = DataValidation(
        type="list",
        formula1='"0°C à +4°C,+2°C / +4°C,-18°C,Ambiant"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Température invalide", error="Choisissez une valeur dans la liste.",
    )
    ws.add_data_validation(dv_temp)
    dv_temp.add(f"{temp_letter}4:{temp_letter}500")

    fmt_col    = next(i for i, (k, *_) in enumerate(colonnes, 1) if k == "format_etiquette")
    fmt_letter = ws.cell(row=1, column=fmt_col).column_letter
    dv_fmt = DataValidation(
        type="list",
        formula1='"standard_60x40,grand_100x60,petit_40x30"',
        allow_blank=True, showErrorMessage=True,
        errorTitle="Format invalide", error="Choisissez un format dans la liste.",
    )
    ws.add_data_validation(dv_fmt)
    dv_fmt.add(f"{fmt_letter}4:{fmt_letter}500")

    guide = wb.create_sheet("Mode d'emploi")
    guide.column_dimensions["A"].width = 100
    lignes_guide = [
        ("CATALOGUE DE VENTE — MODE D'EMPLOI", True, "FFFFFF", "8B1A1A", 13),
        ("", False, "333333", None, 11),
        ("VUE D'ENSEMBLE", True, "8B1A1A", "FFF0F0", 11),
        ("• Ce fichier contient 2 onglets : « Catalogue vente » (vos données) et ce mode d'emploi.", False, "333333", "FFF0F0", 11),
        ("• Une ligne = un produit fini. Commencez à saisir à partir de la LIGNE 4.", False, "333333", "FFF0F0", 11),
        ("• La ligne 3 est un exemple — vous pouvez l'écraser ou la supprimer.", False, "333333", "FFF0F0", 11),
        ("• Si un produit existe déjà (même nom), il sera mis à jour.", False, "333333", "FFF0F0", 11),
        ("", False, "333333", None, 11),
        ("COLONNE PAR COLONNE", True, "8B1A1A", None, 12),
        ("", False, "333333", None, 11),
        ("1. Nom du produit fini  [OBLIGATOIRE]", True, "8B1A1A", None, 11),
        ("   Libellé complet et lisible. C'est la clé d'identification à l'import.", False, "333333", None, 11),
        ("   Exemples : « Merguez de bœuf », « Pâté de campagne maison », « Rôti de porc »", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("2. Code de vente  [facultatif]", True, "8B1A1A", None, 11),
        ("   Référence interne, PLU caisse ou code barre. Laissez vide si non applicable.", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("3. Famille  [facultatif]", True, "8B1A1A", None, 11),
        ("   Catégorie principale (liste déroulante) :", False, "333333", None, 11),
        ("   Viande | Charcuterie | Traiteur | Aide culinaire | Hygiène et emballage", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("4. Sous-famille  [facultatif]", True, "8B1A1A", None, 11),
        ("   Sous-catégorie selon la famille choisie.", False, "333333", None, 11),
        ("   Exemples pour Viande : Boeuf | Veau | Agneau | Porc | Volaille | Cheval", False, "8A6D3B", "FFF8E1", 10),
        ("   Exemples pour Charcuterie : Jambon | Pâté, Terrine et Rillette | Salaison...", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("5. Prix de vente TTC (€)  [facultatif]", True, "8B1A1A", None, 11),
        ("   Prix affiché en rayon, toutes taxes comprises.", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("6. TVA (%)  [facultatif, défaut = 5.5]", True, "8B1A1A", None, 11),
        ("   5.5 pour les produits alimentaires de base, 10 pour la restauration, 20 sinon.", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("7. DLC (jours)  [OBLIGATOIRE]", True, "8B1A1A", None, 11),
        ("   Durée de vie en jours à compter de la date de fabrication.", False, "333333", None, 11),
        ("   Exemples : 3 (merguez fraîches), 7 (pâté sous-vide), 30 (saucisson sec)", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("8. Température de conservation  [facultatif, défaut = 0°C à +4°C]", True, "8B1A1A", None, 11),
        ("   Valeurs autorisées (liste déroulante) :", False, "333333", None, 11),
        ("   0°C à +4°C | +2°C / +4°C | -18°C | Ambiant", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("9. Format étiquette  [facultatif, défaut = standard_60x40]", True, "8B1A1A", None, 11),
        ("   standard_60x40 | grand_100x60 | petit_40x30", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("Merci ! En cas de doute, contactez-nous.", True, "8B1A1A", "FFF0F0", 11),
    ]
    for i, (txt, bold, color, bg, size) in enumerate(lignes_guide, 1):
        c = guide.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        guide.row_dimensions[i].height = 18 if txt else 8

    wb.active = wb.sheetnames.index("Catalogue vente")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_catalogue_vente.xlsx"},
    )


@router.post("/catalogue/import", status_code=200)
async def import_catalogue_vente(fichier: UploadFile = File(...), _=Depends(require_admin)):
    """Import Excel catalogue vente — UPSERT sur le nom du produit."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl requis")

    content = await fichier.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["Catalogue vente"] if "Catalogue vente" in wb.sheetnames else wb.active

    LIBELLE_VERS_CLE = {
        "nom du produit fini": "nom", "nom": "nom",
        "code de vente": "code_vente", "code_vente": "code_vente",
        "famille": "famille",
        "sous-famille": "sous_famille", "sous famille": "sous_famille", "sous_famille": "sous_famille",
        "prix de vente ttc (€)": "prix_vente_ttc", "prix vente ttc": "prix_vente_ttc",
        "tva (%)": "tva_percent", "tva": "tva_percent",
        "dlc (jours)": "dlc_jours", "dlc jours": "dlc_jours", "dlc_jours": "dlc_jours",
        "température de conservation": "temperature_conservation",
        "temperature_conservation": "temperature_conservation",
        "format étiquette": "format_etiquette", "format_etiquette": "format_etiquette",
    }

    def _cle(libelle):
        return LIBELLE_VERS_CLE.get(libelle.strip().lower(), libelle.strip().lower())

    headers = [_cle(str(ws.cell(row=1, column=c).value or "")) for c in range(1, ws.max_column + 1)]
    if "nom" not in headers:
        raise HTTPException(400, "Colonne « Nom du produit fini » manquante")

    def col(row, name):
        if name not in headers:
            return ""
        idx = headers.index(name) + 1
        v = ws.cell(row=row, column=idx).value
        return str(v).strip() if v is not None else ""

    TEMP_VALIDES   = {"0°C à +4°C", "+2°C / +4°C", "-18°C", "Ambiant"}
    FORMAT_VALIDES = {"standard_60x40", "grand_100x60", "petit_40x30"}
    EXEMPLE_NOMS   = {"merguez de bœuf", "merguez de boeuf"}
    stats = {"crees": 0, "mis_a_jour": 0, "erreurs": []}

    async with get_db() as db:
        for row_num in range(4, ws.max_row + 1):
            nom = col(row_num, "nom")
            if not nom or nom.strip().lower() in EXEMPLE_NOMS:
                continue

            try:
                dlc = int(float(col(row_num, "dlc_jours") or 3))
            except ValueError:
                dlc = 3

            prix_raw = col(row_num, "prix_vente_ttc")
            try:
                prix = float(prix_raw.replace(",", ".")) if prix_raw else None
            except ValueError:
                prix = None

            tva_raw = col(row_num, "tva_percent")
            try:
                tva = float(tva_raw.replace(",", ".")) if tva_raw else 5.5
            except ValueError:
                tva = 5.5

            temp = col(row_num, "temperature_conservation") or "0°C à +4°C"
            if temp not in TEMP_VALIDES:
                temp = "0°C à +4°C"

            fmt = col(row_num, "format_etiquette") or "standard_60x40"
            if fmt not in FORMAT_VALIDES:
                fmt = "standard_60x40"

            famille      = col(row_num, "famille") or None
            sous_famille = col(row_num, "sous_famille") or None
            code_vente   = col(row_num, "code_vente") or None

            cur = await db.execute(
                "SELECT id FROM catalogue_vente WHERE boutique_id = 1 AND LOWER(TRIM(nom)) = LOWER(TRIM(?))",
                (nom,)
            )
            existing = await cur.fetchone()
            if existing:
                await db.execute(
                    """UPDATE catalogue_vente
                       SET code_vente=?, prix_vente_ttc=?, tva_percent=?, dlc_jours=?,
                           temperature_conservation=?, format_etiquette=?,
                           famille=?, sous_famille=?
                       WHERE id=?""",
                    (code_vente, prix, tva, dlc, temp, fmt, famille, sous_famille, existing["id"])
                )
                stats["mis_a_jour"] += 1
            else:
                await db.execute(
                    """INSERT INTO catalogue_vente
                           (boutique_id, nom, code_vente, prix_vente_ttc, tva_percent,
                            dlc_jours, temperature_conservation, format_etiquette,
                            famille, sous_famille)
                       VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (nom, code_vente, prix, tva, dlc, temp, fmt, famille, sous_famille)
                )
                stats["crees"] += 1
        await db.commit()

    return stats


@router.get("/catalogue/{produit_id}")
async def detail_produit_vente(produit_id: int):
    async with get_db() as db:
        cur = await db.execute("SELECT * FROM catalogue_vente WHERE id = ?", (produit_id,))
        row = await cur.fetchone()
    if not row:
        raise HTTPException(404, "Produit de vente introuvable")
    return dict(row)


@router.post("/catalogue", status_code=201)
async def creer_produit_vente(body: ProduitVenteCreate, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute(
            """INSERT INTO catalogue_vente
                   (boutique_id, nom, code_vente, prix_vente_ttc, tva_percent,
                    dlc_jours, temperature_conservation, format_etiquette, famille, sous_famille)
               VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (body.nom, body.code_vente, body.prix_vente_ttc, body.tva_percent,
             body.dlc_jours, body.temperature_conservation, body.format_etiquette,
             body.famille, body.sous_famille),
        )
        await db.commit()
        cur2 = await db.execute("SELECT * FROM catalogue_vente WHERE id = ?", (cur.lastrowid,))
        return dict(await cur2.fetchone())


@router.put("/catalogue/{produit_id}")
async def modifier_produit_vente(produit_id: int, body: ProduitVenteUpdate, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute("SELECT id FROM catalogue_vente WHERE id = ?", (produit_id,))
        if not await cur.fetchone():
            raise HTTPException(404, "Produit de vente introuvable")

        fields = {k: v for k, v in body.model_dump().items() if v is not None}
        if not fields:
            raise HTTPException(400, "Aucun champ à modifier")

        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [produit_id]
        await db.execute(f"UPDATE catalogue_vente SET {set_clause} WHERE id = ?", values)
        await db.commit()

        cur2 = await db.execute("SELECT * FROM catalogue_vente WHERE id = ?", (produit_id,))
        return dict(await cur2.fetchone())


@router.delete("/catalogue/{produit_id}", status_code=200)
async def supprimer_produit_vente(
    produit_id: int, permanent: bool = Query(False), _=Depends(require_admin)
):
    async with get_db() as db:
        if permanent:
            # Une recette peut référencer ce produit fini : on délie avant suppression.
            await db.execute(
                "UPDATE recettes SET catalogue_vente_id = NULL WHERE catalogue_vente_id = ?",
                (produit_id,),
            )
            await db.execute("DELETE FROM catalogue_vente WHERE id = ?", (produit_id,))
        else:
            await db.execute("UPDATE catalogue_vente SET actif = 0 WHERE id = ?", (produit_id,))
        await db.commit()
    return {"ok": True}
