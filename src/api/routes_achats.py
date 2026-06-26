"""
routes_achats.py — Module Achats : Fournisseurs, Catalogue, Commandes

GET    /api/achats/fournisseurs                          → liste fournisseurs enrichis
POST   /api/achats/fournisseurs                          → créer fournisseur
PUT    /api/achats/fournisseurs/{id}                     → modifier fournisseur

GET    /api/achats/catalogue                             → catalogue fournisseur (filtres)
GET    /api/achats/catalogue/{id}                        → détail article
POST   /api/achats/catalogue                             → créer article
PUT    /api/achats/catalogue/{id}                        → modifier article
DELETE /api/achats/catalogue/{id}                        → désactiver article
POST   /api/achats/catalogue/import                      → import Excel
GET    /api/achats/catalogue/template                    → télécharger template Excel

GET    /api/achats/commandes                             → liste commandes (filtres)
GET    /api/achats/commandes/{id}                        → détail commande + lignes
POST   /api/achats/commandes                             → créer commande
PUT    /api/achats/commandes/{id}                        → modifier commande
POST   /api/achats/commandes/{id}/envoyer                → envoyer mail fournisseur
POST   /api/achats/commandes/{id}/dupliquer              → dupliquer commande
GET    /api/achats/commandes/{id}/lignes                 → lignes de commande
POST   /api/achats/commandes/{id}/lignes                 → ajouter ligne
PUT    /api/achats/commandes/{id}/lignes/{ligne_id}      → modifier ligne
DELETE /api/achats/commandes/{id}/lignes/{ligne_id}      → supprimer ligne

GET    /api/achats/factures                              → liste factures (filtres)
GET    /api/achats/factures/{id}                         → détail facture + lignes + écarts
POST   /api/achats/factures/depuis-reception/{rid}       → facture pré-remplie (réception + commande)
POST   /api/achats/factures                              → créer facture manuelle
PUT    /api/achats/factures/{id}                         → modifier entête (n°, date, statut)
PUT    /api/achats/factures/{id}/lignes/{ligne_id}       → saisir poids/prix facturé → recalcul écarts
DELETE /api/achats/factures/{id}/lignes/{ligne_id}       → supprimer ligne
DELETE /api/achats/factures/{id}                         → supprimer facture

GET    /api/achats/panier                                → lignes du panier sauvegardé
PUT    /api/achats/panier                                → sauvegarder le panier
DELETE /api/achats/panier                                → vider le panier
POST   /api/achats/panier/generer                        → générer les commandes (1/fournisseur)
GET    /api/achats/panier/references                     → lignes d'achat ⭐ du comparatif
GET    /api/achats/panier/suggestions                    → suggestions (récurrence + score)
GET    /api/achats/panier/cadencier                      → cadencier par semaine/mois

GET    /api/achats/pilotage/ca                            → historique CA (filtre période)
GET    /api/achats/pilotage/ca/{date_ca}                  → CA d'un jour (ou null)
POST   /api/achats/pilotage/ca                            → enregistrer/corriger le CA d'un jour (upsert)
GET    /api/achats/pilotage/ca/stats/resume               → totaux & moyennes (mois courant + 30j)
GET    /api/achats/pilotage/ca/stats/comparatif            → période vs période décalée (j/s/m/a × n)
GET    /api/achats/pilotage/ca/stats/comparer-dates        → comparaison de deux jours précis
"""

import io
import logging
import unicodedata
from datetime import date, datetime, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel

from src.api.routes_auth import require_admin
from src.database import get_db

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/achats", tags=["achats"])


# ---------------------------------------------------------------------------
# Modèles Pydantic
# ---------------------------------------------------------------------------

class FournisseurCreate(BaseModel):
    nom: str
    nom_commercial: Optional[str] = None
    email_commercial: Optional[str] = None
    emails_copie: Optional[str] = None  # JSON list ex: '["cc1@x.fr","cc2@x.fr"]'
    telephone: Optional[str] = None
    adresse: Optional[str] = None
    conditions_paiement: Optional[str] = None
    delai_paiement_jours: Optional[int] = None
    jours_livraison: Optional[str] = None
    rythme_livraison: Optional[str] = None
    heure_limite_commande: Optional[str] = None
    heure_livraison: Optional[str] = None
    commentaire: Optional[str] = None
    actif: Optional[bool] = True


class FournisseurUpdate(BaseModel):
    nom: Optional[str] = None
    nom_commercial: Optional[str] = None
    email_commercial: Optional[str] = None
    emails_copie: Optional[str] = None  # JSON list ex: '["cc1@x.fr","cc2@x.fr"]'
    telephone: Optional[str] = None
    adresse: Optional[str] = None
    conditions_paiement: Optional[str] = None
    delai_paiement_jours: Optional[int] = None
    jours_livraison: Optional[str] = None
    rythme_livraison: Optional[str] = None
    heure_limite_commande: Optional[str] = None
    heure_livraison: Optional[str] = None
    commentaire: Optional[str] = None
    actif: Optional[bool] = None


class CatalogueArticleCreate(BaseModel):
    fournisseur_id: int
    code_article: str
    designation: str
    prix_achat_ht: float
    format_prix: Optional[str] = "kg"        # 'kg' (prix au kilo) | 'colis' (prix au colis/pièce)
    qte_par_colis: Optional[float] = None     # nb de pièces par colis (brut, si format 'colis')
    poids_unitaire_kg: Optional[float] = None # poids d'une pièce en kg (brut, si format 'colis')
    tva_percent: Optional[float] = 5.5
    conditionnement: Optional[str] = None
    unites_autorisees: Optional[str] = None   # CSV parmi kg/piece/colis
    famille: Optional[str] = None
    sous_famille: Optional[str] = None
    dlc_type: Optional[str] = "dlc"


class CatalogueArticleUpdate(BaseModel):
    fournisseur_id: Optional[int] = None
    code_article: Optional[str] = None
    designation: Optional[str] = None
    prix_achat_ht: Optional[float] = None
    format_prix: Optional[str] = None
    qte_par_colis: Optional[float] = None
    poids_unitaire_kg: Optional[float] = None
    tva_percent: Optional[float] = None
    conditionnement: Optional[str] = None
    unites_autorisees: Optional[str] = None   # CSV parmi kg/piece/colis
    famille: Optional[str] = None
    sous_famille: Optional[str] = None
    dlc_type: Optional[str] = None
    actif: Optional[bool] = None


class ComparatifGroupeCreate(BaseModel):
    nom: str
    sous_famille: Optional[str] = None


class ComparatifGroupeUpdate(BaseModel):
    nom: Optional[str] = None
    sous_famille: Optional[str] = None


class ComparatifLigneAdd(BaseModel):
    catalogue_fournisseur_id: int


class ComparatifFromCluster(BaseModel):
    nom: str
    catalogue_fournisseur_ids: List[int]


class ComparatifVenteLink(BaseModel):
    # Associer un produit de vente au groupe (1 groupe → N ventes).
    catalogue_vente_id: int


class ComparatifVenteUpdate(BaseModel):
    # Édition d'un produit de vente associé depuis le comparateur (simulation marge) :
    # prix TTC, unité de vente (kg|piece), poids d'une pièce, et classification
    # famille/sous-famille (utile pour les viandes type « Collier » non classées).
    # Champs optionnels ; l'endpoint regarde `model_fields_set` pour ne toucher qu'aux fournis.
    prix_vente_ttc: Optional[float] = None
    unite_vente: Optional[str] = None       # 'kg' | 'piece'
    poids_piece_kg: Optional[float] = None
    famille: Optional[str] = None
    sous_famille: Optional[str] = None


class ComparatifReferenceUpdate(BaseModel):
    # La ligne fournisseur de référence (⭐). null = retirer l'étoile.
    ligne_choisie_id: Optional[int] = None


class CommandeLigneCreate(BaseModel):
    catalogue_fournisseur_id: Optional[int] = None
    code_article: Optional[str] = None
    designation: str
    prix_unitaire_ht: float = 0.0
    quantite_commandee: float = 1.0
    unite: Optional[str] = "kg"
    commentaire_ligne: Optional[str] = None


class CommandeLigneUpdate(BaseModel):
    quantite_commandee: Optional[float] = None
    prix_unitaire_ht: Optional[float] = None
    unite: Optional[str] = None
    commentaire_ligne: Optional[str] = None


class CommandeCreate(BaseModel):
    fournisseur_id: int
    date_commande: Optional[str] = None
    date_livraison_prevue: Optional[str] = None
    commentaire: Optional[str] = None
    personnel_id: Optional[int] = None
    lignes: Optional[List[CommandeLigneCreate]] = []


class CommandeUpdate(BaseModel):
    date_livraison_prevue: Optional[str] = None
    statut: Optional[str] = None
    commentaire: Optional[str] = None


# --- Factures (sous-module rapprochement commande ↔ réception ↔ facture) ---

class FactureUpdate(BaseModel):
    numero_facture: Optional[str] = None
    date_facture: Optional[str] = None
    statut: Optional[str] = None          # brouillon|validee|litige
    commentaire: Optional[str] = None


class FactureLigneUpdate(BaseModel):
    poids_facture_kg: Optional[float] = None
    prix_facture_ht: Optional[float] = None
    statut_ligne: Optional[str] = None    # ok|litige
    commentaire_litige: Optional[str] = None


class FactureLigneCreate(BaseModel):
    catalogue_fournisseur_id: Optional[int] = None
    reception_ligne_id: Optional[int] = None
    code_article: Optional[str] = None
    designation: str
    unite: Optional[str] = "kg"
    poids_recu_kg: Optional[float] = None
    prix_commande_ht: Optional[float] = None
    quantite_commandee: Optional[float] = None
    poids_facture_kg: Optional[float] = None
    prix_facture_ht: Optional[float] = None


class FactureCreate(BaseModel):
    """Création manuelle d'une facture (ligne par ligne) — usage libre.

    Pour le cas standard, utiliser POST /factures/depuis-reception/{reception_id}
    qui pré-remplit les lignes à partir de la réception et de la commande mappée.
    """
    fournisseur_id: int
    reception_id: Optional[int] = None
    commande_id: Optional[int] = None
    numero_facture: Optional[str] = None
    date_facture: Optional[str] = None
    personnel_id: Optional[int] = None
    commentaire: Optional[str] = None
    lignes: Optional[List[FactureLigneCreate]] = []


class CaJournalierUpsert(BaseModel):
    """Saisie (ou correction) du chiffre d'affaires d'un jour, ventilé en deux
    sections (matin 9-13h / soir 16-19h30).

    Une seule ligne par date (UPSERT sur date_ca). Le total du jour est la somme
    des deux sections (calculé côté serveur). Les tickets sont facultatifs et
    servent au calcul du panier moyen (global et par section).
    """
    date_ca:           str                      # YYYY-MM-DD
    montant_ttc_matin: float = 0.0
    nb_tickets_matin:  Optional[int] = None
    montant_ttc_soir:  float = 0.0
    nb_tickets_soir:   Optional[int] = None
    meteo:             Optional[str] = None   # 'soleil' | 'pluie' | None
    commentaire:       Optional[str] = None
    personnel_id:      Optional[int] = None


# ---------------------------------------------------------------------------
# Helpers catalogue
# ---------------------------------------------------------------------------

def _normaliser_format_prix(valeur) -> str:
    """Ramène le format de prix à 'kg' ou 'colis'.

    Tolère les anciennes valeurs ('piece' → 'colis') et les libellés saisis
    par un fournisseur dans le template ('au kilo', 'au colis', 'pièce'...).
    """
    v = (str(valeur) if valeur is not None else "").strip().lower()
    if v in ("kg", "kilo", "au kilo", "kilogramme", "/kg"):
        return "kg"
    if v in ("colis", "piece", "pièce", "au colis", "carton", "unite", "unité"):
        return "colis"
    return "kg"  # défaut prudent : au kilo


def _calc_poids_colis_kg(qte_par_colis, poids_unitaire_kg):
    """Champ généré : poids total d'un colis = qte_par_colis × poids_unitaire_kg.

    Retourne None si une des deux données brutes manque (cas prix au kg, où le
    poids du colis n'est pas fixe).
    """
    try:
        if qte_par_colis in (None, "") or poids_unitaire_kg in (None, ""):
            return None
        return round(float(qte_par_colis) * float(poids_unitaire_kg), 3)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Fournisseurs
# ---------------------------------------------------------------------------

@router.get("/fournisseurs")
async def get_fournisseurs_achats(actif_only: bool = Query(True)):
    async with get_db() as db:
        q = "SELECT * FROM fournisseurs WHERE boutique_id = 1"
        if actif_only:
            q += " AND actif = 1"
        q += " ORDER BY nom"
        cur = await db.execute(q)
        rows = await cur.fetchall()
        fournisseurs = [dict(r) for r in rows]

        # Ajouter le nombre d'articles catalogue pour chaque fournisseur
        for f in fournisseurs:
            cur2 = await db.execute(
                "SELECT COUNT(*) FROM catalogue_fournisseur WHERE fournisseur_id = ? AND actif = 1",
                (f["id"],)
            )
            f["nb_articles"] = (await cur2.fetchone())[0]

        return fournisseurs


@router.post("/fournisseurs", status_code=201)
async def create_fournisseur_achats(body: FournisseurCreate, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute(
            """INSERT INTO fournisseurs
               (boutique_id, nom, nom_commercial, email_commercial, emails_copie, telephone, adresse,
                conditions_paiement, delai_paiement_jours, jours_livraison,
                rythme_livraison, heure_limite_commande, heure_livraison,
                commentaire, actif)
               VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (body.nom, body.nom_commercial, body.email_commercial, body.emails_copie, body.telephone, body.adresse,
             body.conditions_paiement, body.delai_paiement_jours, body.jours_livraison,
             body.rythme_livraison, body.heure_limite_commande, body.heure_livraison,
             body.commentaire, 1 if body.actif else 0)
        )
        await db.commit()
        fid = cur.lastrowid
        cur2 = await db.execute("SELECT * FROM fournisseurs WHERE id = ?", (fid,))
        return dict(await cur2.fetchone())


@router.put("/fournisseurs/{fid}")
async def update_fournisseur_achats(fid: int, body: FournisseurUpdate, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute("SELECT * FROM fournisseurs WHERE id = ?", (fid,))
        existing = await cur.fetchone()
        if not existing:
            raise HTTPException(404, "Fournisseur introuvable")

        dump = body.model_dump()
        # emails_copie peut légitimement être None (vider la liste) → toujours inclure si présent dans le body
        fields = {k: v for k, v in dump.items() if v is not None or k == "emails_copie"}
        if not fields:
            return dict(existing)

        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [fid]
        await db.execute(f"UPDATE fournisseurs SET {set_clause} WHERE id = ?", values)
        await db.commit()

        cur2 = await db.execute("SELECT * FROM fournisseurs WHERE id = ?", (fid,))
        return dict(await cur2.fetchone())


@router.delete("/fournisseurs/{fid}", status_code=200)
async def delete_fournisseur(fid: int, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute("SELECT id, nom FROM fournisseurs WHERE id = ?", (fid,))
        existing = await cur.fetchone()
        if not existing:
            raise HTTPException(404, "Fournisseur introuvable")
        cur2 = await db.execute(
            "SELECT COUNT(*) AS nb FROM catalogue_fournisseur WHERE fournisseur_id = ?", (fid,)
        )
        nb_articles = (await cur2.fetchone())["nb"]
        if nb_articles > 0:
            raise HTTPException(
                409,
                f"Impossible de supprimer : ce fournisseur a {nb_articles} article(s) dans le catalogue. "
                "Supprimez d'abord les articles ou réassignez-les."
            )
        await db.execute("DELETE FROM fournisseurs WHERE id = ?", (fid,))
        await db.commit()
        return {"ok": True, "nom": existing["nom"]}


# ---------------------------------------------------------------------------
# Catalogue fournisseur
# ---------------------------------------------------------------------------

@router.get("/catalogue")
async def get_catalogue(
    fournisseur_id: Optional[int] = Query(None),
    q: Optional[str] = Query(None),
    famille: Optional[str] = Query(None),
    sous_famille: Optional[str] = Query(None),
    actif_only: bool = Query(True),
    avec_stock: bool = Query(False),
):
    async with get_db() as db:
        sql = """
            SELECT c.*, f.nom AS fournisseur_nom
            FROM catalogue_fournisseur c
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            WHERE f.boutique_id = 1
        """
        params = []
        if fournisseur_id:
            sql += " AND c.fournisseur_id = ?"
            params.append(fournisseur_id)
        if actif_only:
            sql += " AND c.actif = 1"
        if q:
            sql += " AND (c.designation LIKE ? OR c.code_article LIKE ?)"
            params += [f"%{q}%", f"%{q}%"]
        if famille:
            sql += " AND c.famille = ?"
            params.append(famille)
        if sous_famille:
            sql += " AND c.sous_famille = ?"
            params.append(sous_famille)
        sql += " ORDER BY f.nom, c.famille, c.sous_famille, c.designation"
        cur = await db.execute(sql, params)
        articles = [dict(r) for r in await cur.fetchall()]

        # Prix au kilo normalisé (None si incalculable) — attendu par le comparateur et
        # tout consommateur de ce catalogue ; les écrans qui l'ignorent ne sont pas affectés.
        for a in articles:
            a["prix_kg"] = _calc_prix_kg(
                a.get("format_prix"), a.get("prix_achat_ht"), a.get("poids_colis_kg"), a.get("famille")
            )

        if avec_stock and articles:
            stocks = await _compter_stock_par_article(db)
            for a in articles:
                a["stock"] = stocks.get(a["id"], 0)

        return articles


async def _compter_stock_par_article(db) -> dict:
    """Compte le nombre de lignes de réception encore en stock, groupées par
    catalogue_fournisseur_id. « En stock » = réception clôturée + conforme +
    non refusée + DLC non dépassée + pas encore sortie (dlc_devenir).

    Une ligne de réception = une unité reçue (carcasse, colis, pièce…) toujours
    présente. Retourne { catalogue_fournisseur_id: nombre }.
    """
    today = date.today().isoformat()
    cur = await db.execute(
        """
        SELECT rl.catalogue_fournisseur_id AS cat_id, COUNT(*) AS nb
        FROM reception_lignes rl
        JOIN receptions r ON r.id = rl.reception_id
        WHERE rl.catalogue_fournisseur_id IS NOT NULL
          AND r.statut = 'cloturee'
          AND rl.conforme = 1
          AND r.livraison_refusee = 0
          AND (COALESCE(rl.dlc, rl.dluo) IS NULL OR COALESCE(rl.dlc, rl.dluo) >= ?)
          AND NOT EXISTS (
              SELECT 1 FROM dlc_devenir dd
              WHERE dd.source_type = 'reception_ligne' AND dd.source_id = rl.id
          )
        GROUP BY rl.catalogue_fournisseur_id
        """,
        (today,),
    )
    return {row["cat_id"]: row["nb"] for row in await cur.fetchall()}


@router.get("/catalogue/export")
async def export_catalogue(fournisseur_id: Optional[int] = Query(None)):
    """Exporte le catalogue fournisseur en Excel (filtre optionnel par fournisseur)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl requis")

    async with get_db() as db:
        sql = """
            SELECT c.*, f.nom AS fournisseur_nom
            FROM catalogue_fournisseur c
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            WHERE f.boutique_id = 1 AND c.actif = 1
        """
        params = []
        if fournisseur_id:
            sql += " AND c.fournisseur_id = ?"
            params.append(fournisseur_id)
        sql += " ORDER BY f.nom, c.famille, c.sous_famille, c.designation"
        cur = await db.execute(sql, params)
        articles = [dict(r) for r in await cur.fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue Fournisseur"

    cols = [
        ("fournisseur_nom",   "Fournisseur"),
        ("code_article",      "Code article"),
        ("designation",       "Désignation"),
        ("prix_achat_ht",     "Prix achat HT (€)"),
        ("format_prix",       "Prix au (kg / colis)"),
        ("qte_par_colis",     "Qté par colis"),
        ("poids_unitaire_kg", "Poids unitaire (kg)"),
        ("poids_colis_kg",    "Poids total colis (kg)"),
        ("tva_percent",       "TVA (%)"),
        ("unites_autorisees",  "Unités de commande"),
        ("famille",           "Famille"),
        ("sous_famille",      "Sous-famille"),
        ("dlc_type",          "Type DLC"),
    ]

    header_fill = PatternFill("solid", fgColor="2D7D46")
    for col_idx, (_, label) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    for row_idx, art in enumerate(articles, 2):
        for col_idx, (key, _) in enumerate(cols, 1):
            ws.cell(row=row_idx, column=col_idx, value=art.get(key) or "")

    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(cols)).column_letter}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "catalogue_fournisseur.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@router.get("/catalogue/template")
async def download_template():
    """Télécharger le template Excel d'import catalogue fournisseur."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl requis pour générer le template")

    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue"

    # Chaque colonne : (clé, en-tête lisible, note d'aide, ex. au kg, ex. au colis)
    colonnes = [
        ("fournisseur_nom",   "Fournisseur",
         "Nom exact de votre société (tel qu'enregistré chez le client).",
         "Boucherie Martin", "Boucherie Martin"),
        ("code_article",      "Code article",
         "Votre référence produit. Obligatoire.",
         "CARC-BF", "STK-185"),
        ("designation",       "Désignation",
         "Libellé complet du produit. Obligatoire.",
         "Carcasse boeuf", "Steak haché 185g"),
        ("prix_achat_ht",     "Prix achat HT (€)",
         "Prix HT. ATTENTION : prix AU KILO ou prix AU COLIS selon la colonne suivante. Obligatoire.",
         "9.70", "18.00"),
        ("format_prix",       "Prix au (kg / colis)",
         "Écrire 'kg' si le prix ci-contre est au kilo, ou 'colis' s'il est pour un colis/une pièce entière. Obligatoire.",
         "kg", "colis"),
        ("qte_par_colis",     "Qté par colis",
         "Nombre de pièces dans un colis. LAISSER VIDE si le prix est au kg.",
         "", "10"),
        ("poids_unitaire_kg", "Poids unitaire (kg)",
         "Poids d'UNE pièce en kg (ex: 0.185 pour 185 g). LAISSER VIDE si le prix est au kg.",
         "", "0.185"),
        ("tva_percent",       "TVA (%)",
         "Taux de TVA : 5.5 pour la viande, 20 sinon.",
         "5.5", "5.5"),
        ("unites_autorisees", "Unités de commande",
         "Unités que le fournisseur accepte. Séparer par virgule parmi : kg, piece, colis. Laisser vide = tout autorisé.",
         "kg,piece,colis", "colis"),
        ("famille",           "Famille",
         "Catégorie : Viande | Charcuterie | Traiteur | Aide culinaire | Hygiène et emballage. Facultatif.",
         "Viande", "Viande"),
        ("sous_famille",      "Sous-famille",
         "Sous-catégorie selon la famille (ex pour Viande : Boeuf, Veau, Agneau, Porc, Volaille, Cheval). Facultatif.",
         "Boeuf", "Boeuf"),
        ("dlc_type",          "Type de DLC",
         "dlc = date limite classique | date_abattage = produit carcasse daté à l'abattage | no_dlc = sans DLC.",
         "date_abattage", "dlc"),
    ]

    header_fill = PatternFill("solid", fgColor="2D7D46")
    note_fill   = PatternFill("solid", fgColor="E8F5E9")
    ex_fill     = PatternFill("solid", fgColor="FFF8E1")
    wrap_top    = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Colonnes numériques → format Excel appliqué aux cellules de données.
    # Les exemples sont convertis en nombre pour ne pas rester du texte.
    FORMATS_NUM = {
        "prix_achat_ht":     "0.00",    # D — prix : 2 décimales
        "qte_par_colis":     "0",       # F — quantité : entier
        "poids_unitaire_kg": "0.000",   # G — poids : 3 décimales
        "tva_percent":       "0.0",     # H — TVA : 1 décimale
    }

    def _to_num(v):
        """Convertit une valeur d'exemple en nombre, ou la laisse vide/texte."""
        if v in (None, ""):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return v

    DERNIERE_LIGNE = 500  # plage de saisie sur laquelle on applique le format

    for col, (key, header, note, ex_kg, ex_colis) in enumerate(colonnes, 1):
        letter = ws.cell(row=1, column=col).column_letter
        est_num = key in FORMATS_NUM

        # Ligne 1 : en-tête lisible
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Le nom technique reste en commentaire (pour l'import, insensible à l'ordre)
        cell.comment = Comment(f"Colonne technique : {key}", "HACCP")

        # Ligne 2 : note d'aide
        note_cell = ws.cell(row=2, column=col, value=note)
        note_cell.fill = note_fill
        note_cell.font = Font(italic=True, size=9, color="2D5A3A")
        note_cell.alignment = wrap_top

        # Lignes 3 et 4 : deux exemples concrets (au kg / au colis)
        v3, v4 = (_to_num(ex_kg), _to_num(ex_colis)) if est_num else (ex_kg, ex_colis)
        c3 = ws.cell(row=3, column=col, value=v3)
        c4 = ws.cell(row=4, column=col, value=v4)
        for c in (c3, c4):
            c.fill = ex_fill
            c.font = Font(size=10, color="8A6D3B")
            c.alignment = wrap_top

        # Format nombre appliqué aux exemples ET à toute la zone de saisie
        if est_num:
            for r in range(3, DERNIERE_LIGNE + 1):
                ws.cell(row=r, column=col).number_format = FORMATS_NUM[key]

        ws.column_dimensions[letter].width = 20

    # Étiquettes des lignes d'exemple, dans une colonne hors tableau (à droite)
    note_col = len(colonnes) + 2
    ws.cell(row=3, column=note_col, value="◀ Exemple : produit vendu AU KILO (carcasse)").font = Font(italic=True, size=9, color="8A6D3B")
    ws.cell(row=4, column=note_col, value="◀ Exemple : produit vendu AU COLIS (10 steaks de 185g)").font = Font(italic=True, size=9, color="8A6D3B")

    ws.row_dimensions[2].height = 58
    ws.freeze_panes = "A5"  # données à saisir à partir de la ligne 5

    # Colonne E — format_prix : kg / colis  [OBLIGATOIRE]
    fmt_letter = ws.cell(row=1, column=5).column_letter
    dv = DataValidation(type="list", formula1='"kg,colis"', allow_blank=False,
                        showErrorMessage=True, errorTitle="Valeur requise",
                        error="Choisissez « kg » ou « colis » dans la liste.")
    dv.prompt = "Choisir : kg (prix au kilo) ou colis (prix au colis/pièce)"
    dv.promptTitle = "Le prix est au..."
    ws.add_data_validation(dv)
    dv.add(f"{fmt_letter}5:{fmt_letter}500")

    # Colonne J — famille  [OBLIGATOIRE]
    fam_col = next(i for i, (k, *_rest) in enumerate(colonnes, 1) if k == "famille")
    fam_letter = ws.cell(row=1, column=fam_col).column_letter
    dv_fam = DataValidation(
        type="list",
        formula1='"Viande,Charcuterie,Traiteur,Aide culinaire,Hygiène et emballage"',
        allow_blank=False,
        showErrorMessage=True, errorTitle="Valeur requise",
        error="Choisissez une famille dans la liste.",
    )
    dv_fam.prompt = "Choisir une famille de produit"
    dv_fam.promptTitle = "Famille"
    ws.add_data_validation(dv_fam)
    dv_fam.add(f"{fam_letter}5:{fam_letter}500")

    # Onglet Listes (caché) : sous-familles par colonne + plages nommées pour INDIRECT
    FAMILLES_SF = {
        'Viande':               ['Boeuf', 'Veau', 'Agneau', 'Porc', 'Volaille', 'Cheval'],
        'Charcuterie':          ['Jambon', 'Pâté, Terrine et Rillette', 'Salaison et Pièce séchée',
                                 'Saucisse à cuire et Saucisson cuit', 'Spécialité charcutière'],
        'Traiteur':             ['Crudité', 'Fromage', 'Plat préparé', 'Accompagnement', 'Pané', 'Dessert'],
        'Aide culinaire':       ['Épices et Aromates', 'Boyaux et Ficellerie', 'Marinades, Sauces et huile',
                                 'Bases et Liants', 'Fruits secs et Inclusions', 'Alcools de cuisson'],
        'Hygiène et emballage': ['Hygiène', 'Emballage'],
    }
    ws_listes = wb.create_sheet("Listes")
    ws_listes.sheet_state = "hidden"
    for col_l, (fam, sfs) in enumerate(FAMILLES_SF.items(), 1):
        for row_l, sf in enumerate(sfs, 1):
            ws_listes.cell(row=row_l, column=col_l, value=sf)
        col_letter_l = ws_listes.cell(row=1, column=col_l).column_letter
        range_ref    = f"Listes!${col_letter_l}$1:${col_letter_l}${len(sfs)}"
        safe_name    = fam.replace(" ", "_").replace("é", "e").replace("è", "e").replace("ê", "e").replace("î", "i").replace("ô", "o").replace("â", "a").replace("û", "u")
        wb.defined_names[safe_name] = DefinedName(safe_name, attr_text=range_ref)

    # Colonne K — sous_famille : liste dépendante de la famille via INDIRECT
    sf_col    = next(i for i, (k, *_rest) in enumerate(colonnes, 1) if k == "sous_famille")
    sf_letter = ws.cell(row=1, column=sf_col).column_letter
    dv_sf = DataValidation(
        type="list",
        formula1=f'INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE({fam_letter}5,"é","e"),"è","e")," ","_"))',
        allow_blank=True, showErrorMessage=False,
    )
    dv_sf.promptTitle = "Sous-famille"
    dv_sf.prompt = "Choisissez d'abord une famille (colonne Famille)"
    ws.add_data_validation(dv_sf)
    dv_sf.add(f"{sf_letter}5:{sf_letter}500")

    # Masquer la colonne L (dlc_type) — avancé, rarement saisi manuellement
    for hidden_key in ("dlc_type",):
        hidden_col = next(i for i, (k, *_rest) in enumerate(colonnes, 1) if k == hidden_key)
        hidden_letter = ws.cell(row=1, column=hidden_col).column_letter
        ws.column_dimensions[hidden_letter].hidden = True

    # --- Colonnes F et G bloquées si E = kg (validation custom) ---
    # Excel évalue la formule sur la cellule cible : autorisé seulement si E de la même ligne = "colis"
    for blocked_key in ("qte_par_colis", "poids_unitaire_kg"):
        b_col = next(i for i, (k, *_rest) in enumerate(colonnes, 1) if k == blocked_key)
        b_letter = ws.cell(row=1, column=b_col).column_letter
        dv_block = DataValidation(
            type="custom",
            formula1=f'=EXACT(E5,"colis")',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Colonne non applicable",
            error='Cette colonne n\'est à remplir que si "Prix au" (colonne E) vaut "colis". Laissez vide pour un produit au kilo.',
        )
        dv_block.showInputMessage = True
        dv_block.promptTitle = "Au kilo ?"
        dv_block.prompt = "Laisser vide si le prix est au kilo (colonne E = kg)."
        ws.add_data_validation(dv_block)
        dv_block.add(f"{b_letter}5:{b_letter}500")


    # --- Onglet « Mode d'emploi » -------------------------------------------
    guide = wb.create_sheet("Mode d'emploi")
    guide.column_dimensions["A"].width = 100
    title_fill  = PatternFill("solid", fgColor="2D7D46")
    section_fill = PatternFill("solid", fgColor="E8F5E9")
    ex_fill2    = PatternFill("solid", fgColor="FFF8E1")

    lignes_guide = [
        # (texte, gras, couleur_texte, couleur_fond, taille)
        ("COMMENT REMPLIR CE CATALOGUE — MODE D'EMPLOI", True,  "FFFFFF", "2D7D46", 13),
        ("", False, "333333", None, 11),
        ("VUE D'ENSEMBLE", True, "2D7D46", "E8F5E9", 11),
        ("• Ce fichier contient 2 onglets : « Catalogue » (vos données) et ce mode d'emploi.", False, "333333", "E8F5E9", 11),
        ("• Une ligne = un produit. Commencez à saisir à partir de la LIGNE 5 de l'onglet « Catalogue ».", False, "333333", "E8F5E9", 11),
        ("• Les lignes 3 et 4 sont des exemples — vous pouvez les écraser ou les supprimer.", False, "333333", "E8F5E9", 11),
        ("• Les colonnes marquées OBLIGATOIRE doivent être remplies, les autres sont facultatives.", False, "333333", "E8F5E9", 11),
        ("", False, "333333", None, 11),
        ("COLONNE PAR COLONNE", True, "2D7D46", None, 12),
        ("", False, "333333", None, 11),
        ("1. Fournisseur  [OBLIGATOIRE]", True, "2D5A3A", None, 11),
        ("   Le nom exact de votre société, tel qu'il est enregistré dans notre système.", False, "333333", None, 11),
        ("   Exemple : « Boucherie Martin »", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("2. Code article  [OBLIGATOIRE]", True, "2D5A3A", None, 11),
        ("   Votre référence interne du produit. Doit être unique par fournisseur.", False, "333333", None, 11),
        ("   Exemples : « CARC-BF », « STK-185 », « PORC-ECH-001 »", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("3. Désignation  [OBLIGATOIRE]", True, "2D5A3A", None, 11),
        ("   Libellé complet et lisible du produit.", False, "333333", None, 11),
        ("   Exemples : « Carcasse bœuf », « Steak haché 185g », « Épaule d'agneau »", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("4. Prix achat HT (€)  [OBLIGATOIRE]", True, "2D5A3A", None, 11),
        ("   Le prix hors-taxe. ATTENTION : ce prix est-il AU KILO ou AU COLIS ?", False, "333333", None, 11),
        ("   C'est la colonne suivante (« Prix au ») qui le précise.", False, "333333", None, 11),
        ("   Exemples : 9.70  (au kilo)  |  18.00  (au colis)", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("5. Prix au (kg / colis)  [OBLIGATOIRE] — LA COLONNE CLÉ", True, "C0392B", None, 11),
        ("   Écrivez exactement « kg » ou « colis » (liste déroulante disponible) :", False, "333333", None, 11),
        ("     • kg    → le prix de la colonne 4 est le prix d'UN KILO.", False, "333333", None, 11),
        ("     • colis → le prix de la colonne 4 est le prix d'UN COLIS entier (ou d'une pièce).", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("6. Qté par colis  [si prix au colis]", True, "2D5A3A", None, 11),
        ("   Nombre de pièces contenues dans un colis.", False, "333333", None, 11),
        ("   → Laisser VIDE si le prix est au kilo.", False, "333333", None, 11),
        ("   Exemple : 10  (un carton de 10 steaks)", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("7. Poids unitaire (kg)  [si prix au colis]", True, "2D5A3A", None, 11),
        ("   Poids d'UNE SEULE pièce, en kilos. Utilisez un point comme séparateur décimal.", False, "333333", None, 11),
        ("   → Laisser VIDE si le prix est au kilo.", False, "333333", None, 11),
        ("   Exemples : 0.185  (185 g)  |  2.500  (2,5 kg)  |  0.350  (350 g)", False, "8A6D3B", "FFF8E1", 10),
        ("   Note : la virgule est aussi acceptée (0,185 fonctionne).", False, "888888", None, 10),
        ("", False, "333333", None, 11),
        ("8. TVA (%)  [OBLIGATOIRE]", True, "2D5A3A", None, 11),
        ("   5.5  pour la viande, charcuterie et produits alimentaires de base.", False, "333333", None, 11),
        ("   20   pour tout le reste (emballages, produits d'hygiène, etc.).", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("9. Unités de commande  [facultatif, défaut = tout autorisé]", True, "2D5A3A", None, 11),
        ("   Indique quelles unités le fournisseur accepte en commande.", False, "333333", None, 11),
        ("   Saisir une ou plusieurs valeurs séparées par des virgules parmi : kg, piece, colis", False, "333333", None, 11),
        ("   Laisser vide = toutes les unités sont autorisées (équivalent à « kg,piece,colis »).", False, "333333", None, 11),
        ("   Exemples :", False, "333333", None, 11),
        ("   « colis »          → fournisseur vend uniquement au colis entier (commande à la pièce interdite)", False, "8A6D3B", "FFF8E1", 10),
        ("   « kg,colis »       → fournisseur vend au kilo ou au colis, mais pas à la pièce", False, "8A6D3B", "FFF8E1", 10),
        ("   « kg,piece,colis » → toutes les conversions autorisées (ou laisser vide)", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("10. Famille  [facultatif]", True, "2D5A3A", None, 11),
        ("    Catégorie principale du produit (liste déroulante) :", False, "333333", None, 11),
        ("    Viande | Charcuterie | Traiteur | Aide culinaire | Hygiène et emballage", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("11. Sous-famille  [facultatif]", True, "2D5A3A", None, 11),
        ("    Sous-catégorie selon la famille choisie.", False, "333333", None, 11),
        ("    Exemples pour Viande : Boeuf | Veau | Agneau | Porc | Volaille | Cheval", False, "8A6D3B", "FFF8E1", 10),
        ("    Exemples pour Charcuterie : Saucisse | Jambon | Pâté | Rillettes", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("12. Type de DLC  [facultatif, défaut = dlc]", True, "2D5A3A", None, 11),
        ("    Indique comment la date de consommation est gérée pour ce produit :", False, "333333", None, 11),
        ("    • dlc           → Date Limite de Consommation classique (étiquette apposée).", False, "333333", None, 11),
        ("    • date_abattage → Produit carcasse : la date est celle de l'abattage (tracé par lot).", False, "333333", None, 11),
        ("    • no_dlc        → Produit sans DLC (ex : emballage, sel, épice...).", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("EXEMPLES COMPLETS", True, "2D7D46", "E8F5E9", 12),
        ("", False, "333333", None, 11),
        ("Exemple A — Carcasse bœuf vendue AU KILO :", True, "2D5A3A", None, 11),
        ("   Fournisseur=Boucherie Martin | Code=CARC-BF | Désignation=Carcasse bœuf", False, "8A6D3B", "FFF8E1", 10),
        ("   Prix=9.70 | Prix au=kg | Qté colis=(vide) | Poids unitaire=(vide)", False, "8A6D3B", "FFF8E1", 10),
        ("   TVA=5.5 | Unités=kg,piece,colis | Famille=Viande | Sous-famille=Boeuf | DLC=date_abattage", False, "8A6D3B", "FFF8E1", 10),
        ("", False, "333333", None, 11),
        ("Exemple B — Steaks hachés vendus AU COLIS (carton de 10 × 185 g), commande au colis uniquement :", True, "2D5A3A", None, 11),
        ("   Fournisseur=Boucherie Martin | Code=STK-185 | Désignation=Steak haché 185g", False, "8A6D3B", "FFF8E1", 10),
        ("   Prix=18.00 | Prix au=colis | Qté colis=10 | Poids unitaire=0.185", False, "8A6D3B", "FFF8E1", 10),
        ("   TVA=5.5 | Unités=colis | Famille=Viande | Sous-famille=Boeuf | DLC=dlc", False, "8A6D3B", "FFF8E1", 10),
        ("   → Poids total du colis calculé automatiquement : 10 × 0,185 = 1,85 kg", False, "555555", None, 10),
        ("   → Unités=colis : seule la commande au colis entier est autorisée pour ce fournisseur", False, "555555", None, 10),
        ("", False, "333333", None, 11),
        ("QUESTIONS FRÉQUENTES", True, "2D7D46", None, 12),
        ("", False, "333333", None, 11),
        ("Q : Mon produit est une pièce unique (ex : épaule entière à 12 €). Comment le saisir ?", True, "2D5A3A", None, 11),
        ("R : Prix au = « colis », Qté par colis = 1, Poids unitaire = poids moyen de la pièce.", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("Q : J'ai des produits sans poids fixe (ex : pièces au kilo variable). Comment faire ?", True, "2D5A3A", None, 11),
        ("R : Utilisez « kg » comme format de prix. Les colonnes Qté et Poids unitaire restent vides.", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("Q : La colonne « Poids total colis » est calculée — que dois-je saisir ?", True, "2D5A3A", None, 11),
        ("R : Rien. Elle est calculée automatiquement à l'import (Qté × Poids unitaire). Laissez-la vide.", False, "333333", None, 11),
        ("", False, "333333", None, 11),
        ("Merci ! En cas de doute, contactez-nous — nous vous aiderons à compléter le fichier.", True, "2D7D46", "E8F5E9", 11),
    ]

    for i, row_data in enumerate(lignes_guide, 1):
        txt, bold, color, bg, size = row_data
        c = guide.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=size, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        guide.row_dimensions[i].height = 18 if txt else 8

    wb.active = wb.sheetnames.index("Catalogue")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_catalogue_fournisseur.xlsx"}
    )


@router.post("/catalogue/import")
async def import_catalogue(fichier: bytes = None):
    """Import Excel catalogue fournisseur — voir /catalogue/template pour le format."""
    raise HTTPException(501, "Utiliser POST multipart/form-data avec champ 'fichier'")


@router.post("/catalogue/import/xlsx", status_code=200)
async def import_catalogue_xlsx(fichier: bytes):
    from fastapi import UploadFile, File
    raise HTTPException(501, "Route implémentée via form upload — voir import_catalogue_form")


from fastapi import UploadFile, File

@router.post("/catalogue/import/upload", status_code=200)
async def import_catalogue_upload(fichier: UploadFile = File(...), _=Depends(require_admin)):
    """Import Excel catalogue fournisseur."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl requis")

    content = await fichier.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    # L'onglet de données s'appelle "Catalogue" dans le template ; sinon onglet actif.
    ws = wb["Catalogue"] if "Catalogue" in wb.sheetnames else wb.active

    # Correspondance libellé lisible (template fournisseur) → clé technique.
    # On accepte aussi directement les clés techniques (anciens exports).
    LIBELLE_VERS_CLE = {
        "fournisseur": "fournisseur_nom",
        "code article": "code_article",
        "désignation": "designation", "designation": "designation",
        "prix achat ht (€)": "prix_achat_ht", "prix achat ht": "prix_achat_ht",
        "prix au (kg / colis)": "format_prix", "prix au": "format_prix",
        "qté par colis": "qte_par_colis", "qte par colis": "qte_par_colis",
        "poids unitaire (kg)": "poids_unitaire_kg", "poids unitaire": "poids_unitaire_kg",
        "poids total colis (kg)": "poids_colis_kg",
        "tva (%)": "tva_percent", "tva": "tva_percent",
        "conditionnement (texte libre)": "conditionnement", "conditionnement": "conditionnement",
        "unités de commande": "unites_autorisees", "unites de commande": "unites_autorisees",
        "unites_autorisees": "unites_autorisees",
        "famille": "famille",
        "sous-famille": "sous_famille", "sous famille": "sous_famille", "sous_famille": "sous_famille",
        "type de dlc": "dlc_type", "type dlc": "dlc_type",
    }

    def _cle(libelle):
        l = libelle.strip().lower()
        return LIBELLE_VERS_CLE.get(l, l)  # libellé connu → clé, sinon tel quel

    headers = [_cle(str(ws.cell(row=1, column=c).value or "")) for c in range(1, ws.max_column + 1)]
    required = {"fournisseur_nom", "code_article", "designation", "prix_achat_ht"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(400, f"Colonnes manquantes : {missing}")

    def col(row, name):
        idx = headers.index(name) + 1
        v = ws.cell(row=row, column=idx).value
        return str(v).strip() if v is not None else ""

    stats = {"crees": 0, "mis_a_jour": 0, "erreurs": []}

    # Lignes 1=en-têtes, 2=notes, 3-4=exemples → vraies données à partir de la ligne 5.
    # On saute aussi toute ligne d'exemple résiduelle laissée par le fournisseur.
    EXEMPLES_CODES = {"carc-bf", "stk-185"}

    async with get_db() as db:
        for row_num in range(5, ws.max_row + 1):
            nom_fourn = col(row_num, "fournisseur_nom")
            code = col(row_num, "code_article")
            if not nom_fourn or not code:
                continue
            # Ignorer une ligne d'exemple que le fournisseur aurait laissée en place
            if code.strip().lower() in EXEMPLES_CODES and nom_fourn.strip().lower() == "boucherie martin":
                continue

            # Trouver le fournisseur
            cur = await db.execute(
                "SELECT id FROM fournisseurs WHERE boutique_id = 1 AND LOWER(TRIM(nom)) = LOWER(TRIM(?))",
                (nom_fourn,)
            )
            fourn = await cur.fetchone()
            if not fourn:
                stats["erreurs"].append(f"Ligne {row_num} : fournisseur '{nom_fourn}' introuvable")
                continue

            designation = col(row_num, "designation")
            try:
                prix = float(col(row_num, "prix_achat_ht") or 0)
            except ValueError:
                stats["erreurs"].append(f"Ligne {row_num} : prix invalide")
                continue

            tva_raw = col(row_num, "tva_percent") if "tva_percent" in headers else "5.5"
            try:
                tva = float(tva_raw or 5.5)
            except ValueError:
                tva = 5.5

            format_prix       = _normaliser_format_prix(col(row_num, "format_prix")) if "format_prix" in headers else "kg"
            conditionnement   = col(row_num, "conditionnement")   if "conditionnement"   in headers else None
            unites_autorisees = col(row_num, "unites_autorisees") if "unites_autorisees" in headers else None
            famille           = col(row_num, "famille")           if "famille"           in headers else None
            sous_famille    = col(row_num, "sous_famille")    if "sous_famille"    in headers else None
            dlc_type        = col(row_num, "dlc_type")        if "dlc_type"        in headers else "dlc"

            def _num(name):
                """Lit une cellule numérique optionnelle (qte/poids), '' → None."""
                if name not in headers:
                    return None
                raw = col(row_num, name).replace(",", ".")  # tolère la virgule décimale
                if not raw:
                    return None
                try:
                    return float(raw)
                except ValueError:
                    return None

            qte_par_colis     = _num("qte_par_colis")
            poids_unitaire_kg = _num("poids_unitaire_kg")
            poids_colis_kg    = _calc_poids_colis_kg(qte_par_colis, poids_unitaire_kg)

            # UPSERT
            cur2 = await db.execute(
                "SELECT id FROM catalogue_fournisseur WHERE fournisseur_id = ? AND code_article = ?",
                (fourn["id"], code)
            )
            existing = await cur2.fetchone()
            if existing:
                await db.execute(
                    """UPDATE catalogue_fournisseur
                       SET designation=?, prix_achat_ht=?, format_prix=?,
                           qte_par_colis=?, poids_unitaire_kg=?, poids_colis_kg=?,
                           tva_percent=?, conditionnement=?, unites_autorisees=?,
                           famille=?, sous_famille=?, dlc_type=?, date_maj=CURRENT_TIMESTAMP
                       WHERE id=?""",
                    (designation, prix, format_prix,
                     qte_par_colis, poids_unitaire_kg, poids_colis_kg,
                     tva, conditionnement or None, unites_autorisees or None,
                     famille or None, sous_famille or None,
                     dlc_type or "dlc", existing["id"])
                )
                stats["mis_a_jour"] += 1
            else:
                await db.execute(
                    """INSERT INTO catalogue_fournisseur
                       (fournisseur_id, code_article, designation, prix_achat_ht, format_prix,
                        qte_par_colis, poids_unitaire_kg, poids_colis_kg, tva_percent, conditionnement,
                        unites_autorisees, famille, sous_famille, dlc_type)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (fourn["id"], code, designation, prix, format_prix,
                     qte_par_colis, poids_unitaire_kg, poids_colis_kg,
                     tva, conditionnement or None, unites_autorisees or None,
                     famille or None, sous_famille or None,
                     dlc_type or "dlc")
                )
                stats["crees"] += 1

        await db.commit()

    return stats


@router.get("/catalogue/{article_id}")
async def get_article(article_id: int):
    async with get_db() as db:
        cur = await db.execute(
            "SELECT c.*, f.nom AS fournisseur_nom FROM catalogue_fournisseur c JOIN fournisseurs f ON f.id = c.fournisseur_id WHERE c.id = ?",
            (article_id,)
        )
        row = await cur.fetchone()
        if not row:
            raise HTTPException(404, "Article introuvable")
        return dict(row)


@router.post("/catalogue", status_code=201)
async def create_article(body: CatalogueArticleCreate, _=Depends(require_admin)):
    async with get_db() as db:
        cur_f = await db.execute("SELECT id FROM fournisseurs WHERE id = ? AND boutique_id = 1", (body.fournisseur_id,))
        if not await cur_f.fetchone():
            raise HTTPException(404, "Fournisseur introuvable")

        cur_dup = await db.execute(
            "SELECT id FROM catalogue_fournisseur WHERE fournisseur_id = ? AND code_article = ?",
            (body.fournisseur_id, body.code_article)
        )
        if await cur_dup.fetchone():
            raise HTTPException(409, f"Code article '{body.code_article}' déjà existant pour ce fournisseur")

        format_prix = _normaliser_format_prix(body.format_prix)
        poids_colis = _calc_poids_colis_kg(body.qte_par_colis, body.poids_unitaire_kg)
        cur = await db.execute(
            """INSERT INTO catalogue_fournisseur
               (fournisseur_id, code_article, designation, prix_achat_ht, format_prix,
                qte_par_colis, poids_unitaire_kg, poids_colis_kg, tva_percent, conditionnement,
                unites_autorisees, famille, sous_famille, dlc_type)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (body.fournisseur_id, body.code_article, body.designation, body.prix_achat_ht,
             format_prix, body.qte_par_colis, body.poids_unitaire_kg,
             poids_colis, body.tva_percent, body.conditionnement,
             body.unites_autorisees or 'kg,piece,colis',
             body.famille, body.sous_famille, body.dlc_type)
        )
        await db.commit()
        cur2 = await db.execute("SELECT * FROM catalogue_fournisseur WHERE id = ?", (cur.lastrowid,))
        return dict(await cur2.fetchone())


@router.put("/catalogue/{article_id}")
async def update_article(article_id: int, body: CatalogueArticleUpdate, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute("SELECT * FROM catalogue_fournisseur WHERE id = ?", (article_id,))
        existing = await cur.fetchone()
        if not existing:
            raise HTTPException(404, "Article introuvable")

        fields = {k: v for k, v in body.model_dump().items() if v is not None}
        if not fields:
            raise HTTPException(400, "Aucun champ à modifier")

        if "format_prix" in fields:
            fields["format_prix"] = _normaliser_format_prix(fields["format_prix"])

        # Recalcule le poids du colis si une des deux données brutes est modifiée,
        # en repartant des valeurs existantes pour celle qui ne change pas.
        if "qte_par_colis" in fields or "poids_unitaire_kg" in fields:
            qte = fields.get("qte_par_colis", existing["qte_par_colis"])
            pu = fields.get("poids_unitaire_kg", existing["poids_unitaire_kg"])
            fields["poids_colis_kg"] = _calc_poids_colis_kg(qte, pu)

        fields["date_maj"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [article_id]
        await db.execute(f"UPDATE catalogue_fournisseur SET {set_clause} WHERE id = ?", values)
        await db.commit()

        cur2 = await db.execute("SELECT * FROM catalogue_fournisseur WHERE id = ?", (article_id,))
        return dict(await cur2.fetchone())


@router.delete("/catalogue/{article_id}", status_code=200)
async def delete_article(article_id: int, permanent: bool = Query(False), _=Depends(require_admin)):
    async with get_db() as db:
        if permanent:
            # Couper les liens FK avant suppression (les lignes historiques restent, juste délié)
            await db.execute(
                "UPDATE commande_lignes SET catalogue_fournisseur_id = NULL WHERE catalogue_fournisseur_id = ?",
                (article_id,)
            )
            await db.execute(
                "UPDATE reception_lignes SET catalogue_fournisseur_id = NULL WHERE catalogue_fournisseur_id = ?",
                (article_id,)
            )
            await db.execute("DELETE FROM catalogue_fournisseur WHERE id = ?", (article_id,))
        else:
            await db.execute("UPDATE catalogue_fournisseur SET actif = 0 WHERE id = ?", (article_id,))
        await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Commandes
# ---------------------------------------------------------------------------

async def _generer_numero_commande(db) -> str:
    today = date.today().strftime("%Y%m%d")
    cur = await db.execute(
        "SELECT COUNT(*) FROM commandes WHERE date_commande = ?",
        (date.today().isoformat(),)
    )
    count = (await cur.fetchone())[0] + 1
    return f"CMD-{today}-{count:03d}"


async def _recalculer_total(db, commande_id: int):
    cur = await db.execute(
        "SELECT COALESCE(SUM(montant_ht), 0) FROM commande_lignes WHERE commande_id = ?",
        (commande_id,)
    )
    total = (await cur.fetchone())[0]
    await db.execute("UPDATE commandes SET montant_total_ht = ? WHERE id = ?", (total, commande_id))


@router.get("/commandes")
async def get_commandes(
    fournisseur_id: Optional[int] = Query(None),
    statut: Optional[str] = Query(None),
    limit: int = Query(50),
    non_liee: bool = Query(False),
):
    async with get_db() as db:
        sql = """
            SELECT c.*, f.nom AS fournisseur_nom, f.email_commercial,
                   p.prenom AS personnel_prenom
            FROM commandes c
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            LEFT JOIN personnel p ON p.id = c.personnel_id
            WHERE c.boutique_id = 1
        """
        params = []
        if fournisseur_id:
            sql += " AND c.fournisseur_id = ?"
            params.append(fournisseur_id)
        if statut:
            sql += " AND c.statut = ?"
            params.append(statut)
        if non_liee:
            # Une commande n'est masquée que si elle est liée à une réception
            # CLÔTURÉE. Tant que la réception liée est 'en_cours' (fiche créée puis
            # abandonnée sans clôture), la commande reste sélectionnable — sinon
            # elle se retrouve piégée à une réception fantôme (ni au stock, ni en
            # attente, ni livrée).
            sql += """ AND NOT EXISTS (
                SELECT 1 FROM commande_receptions_mapping m
                JOIN receptions r ON r.id = m.reception_id
                WHERE m.commande_id = c.id AND r.statut = 'cloturee'
            )"""
        sql += " ORDER BY c.date_commande DESC LIMIT ?"
        params.append(limit)
        cur = await db.execute(sql, params)
        commandes = [dict(r) for r in await cur.fetchall()]

        # Ajouter le nombre de lignes pour chaque commande
        for cmd in commandes:
            cur2 = await db.execute(
                "SELECT COUNT(*) FROM commande_lignes WHERE commande_id = ?", (cmd["id"],)
            )
            cmd["nb_lignes"] = (await cur2.fetchone())[0]

        return commandes


@router.get("/commandes/{commande_id}")
async def get_commande(commande_id: int):
    async with get_db() as db:
        cur = await db.execute(
            """SELECT c.*, f.nom AS fournisseur_nom, f.email_commercial, f.telephone, f.adresse,
                      p.prenom AS personnel_prenom
               FROM commandes c
               JOIN fournisseurs f ON f.id = c.fournisseur_id
               LEFT JOIN personnel p ON p.id = c.personnel_id
               WHERE c.id = ?""",
            (commande_id,)
        )
        commande = await cur.fetchone()
        if not commande:
            raise HTTPException(404, "Commande introuvable")
        result = dict(commande)

        cur2 = await db.execute(
            """SELECT cl.*, cf.dlc_type AS dlc_type, cf.tva_percent AS tva_percent
               FROM commande_lignes cl
               LEFT JOIN catalogue_fournisseur cf ON cf.id = cl.catalogue_fournisseur_id
               WHERE cl.commande_id = ? ORDER BY cl.id""",
            (commande_id,)
        )
        result["lignes"] = [dict(r) for r in await cur2.fetchall()]

        # Réception rapprochée (mapping) + facture éventuelle — pour le bouton "Saisir la facture".
        cur3 = await db.execute(
            """SELECT reception_id FROM commande_receptions_mapping
               WHERE commande_id = ? ORDER BY date_liaison DESC LIMIT 1""",
            (commande_id,),
        )
        map_row = await cur3.fetchone()
        result["reception_id"] = map_row["reception_id"] if map_row else None

        cur4 = await db.execute(
            "SELECT id FROM factures WHERE commande_id = ? OR reception_id = ? LIMIT 1",
            (commande_id, result["reception_id"]),
        )
        fac_row = await cur4.fetchone()
        result["facture_id"] = fac_row["id"] if fac_row else None

        return result


@router.post("/commandes", status_code=201)
async def create_commande(body: CommandeCreate):
    async with get_db() as db:
        cur_f = await db.execute(
            "SELECT * FROM fournisseurs WHERE id = ? AND boutique_id = 1",
            (body.fournisseur_id,)
        )
        if not await cur_f.fetchone():
            raise HTTPException(404, "Fournisseur introuvable")

        numero = await _generer_numero_commande(db)
        date_cmd = body.date_commande or date.today().isoformat()

        cur = await db.execute(
            """INSERT INTO commandes (boutique_id, fournisseur_id, numero_commande, date_commande,
                                      date_livraison_prevue, commentaire, personnel_id)
               VALUES (1, ?, ?, ?, ?, ?, ?)""",
            (body.fournisseur_id, numero, date_cmd, body.date_livraison_prevue,
             body.commentaire, body.personnel_id)
        )
        await db.commit()
        commande_id = cur.lastrowid

        # Insérer les lignes si fournies
        for ligne in (body.lignes or []):
            montant = ligne.quantite_commandee * ligne.prix_unitaire_ht
            await db.execute(
                """INSERT INTO commande_lignes
                   (commande_id, catalogue_fournisseur_id, code_article, designation,
                    prix_unitaire_ht, quantite_commandee, unite, montant_ht, commentaire_ligne)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (commande_id, ligne.catalogue_fournisseur_id, ligne.code_article or '',
                 ligne.designation, ligne.prix_unitaire_ht, ligne.quantite_commandee,
                 ligne.unite, montant, ligne.commentaire_ligne)
            )

        await _recalculer_total(db, commande_id)
        await db.commit()

        return await get_commande(commande_id)


@router.put("/commandes/{commande_id}")
async def update_commande(commande_id: int, body: CommandeUpdate):
    async with get_db() as db:
        cur = await db.execute("SELECT * FROM commandes WHERE id = ?", (commande_id,))
        if not await cur.fetchone():
            raise HTTPException(404, "Commande introuvable")

        fields = {k: v for k, v in body.model_dump().items() if v is not None}
        if not fields:
            raise HTTPException(400, "Aucun champ à modifier")

        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [commande_id]
        await db.execute(f"UPDATE commandes SET {set_clause} WHERE id = ?", values)
        await db.commit()

        return await get_commande(commande_id)


@router.post("/commandes/{commande_id}/lignes", status_code=201)
async def add_ligne(commande_id: int, body: CommandeLigneCreate):
    async with get_db() as db:
        cur = await db.execute("SELECT statut FROM commandes WHERE id = ?", (commande_id,))
        cmd = await cur.fetchone()
        if not cmd:
            raise HTTPException(404, "Commande introuvable")
        if cmd["statut"] not in ("brouillon",):
            raise HTTPException(400, "Impossible de modifier une commande déjà confirmée")

        montant = body.quantite_commandee * body.prix_unitaire_ht
        cur2 = await db.execute(
            """INSERT INTO commande_lignes
               (commande_id, catalogue_fournisseur_id, code_article, designation,
                prix_unitaire_ht, quantite_commandee, unite, montant_ht, commentaire_ligne)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (commande_id, body.catalogue_fournisseur_id, body.code_article, body.designation,
             body.prix_unitaire_ht, body.quantite_commandee, body.unite, montant, body.commentaire_ligne)
        )
        await _recalculer_total(db, commande_id)
        await db.commit()

        cur3 = await db.execute("SELECT * FROM commande_lignes WHERE id = ?", (cur2.lastrowid,))
        return dict(await cur3.fetchone())


@router.put("/commandes/{commande_id}/lignes/{ligne_id}")
async def update_ligne(commande_id: int, ligne_id: int, body: CommandeLigneUpdate):
    async with get_db() as db:
        cur = await db.execute(
            "SELECT * FROM commande_lignes WHERE id = ? AND commande_id = ?",
            (ligne_id, commande_id)
        )
        ligne = await cur.fetchone()
        if not ligne:
            raise HTTPException(404, "Ligne introuvable")

        fields = {k: v for k, v in body.model_dump().items() if v is not None}
        if not fields:
            raise HTTPException(400, "Aucun champ à modifier")

        # Recalculer montant si quantité ou prix change
        qte = fields.get("quantite_commandee", ligne["quantite_commandee"])
        prix = fields.get("prix_unitaire_ht", ligne["prix_unitaire_ht"])
        fields["montant_ht"] = qte * prix

        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [ligne_id]
        await db.execute(f"UPDATE commande_lignes SET {set_clause} WHERE id = ?", values)
        await _recalculer_total(db, commande_id)
        await db.commit()

        cur2 = await db.execute("SELECT * FROM commande_lignes WHERE id = ?", (ligne_id,))
        return dict(await cur2.fetchone())


@router.delete("/commandes/{commande_id}/lignes/{ligne_id}", status_code=204)
async def delete_ligne(commande_id: int, ligne_id: int):
    async with get_db() as db:
        await db.execute(
            "DELETE FROM commande_lignes WHERE id = ? AND commande_id = ?",
            (ligne_id, commande_id)
        )
        await _recalculer_total(db, commande_id)
        await db.commit()


@router.delete("/commandes/{commande_id}", status_code=204)
async def delete_commande(commande_id: int):
    """Supprime une commande et toutes ses lignes (sauf si livrée)."""
    async with get_db() as db:
        cur = await db.execute("SELECT id, statut FROM commandes WHERE id = ?", (commande_id,))
        row = await cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Commande introuvable")
        if row["statut"] == "livree":
            raise HTTPException(status_code=409, detail="Impossible de supprimer une commande déjà livrée")
        await db.execute("DELETE FROM commande_lignes WHERE commande_id = ?", (commande_id,))
        await db.execute("DELETE FROM commande_receptions_mapping WHERE commande_id = ?", (commande_id,))
        await db.execute("DELETE FROM commandes WHERE id = ?", (commande_id,))
        await db.commit()


@router.post("/commandes/{commande_id}/dupliquer", status_code=201)
async def dupliquer_commande(commande_id: int, personnel_id: Optional[int] = None):
    """Duplique une commande existante en brouillon avec les mêmes lignes."""
    async with get_db() as db:
        cur = await db.execute("SELECT * FROM commandes WHERE id = ?", (commande_id,))
        source = await cur.fetchone()
        if not source:
            raise HTTPException(404, "Commande introuvable")

        numero = await _generer_numero_commande(db)
        cur2 = await db.execute(
            """INSERT INTO commandes (boutique_id, fournisseur_id, numero_commande, date_commande,
                                      date_livraison_prevue, commentaire, personnel_id, statut)
               VALUES (1, ?, ?, ?, ?, ?, ?, 'brouillon')""",
            (source["fournisseur_id"], numero, date.today().isoformat(),
             source["date_livraison_prevue"], source["commentaire"], personnel_id or source["personnel_id"])
        )
        await db.commit()
        new_id = cur2.lastrowid

        # Copier les lignes
        cur3 = await db.execute(
            "SELECT * FROM commande_lignes WHERE commande_id = ?", (commande_id,)
        )
        lignes = await cur3.fetchall()
        for l in lignes:
            await db.execute(
                """INSERT INTO commande_lignes
                   (commande_id, catalogue_fournisseur_id, code_article, designation,
                    prix_unitaire_ht, quantite_commandee, unite, montant_ht, commentaire_ligne)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (new_id, l["catalogue_fournisseur_id"], l["code_article"], l["designation"],
                 l["prix_unitaire_ht"], l["quantite_commandee"], l["unite"], l["montant_ht"],
                 l["commentaire_ligne"])
            )

        await _recalculer_total(db, new_id)
        await db.commit()

        return await get_commande(new_id)


# ---------------------------------------------------------------------------
# Panier multi-fournisseurs
# ---------------------------------------------------------------------------

class PanierLigneCreate(BaseModel):
    catalogue_fournisseur_id: Optional[int] = None
    fournisseur_id: int
    fournisseur_nom: str
    code_article: str
    designation: str
    quantite: float
    unite: Optional[str] = "kg"
    prix_ht: float


class PanierSave(BaseModel):
    lignes: List[PanierLigneCreate]


class PanierGenerer(BaseModel):
    date_livraison_prevue: Optional[str] = None
    commentaire: Optional[str] = None


@router.get("/panier")
async def get_panier():
    async with get_db() as db:
        cur = await db.execute(
            "SELECT * FROM panier_lignes WHERE boutique_id = 1 ORDER BY fournisseur_nom, designation"
        )
        return [dict(r) for r in await cur.fetchall()]


@router.put("/panier", status_code=200)
async def save_panier(body: PanierSave):
    async with get_db() as db:
        await db.execute("DELETE FROM panier_lignes WHERE boutique_id = 1")
        for l in body.lignes:
            await db.execute(
                """INSERT INTO panier_lignes
                   (boutique_id, catalogue_fournisseur_id, fournisseur_id, fournisseur_nom,
                    code_article, designation, quantite, unite, prix_ht)
                   VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (l.catalogue_fournisseur_id, l.fournisseur_id, l.fournisseur_nom,
                 l.code_article, l.designation, l.quantite, l.unite, l.prix_ht)
            )
        await db.commit()
    return {"ok": True, "nb_lignes": len(body.lignes)}


@router.delete("/panier", status_code=204)
async def delete_panier():
    async with get_db() as db:
        await db.execute("DELETE FROM panier_lignes WHERE boutique_id = 1")
        await db.commit()


# ---------------------------------------------------------------------------
# Commande semi-automatique : produits de référence + suggestions
# ---------------------------------------------------------------------------
# Squelette du système de commande semi-automatique :
#   1. /panier/references  → les lignes d'achat « de référence » (⭐) arbitrées
#      dans le comparatif fournisseurs (comparatif_groupe_vente.ligne_choisie_id),
#      pour les mettre en avant dans le panier.
#   2. /panier/suggestions → moteur de suggestions basé sur la RÉCURRENCE des
#      commandes passées, avec un score de prédominance par article.

# Pondérations du score de prédominance (somme = 1). Centralisées ici pour
# pouvoir ajuster le moteur sans toucher au calcul. L'activité étant
# irrégulière (commandes du jour au lendemain), le moteur ne prédit PAS quand
# commander : il estime le BESOIN au moment où l'utilisateur ouvre le panier,
# via la consommation induite (quantités commandées ÷ durée d'activité).
SUGGESTION_POIDS_FREQUENCE = 0.60   # part des commandes contenant l'article
SUGGESTION_POIDS_BESOIN    = 0.40   # besoin estimé vs commande type
# Un article est « à commander » quand le besoin estimé atteint cette fraction
# de sa quantité moyenne de commande.
SUGGESTION_SEUIL_BESOIN = 0.5
# Durée minimale d'historique (jours) pour estimer une consommation fiable.
SUGGESTION_MIN_JOURS_CONSO = 7


@router.get("/panier/references")
async def get_panier_references():
    """Lignes d'achat de référence (⭐) choisies dans le comparatif fournisseurs.

    Une ligne catalogue peut servir de référence à plusieurs produits de vente
    (et plusieurs groupes) : on agrège par ligne d'achat.
    """
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT gv.ligne_choisie_id                  AS catalogue_fournisseur_id,
                   GROUP_CONCAT(DISTINCT g.nom)         AS groupes,
                   COUNT(gv.catalogue_vente_id)         AS nb_produits_vente
            FROM comparatif_groupe_vente gv
            JOIN comparatif_groupe g ON g.id = gv.groupe_id
            WHERE gv.ligne_choisie_id IS NOT NULL AND g.boutique_id = 1
            GROUP BY gv.ligne_choisie_id
            """
        )
        return [dict(r) for r in await cur.fetchall()]


async def _calculer_suggestions(db, date_debut: date, fenetre_jours: int):
    """Métriques de récurrence + score de prédominance par article catalogue.

    Partagé entre /panier/suggestions et /panier/cadencier. Ne considère que
    les commandes réellement passées (confirmées ou livrées — les brouillons
    et annulées sont du bruit) depuis date_debut, les articles actifs, et
    exclut la prestation désossage (gérée automatiquement).

    Retourne (suggestions triées par score décroissant, nb_commandes_total,
    occurrences par article : {catalogue_id: [{date_commande, quantite_commandee,
    unite, commande_id}, …]}).
    """
    cur = await db.execute(
        """
        SELECT cl.catalogue_fournisseur_id,
               c.id AS commande_id,
               c.date_commande,
               cl.quantite_commandee,
               cl.unite
        FROM commande_lignes cl
        JOIN commandes c ON c.id = cl.commande_id
        JOIN catalogue_fournisseur cf ON cf.id = cl.catalogue_fournisseur_id
        WHERE c.boutique_id = 1
          AND c.statut IN ('confirmee', 'livree')
          AND c.date_commande >= ?
          AND cl.catalogue_fournisseur_id IS NOT NULL
          AND cf.actif = 1
          AND cf.code_article != ?
        ORDER BY cl.catalogue_fournisseur_id, c.date_commande
        """,
        (date_debut.isoformat(), CODE_PREST_DESOSSAGE),
    )
    lignes = [dict(r) for r in await cur.fetchall()]

    # Nombre total de commandes passées de la fenêtre, pour normaliser la
    # fréquence : un article présent dans toutes les commandes → 1.0.
    cur_n = await db.execute(
        """SELECT COUNT(*) AS n FROM commandes
           WHERE boutique_id = 1 AND statut IN ('confirmee', 'livree')
             AND date_commande >= ?""",
        (date_debut.isoformat(),),
    )
    nb_commandes_total = (await cur_n.fetchone())["n"]

    # Lignes de référence du comparatif (⭐) pour marquage.
    cur_ref = await db.execute(
        """SELECT DISTINCT gv.ligne_choisie_id AS id
           FROM comparatif_groupe_vente gv
           JOIN comparatif_groupe g ON g.id = gv.groupe_id
           WHERE gv.ligne_choisie_id IS NOT NULL AND g.boutique_id = 1"""
    )
    refs = {r["id"] for r in await cur_ref.fetchall()}

    # Infos catalogue pour l'affichage et les filtres (famille / sous-famille).
    cat_ids = {l["catalogue_fournisseur_id"] for l in lignes}
    cat = {}
    if cat_ids:
        placeholders = ",".join("?" * len(cat_ids))
        cur_c = await db.execute(
            f"""SELECT c.id, c.code_article, c.designation, c.fournisseur_id,
                       c.famille, c.sous_famille, f.nom AS fournisseur_nom
                FROM catalogue_fournisseur c
                JOIN fournisseurs f ON f.id = c.fournisseur_id
                WHERE c.id IN ({placeholders})""",
            list(cat_ids),
        )
        cat = {r["id"]: dict(r) for r in await cur_c.fetchall()}

    # Agrégation par article (en Python : intervalles + unité dominante).
    from collections import defaultdict
    par_article = defaultdict(list)
    for l in lignes:
        par_article[l["catalogue_fournisseur_id"]].append(l)

    aujourd_hui = date.today()
    suggestions = []
    for cat_id, occ in par_article.items():
        a = cat.get(cat_id)
        if not a:
            continue

        # Dates de commande distinctes (un article 2× dans la même commande ou
        # le même jour compte une fois pour la récurrence). Déjà triées par SQL.
        dates = sorted({date.fromisoformat(o["date_commande"][:10]) for o in occ})
        nb_cmd = len({o["commande_id"] for o in occ})
        derniere = dates[-1]
        jours_depuis = (aujourd_hui - derniere).days

        # Récurrence : intervalle moyen entre deux commandes successives.
        intervalle_moyen = None
        if len(dates) >= 2:
            ecarts = [(d2 - d1).days for d1, d2 in zip(dates, dates[1:])]
            intervalle_moyen = round(sum(ecarts) / len(ecarts), 1)

        # Quantités dans l'unité la plus fréquente (pas de conversion hasardeuse).
        unites = defaultdict(list)
        for o in occ:
            unites[o["unite"] or "kg"].append(o["quantite_commandee"] or 0)
        unite_dominante = max(unites, key=lambda u: len(unites[u]))
        qtes = unites[unite_dominante]
        quantite_moyenne = round(sum(qtes) / len(qtes), 3)
        qte_max = max(qtes)
        derniere_qte = next(
            (o["quantite_commandee"] or 0 for o in reversed(occ) if (o["unite"] or "kg") == unite_dominante),
            qtes[-1]
        )

        # ── Consommation induite ────────────────────────────────────
        # Le commerce est irrégulier : on ne prédit pas QUAND commander, on
        # estime le besoin MAINTENANT. Tout ce qui a été commandé entre la
        # première commande et aujourd'hui est supposé consommé →
        # conso/jour = total commandé ÷ jours d'activité. Le besoin estimé =
        # conso/jour × jours depuis la dernière commande, plafonné à la plus
        # grosse commande passée (jamais suggérer plus qu'un maximum connu :
        # un long trou — congés, saison — ne doit pas gonfler la suggestion).
        premiere = dates[0]
        jours_activite = (aujourd_hui - premiere).days
        conso_jour = None
        if len(dates) >= 2 and jours_activite >= SUGGESTION_MIN_JOURS_CONSO:
            conso_jour = sum(qtes) / jours_activite
        besoin_estime = None
        if conso_jour:
            besoin_estime = round(min(conso_jour * jours_depuis, qte_max), 3)

        # Quantité suggérée : le besoin estimé si calculable, sinon la
        # commande type (moyenne) — répond à « combien je commande, là ».
        quantite_suggeree = besoin_estime if besoin_estime else quantite_moyenne

        # ── Score de prédominance (0-100) ───────────────────────────
        # fréquence : part des commandes de la fenêtre contenant l'article —
        # « on commande toujours la même chose », c'est elle qui domine.
        frequence = min(1.0, nb_cmd / nb_commandes_total) if nb_commandes_total else 0.0
        # besoin : besoin estimé rapporté à la commande type. Neutre (0.5)
        # quand la consommation n'est pas encore calculable.
        if besoin_estime is not None and quantite_moyenne > 0:
            besoin = min(1.0, besoin_estime / quantite_moyenne)
        else:
            besoin = 0.5
        score = round(100 * (
            SUGGESTION_POIDS_FREQUENCE * frequence
            + SUGGESTION_POIDS_BESOIN * besoin
        ))

        a_commander = bool(
            besoin_estime is not None and quantite_moyenne > 0
            and besoin_estime >= SUGGESTION_SEUIL_BESOIN * quantite_moyenne
        )

        suggestions.append({
            "catalogue_fournisseur_id": cat_id,
            "code_article": a["code_article"],
            "designation": a["designation"],
            "fournisseur_id": a["fournisseur_id"],
            "fournisseur_nom": a["fournisseur_nom"],
            "famille": a["famille"],
            "sous_famille": a["sous_famille"],
            "est_reference": cat_id in refs,
            "nb_commandes": nb_cmd,
            "derniere_commande": derniere.isoformat(),
            "derniere_qte": derniere_qte,
            "jours_depuis": jours_depuis,
            "intervalle_moyen_jours": intervalle_moyen,
            "quantite_moyenne": quantite_moyenne,
            "conso_hebdo": round(conso_jour * 7, 3) if conso_jour else None,
            "besoin_estime": besoin_estime,
            "quantite_suggeree": quantite_suggeree,
            "unite_suggeree": unite_dominante,
            "score": score,
            "composantes": {
                "frequence": round(frequence, 3),
                "besoin": round(besoin, 3),
            },
            "a_commander": a_commander,
        })

    suggestions.sort(key=lambda s: s["score"], reverse=True)
    return suggestions, nb_commandes_total, par_article


@router.get("/panier/suggestions")
async def get_panier_suggestions(fenetre_jours: int = Query(180, ge=7, le=730)):
    """Suggestions « quoi commander et combien », basées sur la consommation
    induite par les commandes passées (confirmées/livrées).

    Pour chaque article commandé dans la fenêtre :
      - conso_hebdo : quantité commandée ÷ durée d'activité (consommation) ;
      - besoin_estime : conso/jour × jours depuis la dernière commande,
        plafonné à la plus grosse commande passée ;
      - quantite_suggeree : le besoin estimé, sinon la commande type (moyenne) ;
      - score (0-100) de prédominance = fréquence (0.6) + besoin (0.4) ;
      - a_commander : besoin estimé ≥ moitié d'une commande type.
    """
    async with get_db() as db:
        suggestions, nb_total, _ = await _calculer_suggestions(
            db, date.today() - timedelta(days=fenetre_jours), fenetre_jours
        )
    return {
        "fenetre_jours": fenetre_jours,
        "nb_commandes_total": nb_total,
        "suggestions": suggestions,
    }


MOIS_COURTS = ["janv", "févr", "mars", "avr", "mai", "juin",
               "juil", "août", "sept", "oct", "nov", "déc"]
JOURS_COURTS = ["lun", "mar", "mer", "jeu", "ven", "sam", "dim"]


@router.get("/panier/cadencier")
async def get_panier_cadencier(
    granularite: str = Query("semaine", pattern="^(jour|semaine|mois)$"),
    periodes: int = Query(12, ge=2, le=31),
):
    """Cadencier « panier intelligent » : quantités commandées par article et
    par période (jour, semaine ISO ou mois civil), enrichies des métriques de
    suggestion (récurrence, score, quantité suggérée).

    Les quantités d'une ligne sont sommées dans l'unité dominante de l'article ;
    si d'autres unités apparaissent dans l'historique, unites_mixtes=True (les
    occurrences dans une autre unité ne sont pas sommées — pas de conversion
    hasardeuse). Tri/filtres (fournisseur, famille, sous-famille) côté front.
    """
    aujourd_hui = date.today()
    if granularite == "jour":
        debuts = [aujourd_hui - timedelta(days=i) for i in range(periodes - 1, -1, -1)]
        buckets = [
            {"label": f"{JOURS_COURTS[d.weekday()]} {d.day}", "debut": d.isoformat(),
             "fin": d.isoformat()}
            for d in debuts
        ]

        def index_periode(d: date) -> int:
            return (d - debuts[0]).days
    elif granularite == "semaine":
        lundi_courant = aujourd_hui - timedelta(days=aujourd_hui.weekday())
        debuts = [lundi_courant - timedelta(weeks=i) for i in range(periodes - 1, -1, -1)]
        buckets = [
            {"label": f"S{d.isocalendar()[1]}", "debut": d.isoformat(),
             "fin": (d + timedelta(days=6)).isoformat()}
            for d in debuts
        ]

        def index_periode(d: date) -> int:
            return ((d - timedelta(days=d.weekday())) - debuts[0]).days // 7
    else:
        debuts = []
        cur_mois = aujourd_hui.replace(day=1)
        for _ in range(periodes):
            debuts.append(cur_mois)
            cur_mois = (cur_mois - timedelta(days=1)).replace(day=1)
        debuts.reverse()

        def fin_mois(d: date) -> date:
            return (d + timedelta(days=32)).replace(day=1) - timedelta(days=1)

        buckets = [
            {"label": f"{MOIS_COURTS[d.month - 1]} {d.year % 100:02d}",
             "debut": d.isoformat(), "fin": fin_mois(d).isoformat()}
            for d in debuts
        ]

        def index_periode(d: date) -> int:
            return (d.year * 12 + d.month) - (debuts[0].year * 12 + debuts[0].month)

    fenetre_jours = (aujourd_hui - debuts[0]).days + 1
    async with get_db() as db:
        suggestions, nb_total, par_article = await _calculer_suggestions(
            db, debuts[0], fenetre_jours
        )

    lignes = []
    for s in suggestions:
        qtes = [0.0] * periodes
        unites_mixtes = False
        for o in par_article.get(s["catalogue_fournisseur_id"], []):
            d = date.fromisoformat(o["date_commande"][:10])
            i = index_periode(d)
            if not (0 <= i < periodes):
                continue
            if (o["unite"] or "kg") == s["unite_suggeree"]:
                qtes[i] += o["quantite_commandee"] or 0
            else:
                unites_mixtes = True
        lignes.append({
            **s,
            "qtes": [round(q, 3) for q in qtes],
            "total": round(sum(qtes), 3),
            "unites_mixtes": unites_mixtes,
        })

    return {
        "granularite": granularite,
        "nb_commandes_total": nb_total,
        "periodes": buckets,
        "lignes": lignes,
    }


# Code de l'article de prestation désossage veau (cf. règle métier ci-dessous).
CODE_PREST_DESOSSAGE = "99864-1"


async def _appliquer_regle_desossage(db, lignes: list) -> list:
    """Règle métier : quand une commande contient du veau (famille=Viande,
    sous-famille=Veau), le fournisseur dont le catalogue contient l'article
    99864-1 reçoit une ligne de prestation désossage dont la quantité (kg) =
    poids total de veau commandé chez CE fournisseur.

    Recalcule à partir du catalogue (source de vérité : famille/sous_famille et
    poids ne sont pas stockés dans panier_lignes). Remplace toute ligne 99864-1
    déjà présente pour rester idempotent. Retourne la liste des lignes augmentée.
    """
    # Charger les infos catalogue nécessaires pour toutes les lignes du panier
    cat_ids = [l["catalogue_fournisseur_id"] for l in lignes if l.get("catalogue_fournisseur_id")]
    cat = {}
    if cat_ids:
        placeholders = ",".join("?" * len(cat_ids))
        cur = await db.execute(
            f"""SELECT id, fournisseur_id, code_article, designation, famille, sous_famille,
                       format_prix, prix_achat_ht, poids_colis_kg, poids_unitaire_kg, qte_par_colis
                FROM catalogue_fournisseur WHERE id IN ({placeholders})""",
            cat_ids,
        )
        cat = {r["id"]: dict(r) for r in await cur.fetchall()}

    def poids_kg(a, qte, unite):
        if not qte:
            return 0.0
        if unite == "kg":
            return qte
        if unite == "piece":
            p = a.get("poids_unitaire_kg")
            return qte * p if p else None
        if unite == "colis":
            p = a.get("poids_colis_kg")
            return qte * p if p else None
        return None

    # Cumuler le poids de veau par fournisseur (hors article prestation lui-même)
    veau_kg = {}
    for l in lignes:
        a = cat.get(l.get("catalogue_fournisseur_id"))
        if not a or a["code_article"] == CODE_PREST_DESOSSAGE:
            continue
        if a.get("famille") == "Viande" and a.get("sous_famille") == "Veau":
            kg = poids_kg(a, l["quantite"], l.get("unite") or "kg")
            if kg:
                veau_kg[a["fournisseur_id"]] = veau_kg.get(a["fournisseur_id"], 0.0) + kg

    # Retirer toute ligne désossage existante (on la régénère proprement)
    lignes = [l for l in lignes if l["code_article"] != CODE_PREST_DESOSSAGE]

    # Ajouter une ligne désossage par fournisseur ayant du veau ET l'article 99864-1
    for fid, kg in veau_kg.items():
        cur = await db.execute(
            """SELECT id, fournisseur_id, code_article, designation, prix_achat_ht
               FROM catalogue_fournisseur
               WHERE fournisseur_id = ? AND code_article = ? AND actif = 1
               LIMIT 1""",
            (fid, CODE_PREST_DESOSSAGE),
        )
        prest = await cur.fetchone()
        if not prest:
            continue  # veau sans prestation au catalogue : ignoré (signalé côté front)
        prest = dict(prest)
        lignes.append({
            "catalogue_fournisseur_id": prest["id"],
            "fournisseur_id": fid,
            "fournisseur_nom": next((l["fournisseur_nom"] for l in lignes if l["fournisseur_id"] == fid), ""),
            "code_article": prest["code_article"],
            "designation": prest["designation"],
            "quantite": round(kg, 3),
            "unite": "kg",
            "prix_ht": prest.get("prix_achat_ht") or 0,
        })
    return lignes


@router.post("/panier/generer", status_code=201)
async def generer_commandes_depuis_panier(body: PanierGenerer):
    """Regroupe le panier par fournisseur et crée une commande brouillon par fournisseur."""
    async with get_db() as db:
        cur = await db.execute(
            "SELECT * FROM panier_lignes WHERE boutique_id = 1 ORDER BY fournisseur_id"
        )
        lignes = [dict(r) for r in await cur.fetchall()]

        if not lignes:
            raise HTTPException(400, "Le panier est vide")

        # Garantir la prestation désossage veau (source de vérité serveur)
        lignes = await _appliquer_regle_desossage(db, lignes)

    # Regrouper par fournisseur
    from collections import defaultdict
    par_fournisseur = defaultdict(list)
    for l in lignes:
        par_fournisseur[l["fournisseur_id"]].append(l)

    commandes_creees = []
    for fournisseur_id, items in par_fournisseur.items():
        cmd_body = CommandeCreate(
            fournisseur_id=fournisseur_id,
            date_livraison_prevue=body.date_livraison_prevue,
            commentaire=body.commentaire,
            lignes=[
                CommandeLigneCreate(
                    catalogue_fournisseur_id=i["catalogue_fournisseur_id"],
                    code_article=i["code_article"],
                    designation=i["designation"],
                    quantite_commandee=i["quantite"],
                    unite=i["unite"],
                    prix_unitaire_ht=i["prix_ht"],
                )
                for i in items
            ]
        )
        cmd = await create_commande(cmd_body)
        commandes_creees.append(cmd)

    # Vider le panier
    async with get_db() as db:
        await db.execute("DELETE FROM panier_lignes WHERE boutique_id = 1")
        await db.commit()

    return {"nb_commandes": len(commandes_creees), "commandes": commandes_creees}


class MappingCreate(BaseModel):
    commande_id: int
    reception_id: int
    personnel_id: Optional[int] = None


@router.post("/commande_receptions_mapping", status_code=201)
async def create_mapping(body: MappingCreate):
    """Lie une réception à une commande (appelé depuis le module réception)."""
    async with get_db() as db:
        try:
            # Nettoyer un éventuel lien vers une réception NON clôturée (fiche
            # abandonnée sans clôture) : on remplace ce lien fantôme par le nouveau.
            await db.execute(
                """DELETE FROM commande_receptions_mapping
                   WHERE commande_id = ?
                     AND reception_id <> ?
                     AND reception_id IN (
                         SELECT id FROM receptions WHERE statut <> 'cloturee'
                     )""",
                (body.commande_id, body.reception_id),
            )
            await db.execute(
                """INSERT OR IGNORE INTO commande_receptions_mapping
                   (commande_id, reception_id, personnel_id)
                   VALUES (?, ?, ?)""",
                (body.commande_id, body.reception_id, body.personnel_id)
            )
            await db.commit()
        except Exception as e:
            raise HTTPException(400, str(e))
        return {"ok": True, "commande_id": body.commande_id, "reception_id": body.reception_id}


@router.post("/commandes/{commande_id}/envoyer")
async def envoyer_commande(commande_id: int):
    """Envoie le récapitulatif de commande par mail au commercial fournisseur."""
    async with get_db() as db:
        cur = await db.execute(
            """SELECT c.*, f.nom AS fournisseur_nom, f.email_commercial, f.emails_copie
               FROM commandes c JOIN fournisseurs f ON f.id = c.fournisseur_id
               WHERE c.id = ?""",
            (commande_id,)
        )
        commande = await cur.fetchone()
        if not commande:
            raise HTTPException(404, "Commande introuvable")

        if not commande["email_commercial"]:
            raise HTTPException(400, "Aucun email commercial renseigné pour ce fournisseur")

        cur2 = await db.execute(
            """SELECT cl.*,
                      cf.prix_achat_ht AS cat_prix_ht,
                      cf.format_prix   AS cat_format_prix
               FROM commande_lignes cl
               LEFT JOIN catalogue_fournisseur cf ON cf.id = cl.catalogue_fournisseur_id
               WHERE cl.commande_id = ? ORDER BY cl.id""",
            (commande_id,)
        )
        lignes = [dict(r) for r in await cur2.fetchall()]

        # Lire config SMTP avant de construire le HTML (from_addr utilisé dans le pied)
        import os, smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        smtp_host = os.getenv("SMTP_HOST", "")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        smtp_user = os.getenv("SMTP_USER", "")
        smtp_pass = os.getenv("SMTP_PASSWORD", "")
        from_addr = os.getenv("SMTP_FROM", smtp_user)

        # Construire le corps du mail (HTML + fallback texte)
        def _fmt_date_fr(d):
            """Date ISO (aaaa-mm-jj) → jj/mm/aa. Renvoie '' si vide/illisible."""
            if not d:
                return ''
            try:
                return datetime.strptime(d[:10], '%Y-%m-%d').strftime('%d/%m/%y')
            except (ValueError, TypeError):
                return d

        def _prix_catalogue(l):
            prix = l.get('cat_prix_ht')
            fmt  = l.get('cat_format_prix') or 'kg'
            if prix is None:
                return '—'
            unite_lbl = '€/kg' if fmt == 'kg' else '€/colis'
            return f"{prix:.2f} {unite_lbl}"

        lignes_rows = "".join(
            f"""<tr>
              <td style="padding:10px 12px;border-bottom:1px solid #f1ead9;font-family:monospace;font-size:13px;color:#5a3e28;">{l['code_article'] or '—'}</td>
              <td style="padding:10px 12px;border-bottom:1px solid #f1ead9;font-size:14px;">{l['designation']}</td>
              <td style="padding:10px 12px;border-bottom:1px solid #f1ead9;text-align:right;white-space:nowrap;">{l['quantite_commandee']} {l['unite']}</td>
              <td style="padding:10px 12px;border-bottom:1px solid #f1ead9;text-align:right;white-space:nowrap;">{_prix_catalogue(l)}</td>
            </tr>"""
            for l in lignes
        )
        # Encart commentaire mis en avant, placé AVANT le tableau de commande.
        commentaire_bloc = (
            '<tr><td style="padding:16px 32px 0;">'
            '<table width="100%" cellpadding="0" cellspacing="0" '
            'style="background:#fff7ed;border:1px solid #f0c9a0;border-left:4px solid #d97706;border-radius:6px;">'
            '<tr><td style="padding:12px 16px;">'
            '<p style="margin:0 0 4px;font-size:12px;font-weight:700;color:#b45309;text-transform:uppercase;letter-spacing:.5px;">'
            '&#128221; Instructions particuli&#232;res</p>'
            f'<p style="margin:0;font-size:14px;color:#2d1f0f;line-height:1.5;">{commande["commentaire"]}</p>'
            '</td></tr></table></td></tr>'
        ) if commande['commentaire'] else ""

        app_url = os.getenv("APP_URL", "http://localhost:8000").rstrip("/")
        logo_url = f"{app_url}/static/assets/logo.png"

        corps_html = (
            '<!DOCTYPE html><html lang="fr"><head><meta charset="utf-8"></head>'
            '<body style="margin:0;padding:0;background:#f5f0e8;font-family:Georgia,serif;color:#2d1f0f;">'
            '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f5f0e8;padding:32px 16px;">'
            '<tr><td align="center">'
            '<table width="620" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:10px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1);">'
            '<tr><td style="background:#6b2d0f;padding:24px 32px;text-align:center;">'
            f'<img src="{logo_url}" alt="Au Comptoir des Lilas" style="max-width:140px;height:auto;margin-bottom:12px;display:block;margin-left:auto;margin-right:auto;">'
            '<p style="margin:0;font-size:20px;font-weight:700;color:#fff;letter-spacing:1px;">Au Comptoir des Lilas</p>'
            '<p style="margin:4px 0 0;font-size:12px;color:#f9d5b0;letter-spacing:2px;text-transform:uppercase;">Boucherie &#8226; Charcuterie</p>'
            '</td></tr>'
            '<tr><td style="padding:24px 32px 8px;border-bottom:2px solid #e8d9c4;">'
            '<p style="margin:0;font-size:18px;font-weight:700;color:#6b2d0f;">Bon de commande</p>'
            '<p style="margin:6px 0 0;font-size:14px;color:#2d1f0f;">'
            '&#128204; <strong>Destinataire : ' + commande['fournisseur_nom'] + '</strong>'
            ' &nbsp;&#183;&nbsp; ' + commande['email_commercial'] +
            '</p>'
            '<p style="margin:4px 0 0;font-size:13px;color:#6b7280;">'
            'N&#176; <strong style="color:#2d1f0f;">' + commande['numero_commande'] + '</strong>'
            ' &nbsp;&#183;&nbsp; Date : <strong style="color:#2d1f0f;">' + _fmt_date_fr(commande['date_commande']) + '</strong>'
            ' &nbsp;&#183;&nbsp; Heure d\'envoi : <strong style="color:#2d1f0f;">' + datetime.now().strftime('%H:%M') + '</strong>'
            ' &nbsp;&#183;&nbsp; Livraison souhait&#233;e : <strong style="color:#2d1f0f;">' + (_fmt_date_fr(commande['date_livraison_prevue']) or '&#192; d&#233;finir') + '</strong>'
            '</p></td></tr>'
            + commentaire_bloc +
            '<tr><td style="padding:0 32px;">'
            '<table width="100%" cellpadding="0" cellspacing="0" style="margin-top:20px;">'
            '<thead><tr style="background:#f3ebdf;">'
            '<th style="padding:10px 12px;text-align:left;font-size:12px;color:#5a3e28;text-transform:uppercase;border-bottom:2px solid #d4c5af;">Code</th>'
            '<th style="padding:10px 12px;text-align:left;font-size:12px;color:#5a3e28;text-transform:uppercase;border-bottom:2px solid #d4c5af;">D&#233;signation</th>'
            '<th style="padding:10px 12px;text-align:right;font-size:12px;color:#5a3e28;text-transform:uppercase;border-bottom:2px solid #d4c5af;">Qt&#233;</th>'
            '<th style="padding:10px 12px;text-align:right;font-size:12px;color:#5a3e28;text-transform:uppercase;border-bottom:2px solid #d4c5af;">Prix HT</th>'
            '</tr></thead>'
            '<tbody>' + lignes_rows + '</tbody>'
            '</table></td></tr>'
            '<tr><td style="background:#f9f4ed;padding:16px 32px;border-top:1px solid #e8d9c4;">'
            '<table width="100%" cellpadding="0" cellspacing="0"><tr>'
            '<td style="font-size:11px;color:#6b7280;line-height:1.6;">'
            '<p style="margin:0 0 8px;font-weight:700;color:#5a3e28;">Au Comptoir des Lilas</p>'
            '<p style="margin:0 0 2px;">122 rue de Paris &nbsp;&#183;&nbsp; 93260 Les Lilas &nbsp;&#183;&nbsp; France</p>'
            '<p style="margin:0;">Tel. 06 88 50 43 41 &nbsp;&#183;&nbsp; ' + from_addr + '</p>'
            '<p style="margin:4px 0 0;font-size:10px;color:#9a7c5a;">SIRET : 103 577 607 00015</p>'
            '</td></tr></table>'
            '</td></tr>'
            '</table></td></tr></table></body></html>'
        )

        corps_txt = "\n".join(
            f"  • {l['code_article']} — {l['designation']} : {l['quantite_commandee']} {l['unite']} "
            f"× {l['prix_unitaire_ht']:.2f}€ HT = {l['montant_ht']:.2f}€ HT"
            for l in lignes
        )
        corps = f"Commande {commande['numero_commande']} du {_fmt_date_fr(commande['date_commande'])}\nDestinataire : {commande['fournisseur_nom']} <{commande['email_commercial']}>\n\n{corps_txt}"

        # Envoi mail via smtplib (config dans variables d'env)
        if not smtp_host or not smtp_user:
            return {
                "envoye": False,
                "message": "Configuration SMTP manquante (SMTP_HOST, SMTP_USER, SMTP_PASSWORD non définis)",
                "destinataire": commande["email_commercial"],
                "sujet": f"Commande {commande['numero_commande']} — Au Comptoir des Lilas",
                "corps": corps,
            }

        import json as _json
        emails_copie = []
        try:
            raw_cc = commande["emails_copie"]
            if raw_cc:
                emails_copie = _json.loads(raw_cc) if isinstance(raw_cc, str) else raw_cc
        except Exception:
            pass

        msg = MIMEMultipart("alternative")
        msg["From"]    = from_addr
        msg["To"]      = commande["email_commercial"]
        if emails_copie:
            msg["Cc"] = ", ".join(emails_copie)
        msg["Subject"] = f"Commande {commande['numero_commande']} — Au Comptoir des Lilas"
        msg.attach(MIMEText(corps, "plain", "utf-8"))
        msg.attach(MIMEText(corps_html, "html", "utf-8"))

        all_recipients = [commande["email_commercial"]] + emails_copie

        try:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.starttls()
                server.login(smtp_user, smtp_pass)
                server.sendmail(from_addr, all_recipients, msg.as_string())

            await db.execute(
                "UPDATE commandes SET statut = 'confirmee', date_envoi_mail = CURRENT_TIMESTAMP WHERE id = ?",
                (commande_id,)
            )
            await db.commit()
            return {"envoye": True, "destinataire": commande["email_commercial"], "copie": emails_copie}

        except Exception as e:
            logger.error("Erreur envoi mail commande %d : %s", commande_id, e)
            raise HTTPException(500, f"Erreur envoi mail : {e}")


# ===========================================================================
# Comparateur fournisseurs
#
# On regroupe plusieurs articles du catalogue (de fournisseurs différents) qui
# désignent le même produit physique, puis on les compare au **prix au kilo**
# normalisé pour arbitrer le meilleur achat. Le €/kg est calculé à la volée
# depuis le catalogue ; quand le poids manque, on n'invente PAS de chiffre.
# ===========================================================================


def _calc_prix_kg(format_prix, prix_achat_ht, poids_colis_kg, famille=None):
    """Prix au kilo normalisé d'un article catalogue.

    - famille 'Viande' : le prix d'achat EST déjà le €/kg (la viande se vend au kilo) ;
      le poids_colis_kg ne sert qu'à la commande, JAMAIS au calcul → pas de division.
    - format 'kg'    : le prix est déjà au kilo → tel quel.
    - format 'colis' : prix du colis / poids total du colis (qté × poids unitaire).
    Retourne None si on ne peut pas le calculer honnêtement (prix absent, ou
    colis sans poids renseigné) → l'UI affiche « €/kg indisponible ».
    """
    # 0 € est un PRIX RÉEL (produit gratuit / offert / échantillon), pas une donnée
    # manquante → €/kg = 0, comparable, marge calculable (100 %). Seul un prix absent
    # (None) ou un colis sans poids reste « indisponible ».
    if prix_achat_ht is None:
        return None
    # Viande : prix d'achat = €/kg direct, quel que soit le format (poids = commande only).
    if (famille or "").strip().lower() == "viande":
        return round(float(prix_achat_ht), 4)
    if format_prix == "kg":
        return round(float(prix_achat_ht), 4)
    if format_prix == "colis" and poids_colis_kg:
        return round(float(prix_achat_ht) / float(poids_colis_kg), 4)
    return None


def _comparer_prix_bl_catalogue(prix_bl, unite_bl, cat_format_prix, cat_prix_ht,
                                cat_poids_colis_kg, cat_famille, seuil_pct):
    """Compare un prix lu sur le BL au prix de référence catalogue, TOUS DEUX en €/kg.

    C'est ici qu'on neutralise le piège kg/colis : on ramène le prix BL ET le prix
    catalogue au €/kg via `_calc_prix_kg` avant de comparer. Sans ça, un prix/colis BL
    confronté à un prix/kg catalogue donnerait un écart absurde.

    Retourne un dict : prix_kg_bl, prix_kg_catalogue, ecart_kg, ecart_pct,
    ecart_significatif (|ecart_pct| > seuil), comparable (False si on ne peut pas
    comparer honnêtement, ex. BL au colis sur de la viande dont le €/kg n'est pas dérivable).
    """
    est_viande = (cat_famille or "").strip().lower() == "viande"

    # €/kg du catalogue (référence) — réutilise la logique existante.
    prix_kg_cat = _calc_prix_kg(cat_format_prix, cat_prix_ht, cat_poids_colis_kg, cat_famille)

    # €/kg du BL : on applique la même normalisation, mais avec l'unité LUE sur le BL.
    # Cas non fiable : prix BL au colis/pièce sur de la viande → _calc_prix_kg prendrait
    # le prix tel quel (règle viande = €/kg direct), ce qui serait faux ici. On le signale.
    comparable = True
    prix_kg_bl = None
    if prix_bl is not None:
        if est_viande and unite_bl in ("colis", "piece"):
            comparable = False  # prix BL pas au kg sur de la viande : non rapprochable sûrement
        else:
            # poids_colis_kg du BL inconnu → on emprunte celui du catalogue (même article).
            prix_kg_bl = _calc_prix_kg(unite_bl or "kg", prix_bl, cat_poids_colis_kg, cat_famille)
            if unite_bl == "colis" and not cat_poids_colis_kg:
                comparable = False  # colis sans poids → €/kg indérivable

    ecart_kg = None
    ecart_pct = None
    significatif = False
    if comparable and prix_kg_bl is not None and prix_kg_cat is not None:
        ecart_kg = round(prix_kg_bl - prix_kg_cat, 4)
        if prix_kg_cat > 0:
            ecart_pct = round((prix_kg_bl - prix_kg_cat) / prix_kg_cat * 100, 2)
            significatif = abs(ecart_pct) > seuil_pct
        elif prix_kg_bl > 0:
            # référence à 0 € mais BL non nul : changement réel, à signaler.
            significatif = True

    return {
        "prix_kg_bl": prix_kg_bl,
        "prix_kg_catalogue": prix_kg_cat,
        "ecart_kg": ecart_kg,
        "ecart_pct": ecart_pct,
        "ecart_significatif": significatif,
        "comparable": comparable,
    }


class ComparerPrixItem(BaseModel):
    catalogue_fournisseur_id: int
    prix_unitaire: Optional[float] = None
    unite_prix: Optional[str] = None         # 'kg' | 'piece' | 'colis' (issu de l'OCR)


class ComparerPrixBody(BaseModel):
    items: List[ComparerPrixItem]
    seuil_pct: float = 2.0                    # écart relatif au-delà duquel on alerte


@router.post("/catalogue/comparer-prix")
async def comparer_prix_catalogue(body: ComparerPrixBody):
    """Compare en lot des prix BL (réception) aux prix de référence catalogue, en €/kg.

    Sert l'alerte d'écart de l'écran réception : pour chaque article du catalogue, on
    renvoie le €/kg du BL, le €/kg de référence, l'écart et s'il est significatif.
    En lot (une réception = N lignes) pour un seul aller-retour réseau (réseau DDNS lent).
    """
    if not body.items:
        return {"resultats": [], "seuil_pct": body.seuil_pct}

    ids = [it.catalogue_fournisseur_id for it in body.items]
    placeholders = ",".join("?" for _ in ids)
    async with get_db() as db:
        cur = await db.execute(
            f"""SELECT id, designation, format_prix, prix_achat_ht, poids_colis_kg, famille
                FROM catalogue_fournisseur WHERE id IN ({placeholders})""",
            ids,
        )
        cat = {r["id"]: dict(r) for r in await cur.fetchall()}

    resultats = []
    for it in body.items:
        art = cat.get(it.catalogue_fournisseur_id)
        if not art:
            resultats.append({
                "catalogue_fournisseur_id": it.catalogue_fournisseur_id,
                "trouve": False,
            })
            continue
        # unite_prix est déjà normalisée par l'OCR (_normaliser_unite_prix → kg|piece|colis).
        # On NE passe PAS par _normaliser_format_prix ici : il écrase 'piece' en 'colis' (OK
        # pour le catalogue, mais on a besoin de distinguer pièce/colis dans la comparaison).
        unite_bl = (it.unite_prix or "").strip().lower() or None
        comp = _comparer_prix_bl_catalogue(
            it.prix_unitaire, unite_bl,
            art["format_prix"], art["prix_achat_ht"], art["poids_colis_kg"], art["famille"],
            body.seuil_pct,
        )
        resultats.append({
            "catalogue_fournisseur_id": it.catalogue_fournisseur_id,
            "trouve": True,
            "designation": art["designation"],
            "prix_achat_ht_catalogue": art["prix_achat_ht"],
            "format_prix_catalogue": art["format_prix"],
            **comp,
        })

    return {"resultats": resultats, "seuil_pct": body.seuil_pct}


def _calc_marge(prix_vente_ttc, tva_percent, achat_ref_kg,
                unite_vente="kg", poids_piece_kg=None):
    """Marge à la volée d'un produit revendu en l'état, au kg OU à la pièce.

    - `achat_ref_kg` = €/kg HT de la ligne fournisseur CHOISIE (⭐), déjà normalisé par
      `_calc_prix_kg`. None → marge indisponible.
    - `prix_vente_ttc` = prix de vente, exprimé dans l'unité de vente (€/kg ou €/pièce).
    - `unite_vente` : 'kg' (défaut) ou 'piece'.
    - `poids_piece_kg` : requis si unité 'piece' (poids d'une pièce vendue) pour ramener
      l'achat €/kg à un coût/pièce. Absent → marge indisponible (jamais inventée).

    Le calcul se fait DANS l'unité de vente :
      - kg    : coût matière = achat €/kg ; marge = vente HT/kg − coût.
      - pièce : coût matière = achat €/kg × poids_piece_kg ; marge = vente HT/pièce − coût.
    Renvoie un dict {unite, base_label, prix_vente_ht, cout_matiere, marge, taux_marge,
    coef, achat_ref_kg, ...} ou None.
    """
    if achat_ref_kg is None or prix_vente_ttc is None:
        return None
    try:
        ttc = float(prix_vente_ttc)
        achat = float(achat_ref_kg)
    except (TypeError, ValueError):
        return None
    if ttc <= 0:
        return None

    tva = float(tva_percent) if tva_percent is not None else 0.0
    prix_vente_ht = ttc / (1.0 + tva / 100.0)

    if unite_vente == "piece":
        if not poids_piece_kg or float(poids_piece_kg) <= 0:
            return None  # poids d'une pièce manquant → marge incalculable
        cout_matiere = achat * float(poids_piece_kg)
        base_label = "€/pièce"
    else:
        cout_matiere = achat  # achat déjà au kg
        base_label = "€/kg"

    marge = prix_vente_ht - cout_matiere
    taux = marge / prix_vente_ht if prix_vente_ht > 0 else None
    coef = ttc / cout_matiere if cout_matiere > 0 else None

    # Pour une vente à la pièce : équivalents au kilo, pour comparer sur la même base
    # que l'achat (qui est en €/kg). vente €/kg = prix_HT/pièce ÷ poids ; achat €/kg = achat_ref.
    if unite_vente == "piece":
        vente_ht_kg = prix_vente_ht / float(poids_piece_kg)
        cout_kg = achat  # le coût matière ramené au kg, c'est l'achat €/kg lui-même
    else:
        vente_ht_kg = prix_vente_ht
        cout_kg = cout_matiere

    return {
        "unite": "piece" if unite_vente == "piece" else "kg",
        "base_label": base_label,
        "prix_vente_ttc": round(ttc, 2),
        "prix_vente_ht": round(prix_vente_ht, 4),
        "tva_percent": round(tva, 2),
        "achat_ref_kg": round(achat, 4),
        "poids_piece_kg": round(float(poids_piece_kg), 4) if poids_piece_kg else None,
        "cout_matiere": round(cout_matiere, 4),
        # Équivalents au kilo (utiles surtout pour la pièce : comparer vente vs achat en €/kg).
        "vente_ht_kg": round(vente_ht_kg, 4),
        "cout_kg": round(cout_kg, 4),
        # 'marge_kg' conservé comme alias générique de la marge dans l'unité (rétrocompat front).
        "marge_kg": round(marge, 4),
        "marge": round(marge, 4),
        "taux_marge": round(taux, 4) if taux is not None else None,
        "coef": round(coef, 4) if coef is not None else None,
    }


def _normaliser_texte(s: str) -> set:
    """Tokens normalisés (minuscule, sans accents, mots de 2+ lettres) d'un libellé,
    pour mesurer la proximité sémantique entre désignations fournisseurs."""
    if not s:
        return set()
    s = unicodedata.normalize("NFKD", s.lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return {t for t in "".join(c if c.isalnum() else " " for c in s).split() if len(t) >= 2}


def _similarite(a: str, b: str) -> float:
    """Indice de Jaccard entre les tokens de deux désignations (0..1)."""
    ta, tb = _normaliser_texte(a), _normaliser_texte(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


@router.get("/comparatif/stats")
async def comparatif_stats(_=Depends(require_admin)):
    """Nombre d'articles actifs du catalogue non encore rattachés à un groupe."""
    async with get_db() as db:
        rows = await db.execute_fetchall(
            """
            SELECT COUNT(*) FROM catalogue_fournisseur c
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            WHERE f.boutique_id = 1 AND c.actif = 1
              AND c.id NOT IN (SELECT catalogue_fournisseur_id FROM comparatif_groupe_ligne)
            """
        )
        return {"articles_non_groupes": rows[0][0] if rows else 0}


@router.get("/comparatif/ventes-non-reliees")
async def comparatif_ventes_non_reliees(_=Depends(require_admin)):
    """Produits du catalogue de VENTE actifs qui ne sont reliés à AUCUN groupe de comparaison
    → aucun suivi de marge. Vue « couverture du catalogue »."""
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT v.id, v.nom, v.prix_vente_ttc, v.unite_vente, v.famille, v.sous_famille
            FROM catalogue_vente v
            WHERE v.boutique_id = 1 AND v.actif = 1
              AND v.id NOT IN (SELECT catalogue_vente_id FROM comparatif_groupe_vente)
            ORDER BY v.famille, v.sous_famille, v.nom
            """
        )
        produits = [dict(r) for r in await cur.fetchall()]
        return {"total": len(produits), "produits": produits}


@router.get("/comparatif/recherche-ventes")
async def comparatif_recherche_ventes(q: str = "", _=Depends(require_admin)):
    """Recherche un produit du catalogue de VENTE par nom et renvoie, pour chacun,
    le groupe de comparaison auquel il est relié (s'il y en a un). Sert à partir du
    produit à vendre pour sauter directement à son VS / sa marge."""
    terme = (q or "").strip()
    async with get_db() as db:
        sql = """
            SELECT v.id, v.nom, v.prix_vente_ttc, v.unite_vente, v.famille, v.sous_famille,
                   gv.groupe_id AS groupe_id, g.nom AS groupe_nom
            FROM catalogue_vente v
            LEFT JOIN comparatif_groupe_vente gv ON gv.catalogue_vente_id = v.id
            LEFT JOIN comparatif_groupe g        ON g.id = gv.groupe_id
            WHERE v.boutique_id = 1 AND v.actif = 1
        """
        params: list = []
        if terme:
            sql += " AND v.nom LIKE ?"
            params.append(f"%{terme}%")
        sql += " ORDER BY v.nom LIMIT 50"
        cur = await db.execute(sql, params)
        produits = [dict(r) for r in await cur.fetchall()]
        return {"total": len(produits), "produits": produits}


@router.post("/comparatif/groupes/from-vente", status_code=201)
async def create_groupe_from_vente(body: ComparatifVenteLink, _=Depends(require_admin)):
    """Crée un groupe de comparaison nommé d'après un produit de vente et lui associe ce produit,
    en un seul appel. Sert à démarrer le suivi de marge d'un produit vendu non encore relié."""
    async with get_db() as db:
        cur_v = await db.execute(
            "SELECT nom, sous_famille FROM catalogue_vente WHERE id = ? AND boutique_id = 1 AND actif = 1",
            (body.catalogue_vente_id,),
        )
        v = await cur_v.fetchone()
        if not v:
            raise HTTPException(404, "Produit de vente introuvable")
        # Refus si déjà relié (cardinalité vente unique).
        cur_d = await db.execute(
            "SELECT 1 FROM comparatif_groupe_vente WHERE catalogue_vente_id = ?",
            (body.catalogue_vente_id,),
        )
        if await cur_d.fetchone():
            raise HTTPException(409, "Ce produit de vente est déjà associé à un groupe")

        cur_g = await db.execute(
            "INSERT INTO comparatif_groupe (boutique_id, nom, sous_famille) VALUES (1, ?, ?)",
            (v["nom"], v["sous_famille"]),
        )
        groupe_id = cur_g.lastrowid
        await db.execute(
            "INSERT INTO comparatif_groupe_vente (groupe_id, catalogue_vente_id) VALUES (?, ?)",
            (groupe_id, body.catalogue_vente_id),
        )
        await db.commit()
        cur2 = await db.execute("SELECT * FROM comparatif_groupe WHERE id = ?", (groupe_id,))
        return dict(await cur2.fetchone())


@router.post("/comparatif/groupes/from-ventes-bulk", status_code=201)
async def create_groupes_from_ventes_bulk(_=Depends(require_admin)):
    """Crée d'un coup un groupe par produit de vente NON ENCORE RELIÉ (nommé d'après le produit,
    associé). Démarre le suivi de marge sur tout le reste du catalogue vente en une fois.
    Idempotent vis-à-vis du re-clic : ne touche pas aux produits déjà reliés."""
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT id, nom, sous_famille FROM catalogue_vente
            WHERE boutique_id = 1 AND actif = 1
              AND id NOT IN (SELECT catalogue_vente_id FROM comparatif_groupe_vente)
            ORDER BY nom
            """
        )
        produits = [dict(r) for r in await cur.fetchall()]
        cree = 0
        for p in produits:
            cur_g = await db.execute(
                "INSERT INTO comparatif_groupe (boutique_id, nom, sous_famille) VALUES (1, ?, ?)",
                (p["nom"], p["sous_famille"]),
            )
            await db.execute(
                "INSERT INTO comparatif_groupe_vente (groupe_id, catalogue_vente_id) VALUES (?, ?)",
                (cur_g.lastrowid, p["id"]),
            )
            cree += 1
        await db.commit()
    return {"crees": cree}


@router.post("/comparatif/viande/reorganiser", status_code=200)
async def reorganiser_viande(_=Depends(require_admin)):
    """Réorganise la VIANDE par sous-famille (Bœuf, Veau…) au lieu d'un groupe par produit.

    1) Supprime les groupes dont le produit de vente associé est de famille « Viande », SANS
       référence d'achat choisie ET SANS aucun article d'achat (vraiment vides → 0 travail perdu).
       Les groupes viande qui ont déjà des achats sont ÉPARGNÉS et listés (epargnes).
    2) Crée un groupe par SOUS-FAMILLE viande ayant ≥ 1 produit de vente, et y associe tous les
       produits de vente viande de cette sous-famille (libérés en 1 ou pas encore reliés).
    Idempotent : relancer ne duplique pas (réutilise un groupe existant de même sous-famille viande).
    """
    async with get_db() as db:
        # — Étape 1 : supprimer les groupes viande vides (sans réf + sans article d'achat) —
        cur = await db.execute(
            """
            SELECT g.id AS groupe_id, gv.catalogue_vente_id, gv.ligne_choisie_id,
                   v.nom AS vente_nom, v.sous_famille,
                   (SELECT COUNT(*) FROM comparatif_groupe_ligne gl WHERE gl.groupe_id = g.id) AS nb_achats
            FROM comparatif_groupe g
            JOIN comparatif_groupe_vente gv ON gv.groupe_id = g.id
            JOIN catalogue_vente v ON v.id = gv.catalogue_vente_id
            WHERE g.boutique_id = 1 AND LOWER(TRIM(COALESCE(v.famille,''))) = 'viande'
            """
        )
        rows = [dict(r) for r in await cur.fetchall()]

        # Sous-familles viande de référence (un groupe nommé ainsi = groupe-cible, à ne PAS supprimer).
        SF_VIANDE = {"boeuf", "bœuf", "veau", "agneau", "porc", "cheval", "volaille"}

        supprimes = 0
        epargnes = []          # groupes viande gardés car ils ont déjà des achats
        produits_a_replacer = []  # (catalogue_vente_id, sous_famille) à regrouper par sous-famille
        for r in rows:
            vide = (r["ligne_choisie_id"] is None) and (r["nb_achats"] == 0)
            # Ne JAMAIS supprimer un groupe-cible (nom = sous-famille viande) : c'est la destination.
            cur_nom = await db.execute("SELECT nom FROM comparatif_groupe WHERE id = ?", (r["groupe_id"],))
            nom_g = (await cur_nom.fetchone())["nom"]
            est_cible = (nom_g or "").strip().lower() in SF_VIANDE
            if est_cible:
                continue  # groupe-cible (destination) : on n'y touche pas, ni suppr ni épargne
            if vide:
                # Supprime le groupe (la liaison vente part en CASCADE) → produit libéré.
                await db.execute("DELETE FROM comparatif_groupe_ligne WHERE groupe_id = ?", (r["groupe_id"],))
                await db.execute("DELETE FROM comparatif_groupe_vente WHERE groupe_id = ?", (r["groupe_id"],))
                await db.execute("DELETE FROM comparatif_groupe WHERE id = ?", (r["groupe_id"],))
                supprimes += 1
                produits_a_replacer.append((r["catalogue_vente_id"], r["sous_famille"]))
            else:
                # Groupe « 1 par produit » avec déjà des achats → on l'épargne (travail réel).
                epargnes.append({"groupe_id": r["groupe_id"], "vente_nom": r["vente_nom"],
                                 "nb_achats": r["nb_achats"]})

        # Ajouter aussi les produits de vente viande NON ENCORE RELIÉS (à regrouper).
        cur_libres = await db.execute(
            """
            SELECT v.id, v.sous_famille FROM catalogue_vente v
            WHERE v.boutique_id = 1 AND v.actif = 1
              AND LOWER(TRIM(COALESCE(v.famille,''))) = 'viande'
              AND v.id NOT IN (SELECT catalogue_vente_id FROM comparatif_groupe_vente)
            """
        )
        for r in await cur_libres.fetchall():
            produits_a_replacer.append((r["id"], r["sous_famille"]))

        # — Étape 2 : (re)grouper par sous-famille —
        # Regrouper les produits à replacer par sous-famille (ignorer ceux sans sous-famille :
        # impossible de les ranger automatiquement → on les laisse non reliés, signalés à part).
        par_sf = {}
        sans_sf = 0
        vus = set()
        for cv_id, sf in produits_a_replacer:
            if cv_id in vus:        # un produit supprimé puis re-capté comme « libre » : 1 seule fois
                continue
            vus.add(cv_id)
            sf_norm = (sf or "").strip()
            if not sf_norm:
                sans_sf += 1
                continue
            par_sf.setdefault(sf_norm, []).append(cv_id)

        groupes_crees = 0
        for sf, cv_ids in par_sf.items():
            # Réutiliser un groupe viande existant de cette sous-famille s'il y en a un
            # (épargné à l'étape 1), sinon en créer un nommé d'après la sous-famille.
            cur_ex = await db.execute(
                """
                SELECT g.id FROM comparatif_groupe g
                WHERE g.boutique_id = 1 AND g.nom = ? AND g.sous_famille = ? LIMIT 1
                """,
                (sf, sf),
            )
            ex = await cur_ex.fetchone()
            if ex:
                groupe_id = ex["id"]
            else:
                cur_g = await db.execute(
                    "INSERT INTO comparatif_groupe (boutique_id, nom, sous_famille) VALUES (1, ?, ?)",
                    (sf, sf),
                )
                groupe_id = cur_g.lastrowid
                groupes_crees += 1
            for cv_id in cv_ids:
                await db.execute(
                    "INSERT OR IGNORE INTO comparatif_groupe_vente (groupe_id, catalogue_vente_id) VALUES (?, ?)",
                    (groupe_id, cv_id),
                )

        await db.commit()
    return {
        "groupes_supprimes": supprimes,
        "groupes_crees": groupes_crees,
        "sous_familles": sorted(par_sf.keys()),
        "produits_sans_sous_famille": sans_sf,
        "groupes_epargnes": epargnes,
    }


@router.get("/comparatif/groupes/{groupe_id}/achats-suggestions")
async def comparatif_achats_suggestions(groupe_id: int, _=Depends(require_admin)):
    """Articles d'ACHAT du catalogue à proposer pour CE groupe, par proximité sémantique
    avec le nom du groupe (= le produit de vente). Sert à remplir un groupe créé depuis un
    produit de vente. Exclut les articles déjà dans le groupe. Trié par similarité décroissante
    puis €/kg croissant. Seuls les articles avec un minimum de proximité (> 0) sont renvoyés.

    Si le produit de vente associé a une sous-famille, les achats de la MÊME sous-famille sont
    BOOSTÉS (remontent en tête) — préférence, pas exclusion : tout reste visible."""
    async with get_db() as db:
        cur = await db.execute(
            "SELECT nom FROM comparatif_groupe WHERE id = ? AND boutique_id = 1", (groupe_id,)
        )
        grp = await cur.fetchone()
        if not grp:
            raise HTTPException(404, "Groupe introuvable")
        nom = grp["nom"]

        # Sous-famille du produit de vente associé (pour le boost de pertinence).
        cur_sf = await db.execute(
            """SELECT v.sous_famille FROM comparatif_groupe_vente gv
               JOIN catalogue_vente v ON v.id = gv.catalogue_vente_id
               WHERE gv.groupe_id = ? LIMIT 1""",
            (groupe_id,),
        )
        row_sf = await cur_sf.fetchone()
        sf_vente = (row_sf["sous_famille"] or "").strip().lower() if row_sf else ""

        cur_d = await db.execute(
            "SELECT catalogue_fournisseur_id FROM comparatif_groupe_ligne WHERE groupe_id = ?",
            (groupe_id,),
        )
        deja = {r[0] for r in await cur_d.fetchall()}

        cur_c = await db.execute(
            """
            SELECT c.*, f.nom AS fournisseur_nom
            FROM catalogue_fournisseur c
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            WHERE f.boutique_id = 1 AND c.actif = 1
            """
        )
        # Cœur sémantique du nom de groupe (mêmes règles que le clustering : sans packaging/unités).
        coeur_nom = _coeur_tokens(nom)
        suggestions = []
        for r in await cur_c.fetchall():
            a = dict(r)
            if a["id"] in deja:
                continue
            score = _similarite(nom, a["designation"])
            # Exiger au moins un mot-cœur commun pour éviter le bruit (ex. juste « de »).
            if score <= 0 or not (coeur_nom & _coeur_tokens(a["designation"])):
                continue
            # Boost si même sous-famille que le produit de vente (préférence).
            meme_sf = bool(sf_vente) and (a.get("sous_famille") or "").strip().lower() == sf_vente
            a["meme_sous_famille"] = meme_sf
            a["score"] = round(score, 3)
            a["prix_kg"] = _calc_prix_kg(a.get("format_prix"), a.get("prix_achat_ht"), a.get("poids_colis_kg"), a.get("famille"))
            suggestions.append(a)

        # Tri : même sous-famille d'abord, puis score, puis €/kg croissant (None en dernier).
        suggestions.sort(key=lambda x: (
            not x["meme_sous_famille"],
            -x["score"],
            x["prix_kg"] is None,
            x["prix_kg"] if x["prix_kg"] is not None else 0.0,
        ))
        return suggestions[:40]


@router.get("/comparatif/marge-incalculable")
async def comparatif_marge_incalculable(_=Depends(require_admin)):
    """Groupes DÉJÀ associés à un produit de vente, mais dont la marge ne peut pas
    être calculée parce qu'il manque une info en amont. On ne signale que là où il y
    a une intention de marge (produit de vente relié) — pas tout le catalogue.

    Une entrée par PRODUIT DE VENTE associé (1 groupe → N ventes) dont la marge est bloquée.
    Motif (premier rencontré) :
      - 'reference'   : aucune ligne fournisseur de référence (⭐) choisie pour le groupe.
      - 'prix_achat'  : la référence existe mais son €/kg est indisponible (prix absent / colis sans poids).
      - 'prix_vente'  : le produit de vente n'a pas de prix de vente.
      - 'poids_piece' : produit vendu à la pièce sans poids unitaire renseigné.
    (0 € d'achat n'est PAS un trou : produit gratuit, marge = 100 %.)
    """
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT g.id AS groupe_id, g.nom AS groupe_nom, gv.ligne_choisie_id,
                   v.id AS catalogue_vente_id, v.nom AS vente_nom, v.prix_vente_ttc,
                   v.unite_vente, v.poids_piece_kg
            FROM comparatif_groupe_vente gv
            JOIN comparatif_groupe g ON g.id = gv.groupe_id
            JOIN catalogue_vente v   ON v.id = gv.catalogue_vente_id
            WHERE g.boutique_id = 1
            ORDER BY g.nom, v.nom
            """
        )
        assoc = [dict(r) for r in await cur.fetchall()]

        # €/kg des références (mémoïsé par ligne_choisie_id).
        ref_cache = {}
        async def ref_kg(ligne_id):
            if ligne_id in ref_cache:
                return ref_cache[ligne_id]
            cur_l = await db.execute(
                "SELECT format_prix, prix_achat_ht, poids_colis_kg, famille FROM catalogue_fournisseur WHERE id = ?",
                (ligne_id,),
            )
            l = await cur_l.fetchone()
            pk = _calc_prix_kg(l["format_prix"], l["prix_achat_ht"], l["poids_colis_kg"], l["famille"]) if l else None
            ref_cache[ligne_id] = pk
            return pk

        bloques = []
        for a in assoc:
            motif = detail = None
            if a["ligne_choisie_id"] is None:
                motif, detail = "reference", "Aucun fournisseur de référence choisi"
            elif await ref_kg(a["ligne_choisie_id"]) is None:
                motif, detail = "prix_achat", "Prix au kilo de la référence indisponible (prix manquant ou colis sans poids)"
            elif a["prix_vente_ttc"] is None:
                motif, detail = "prix_vente", "Pas de prix de vente sur le produit"
            elif (a.get("unite_vente") == "piece") and not a.get("poids_piece_kg"):
                motif, detail = "poids_piece", "Vendu à la pièce mais poids d'une pièce non renseigné"
            if motif:
                bloques.append({
                    "groupe_id": a["groupe_id"],
                    "groupe_nom": a["groupe_nom"],
                    "catalogue_vente_id": a["catalogue_vente_id"],
                    "vente_nom": a["vente_nom"],
                    "motif": motif,
                    "detail": detail,
                })
        return {"total": len(bloques), "groupes": bloques}


@router.get("/comparatif/groupes")
async def list_comparatif_groupes(_=Depends(require_admin)):
    """Liste des groupes de comparaison, avec le nombre d'articles de chacun."""
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT g.*, COUNT(gl.catalogue_fournisseur_id) AS nb_lignes
            FROM comparatif_groupe g
            LEFT JOIN comparatif_groupe_ligne gl ON gl.groupe_id = g.id
            WHERE g.boutique_id = 1
            GROUP BY g.id
            ORDER BY g.nom
            """
        )
        return [dict(r) for r in await cur.fetchall()]


@router.post("/comparatif/groupes", status_code=201)
async def create_comparatif_groupe(body: ComparatifGroupeCreate, _=Depends(require_admin)):
    nom = (body.nom or "").strip()
    if not nom:
        raise HTTPException(400, "Le nom du groupe est obligatoire")
    async with get_db() as db:
        cur = await db.execute(
            "INSERT INTO comparatif_groupe (boutique_id, nom, sous_famille) VALUES (1, ?, ?)",
            (nom, body.sous_famille),
        )
        await db.commit()
        cur2 = await db.execute("SELECT * FROM comparatif_groupe WHERE id = ?", (cur.lastrowid,))
        return dict(await cur2.fetchone())


@router.put("/comparatif/groupes/{groupe_id}")
async def update_comparatif_groupe(groupe_id: int, body: ComparatifGroupeUpdate, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute(
            "SELECT id FROM comparatif_groupe WHERE id = ? AND boutique_id = 1", (groupe_id,)
        )
        if not await cur.fetchone():
            raise HTTPException(404, "Groupe introuvable")
        fields = {k: v for k, v in body.model_dump().items() if v is not None}
        if "nom" in fields:
            fields["nom"] = fields["nom"].strip()
            if not fields["nom"]:
                raise HTTPException(400, "Le nom ne peut pas être vide")
        if not fields:
            raise HTTPException(400, "Aucun champ à modifier")
        set_clause = ", ".join(f"{k} = ?" for k in fields)
        await db.execute(
            f"UPDATE comparatif_groupe SET {set_clause} WHERE id = ?",
            list(fields.values()) + [groupe_id],
        )
        await db.commit()
        cur2 = await db.execute("SELECT * FROM comparatif_groupe WHERE id = ?", (groupe_id,))
        return dict(await cur2.fetchone())


@router.delete("/comparatif/groupes/{groupe_id}")
async def delete_comparatif_groupe(groupe_id: int, _=Depends(require_admin)):
    async with get_db() as db:
        cur = await db.execute(
            "SELECT id FROM comparatif_groupe WHERE id = ? AND boutique_id = 1", (groupe_id,)
        )
        if not await cur.fetchone():
            raise HTTPException(404, "Groupe introuvable")
        await db.execute("DELETE FROM comparatif_groupe_ligne WHERE groupe_id = ?", (groupe_id,))
        await db.execute("DELETE FROM comparatif_groupe WHERE id = ?", (groupe_id,))
        await db.commit()
        return {"supprime": True}


@router.post("/comparatif/groupes/{groupe_id}/lignes", status_code=201)
async def add_comparatif_ligne(groupe_id: int, body: ComparatifLigneAdd, _=Depends(require_admin)):
    """Ajoute un article du catalogue au groupe de comparaison."""
    async with get_db() as db:
        cur = await db.execute(
            "SELECT id FROM comparatif_groupe WHERE id = ? AND boutique_id = 1", (groupe_id,)
        )
        if not await cur.fetchone():
            raise HTTPException(404, "Groupe introuvable")
        cur_a = await db.execute(
            """SELECT c.id FROM catalogue_fournisseur c
               JOIN fournisseurs f ON f.id = c.fournisseur_id
               WHERE c.id = ? AND f.boutique_id = 1""",
            (body.catalogue_fournisseur_id,),
        )
        if not await cur_a.fetchone():
            raise HTTPException(404, "Article catalogue introuvable")
        await db.execute(
            "INSERT OR IGNORE INTO comparatif_groupe_ligne (groupe_id, catalogue_fournisseur_id) VALUES (?, ?)",
            (groupe_id, body.catalogue_fournisseur_id),
        )
        await db.commit()
        return {"ajoute": True}


@router.delete("/comparatif/groupes/{groupe_id}/lignes/{cat_id}")
async def remove_comparatif_ligne(groupe_id: int, cat_id: int, _=Depends(require_admin)):
    async with get_db() as db:
        await db.execute(
            "DELETE FROM comparatif_groupe_ligne WHERE groupe_id = ? AND catalogue_fournisseur_id = ?",
            (groupe_id, cat_id),
        )
        # Nettoyer les références orphelines : tout produit de vente qui pointait cette ligne
        # d'achat perd sa référence (sa marge retombera « indisponible » proprement).
        await db.execute(
            "UPDATE comparatif_groupe_vente SET ligne_choisie_id = NULL WHERE groupe_id = ? AND ligne_choisie_id = ?",
            (groupe_id, cat_id),
        )
        # (Legacy) ancienne référence globale du groupe, si encore renseignée.
        await db.execute(
            "UPDATE comparatif_groupe SET ligne_choisie_id = NULL WHERE id = ? AND ligne_choisie_id = ?",
            (groupe_id, cat_id),
        )
        await db.commit()
        return {"retire": True}


@router.get("/comparatif/groupes/{groupe_id}")
async def get_comparatif(groupe_id: int, _=Depends(require_admin)):
    """Le « VS » : les articles du groupe enrichis du prix €/kg normalisé, le moins
    cher (par €/kg calculable) marqué `meilleur`."""
    async with get_db() as db:
        cur = await db.execute(
            "SELECT * FROM comparatif_groupe WHERE id = ? AND boutique_id = 1", (groupe_id,)
        )
        groupe = await cur.fetchone()
        if not groupe:
            raise HTTPException(404, "Groupe introuvable")

        cur_l = await db.execute(
            """
            SELECT c.*, f.nom AS fournisseur_nom
            FROM comparatif_groupe_ligne gl
            JOIN catalogue_fournisseur c ON c.id = gl.catalogue_fournisseur_id
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            WHERE gl.groupe_id = ?
            ORDER BY f.nom, c.designation
            """,
            (groupe_id,),
        )
        lignes = [dict(r) for r in await cur_l.fetchall()]

        groupe = dict(groupe)

        # €/kg de chaque ligne d'achat (indexé pour la marge par produit de vente).
        prix_kg_par_ligne = {}
        for ligne in lignes:
            pk = _calc_prix_kg(
                ligne.get("format_prix"), ligne.get("prix_achat_ht"), ligne.get("poids_colis_kg"), ligne.get("famille")
            )
            ligne["prix_kg"] = pk
            ligne["meilleur"] = False
            prix_kg_par_ligne[ligne["id"]] = pk

        prix_valides = [l["prix_kg"] for l in lignes if l["prix_kg"] is not None]
        if prix_valides:
            meilleur_prix = min(prix_valides)
            for ligne in lignes:
                if ligne["prix_kg"] == meilleur_prix:
                    ligne["meilleur"] = True

        # Tri des colonnes du moins cher au plus cher (par €/kg normalisé). Les articles
        # sans €/kg calculable (poids colis manquant) passent EN TÊTE pour qu'on pense à
        # les compléter. Tri stable secondaire = ordre SQL (fournisseur, désignation).
        lignes.sort(key=lambda l: (l["prix_kg"] is not None, l["prix_kg"] if l["prix_kg"] is not None else 0.0))

        # Produits de vente associés (1 groupe → N ventes). CHAQUE produit choisit SA propre
        # ligne d'achat de référence (gv.ligne_choisie_id) — 3 cordons bleus vendus = 3 achats
        # différents. Pas de référence choisie → marge indisponible pour ce produit (pas d'héritage).
        cur_v = await db.execute(
            """
            SELECT v.*, gv.ligne_choisie_id AS ligne_choisie_id
            FROM comparatif_groupe_vente gv
            JOIN catalogue_vente v ON v.id = gv.catalogue_vente_id
            WHERE gv.groupe_id = ? AND v.boutique_id = 1
            ORDER BY v.nom
            """,
            (groupe_id,),
        )
        produits_vente = []
        for r in await cur_v.fetchall():
            pv = dict(r)
            ref = pv.get("ligne_choisie_id")
            # La ligne de référence doit appartenir au groupe (sinon ignorée).
            achat_ref_kg = prix_kg_par_ligne.get(ref) if ref in prix_kg_par_ligne else None
            pv["marge"] = _calc_marge(
                pv.get("prix_vente_ttc"),
                pv.get("tva_percent"),
                achat_ref_kg,
                unite_vente=pv.get("unite_vente") or "kg",
                poids_piece_kg=pv.get("poids_piece_kg"),
            )
            produits_vente.append(pv)

        return {
            "groupe": groupe,
            "lignes": lignes,
            "produits_vente": produits_vente,
            # Compat ascendante : 1er produit + sa marge (front legacy / lecture simple).
            "catalogue_vente": produits_vente[0] if produits_vente else None,
            "marge": produits_vente[0]["marge"] if produits_vente else None,
        }


async def _ensure_groupe(db, groupe_id):
    cur = await db.execute(
        "SELECT id FROM comparatif_groupe WHERE id = ? AND boutique_id = 1", (groupe_id,)
    )
    if not await cur.fetchone():
        raise HTTPException(404, "Groupe introuvable")


@router.post("/comparatif/groupes/{groupe_id}/ventes", status_code=201)
async def add_comparatif_vente(
    groupe_id: int, body: ComparatifVenteLink, _=Depends(require_admin)
):
    """Associe un produit de VENTE au groupe (1 groupe → N produits de vente).

    Unicité côté vente : un produit n'appartient qu'à UN groupe. S'il est déjà ailleurs, on le
    DÉPLACE vers ce groupe (retire l'ancienne liaison + sa référence d'achat, devenue caduque).
    """
    async with get_db() as db:
        await _ensure_groupe(db, groupe_id)
        cur_v = await db.execute(
            "SELECT id FROM catalogue_vente WHERE id = ? AND boutique_id = 1 AND actif = 1",
            (body.catalogue_vente_id,),
        )
        if not await cur_v.fetchone():
            raise HTTPException(404, "Produit de vente introuvable")
        # Déjà rattaché ailleurs ? → on déplace (supprime l'ancienne liaison).
        cur_dup = await db.execute(
            "SELECT groupe_id FROM comparatif_groupe_vente WHERE catalogue_vente_id = ?",
            (body.catalogue_vente_id,),
        )
        dup = await cur_dup.fetchone()
        if dup and dup["groupe_id"] != groupe_id:
            await db.execute(
                "DELETE FROM comparatif_groupe_vente WHERE catalogue_vente_id = ?",
                (body.catalogue_vente_id,),
            )
        await db.execute(
            # ligne_choisie_id NULL : la réf d'achat de l'ancien groupe ne vaut plus ici.
            "INSERT OR REPLACE INTO comparatif_groupe_vente (groupe_id, catalogue_vente_id, ligne_choisie_id) VALUES (?, ?, NULL)",
            (groupe_id, body.catalogue_vente_id),
        )
        await db.commit()
    return await get_comparatif(groupe_id, _)


@router.delete("/comparatif/groupes/{groupe_id}/ventes/{cv_id}")
async def remove_comparatif_vente(groupe_id: int, cv_id: int, _=Depends(require_admin)):
    """Délie un produit de vente du groupe (le produit lui-même n'est pas supprimé)."""
    async with get_db() as db:
        await _ensure_groupe(db, groupe_id)
        await db.execute(
            "DELETE FROM comparatif_groupe_vente WHERE groupe_id = ? AND catalogue_vente_id = ?",
            (groupe_id, cv_id),
        )
        await db.commit()
    return await get_comparatif(groupe_id, _)


@router.put("/comparatif/groupes/{groupe_id}/ventes/{cv_id}")
async def update_comparatif_vente(
    groupe_id: int, cv_id: int, body: ComparatifVenteUpdate, _=Depends(require_admin)
):
    """Édite un produit de vente associé depuis le comparateur (prix TTC, unité, poids pièce)
    pour simuler la marge. N'agit que sur un produit RÉELLEMENT rattaché à ce groupe."""
    fournis = body.model_fields_set
    async with get_db() as db:
        await _ensure_groupe(db, groupe_id)
        cur = await db.execute(
            "SELECT 1 FROM comparatif_groupe_vente WHERE groupe_id = ? AND catalogue_vente_id = ?",
            (groupe_id, cv_id),
        )
        if not await cur.fetchone():
            raise HTTPException(404, "Ce produit de vente n'est pas associé à ce groupe")

        sets, vals = [], []
        if "prix_vente_ttc" in fournis:
            sets.append("prix_vente_ttc = ?"); vals.append(body.prix_vente_ttc)
        if "unite_vente" in fournis:
            unite = body.unite_vente if body.unite_vente in ("kg", "piece") else "kg"
            sets.append("unite_vente = ?"); vals.append(unite)
        if "poids_piece_kg" in fournis:
            sets.append("poids_piece_kg = ?"); vals.append(body.poids_piece_kg)
        if "famille" in fournis:
            sets.append("famille = ?"); vals.append(body.famille or None)
        if "sous_famille" in fournis:
            sets.append("sous_famille = ?"); vals.append(body.sous_famille or None)
        if sets:
            vals += [cv_id]
            await db.execute(
                f"UPDATE catalogue_vente SET {', '.join(sets)} WHERE id = ? AND boutique_id = 1",
                vals,
            )
            await db.commit()
    return await get_comparatif(groupe_id, _)


@router.put("/comparatif/groupes/{groupe_id}/ventes/{cv_id}/reference")
async def set_comparatif_vente_reference(
    groupe_id: int, cv_id: int, body: ComparatifReferenceUpdate, _=Depends(require_admin)
):
    """Désigne la ligne d'achat de RÉFÉRENCE propre à CE produit de vente (sa marge se calcule
    dessus). null = retirer la référence. La ligne doit appartenir au groupe."""
    async with get_db() as db:
        await _ensure_groupe(db, groupe_id)
        cur = await db.execute(
            "SELECT 1 FROM comparatif_groupe_vente WHERE groupe_id = ? AND catalogue_vente_id = ?",
            (groupe_id, cv_id),
        )
        if not await cur.fetchone():
            raise HTTPException(404, "Ce produit de vente n'est pas associé à ce groupe")

        ref_id = body.ligne_choisie_id
        if ref_id is not None:
            cur_m = await db.execute(
                """SELECT 1 FROM comparatif_groupe_ligne
                   WHERE groupe_id = ? AND catalogue_fournisseur_id = ?""",
                (groupe_id, ref_id),
            )
            if not await cur_m.fetchone():
                raise HTTPException(400, "Cette ligne d'achat n'appartient pas au groupe")

        await db.execute(
            "UPDATE comparatif_groupe_vente SET ligne_choisie_id = ? WHERE groupe_id = ? AND catalogue_vente_id = ?",
            (ref_id, groupe_id, cv_id),
        )
        await db.commit()
    return await get_comparatif(groupe_id, _)


@router.get("/comparatif/groupes/{groupe_id}/vente-suggestions")
async def comparatif_vente_suggestions(
    groupe_id: int,
    q: Optional[str] = Query(None, description="Recherche libre ; vide = suggestions sémantiques"),
    famille: Optional[str] = Query(None),
    sous_famille: Optional[str] = Query(None),
    _=Depends(require_admin),
):
    """Produits du catalogue de VENTE à proposer pour l'association du groupe.

    - q et/ou filtre famille/sous-famille fourni → recherche filtrée (LIKE + égalité), triée alpha.
    - rien fourni → suggestions SÉMANTIQUES : classe TOUS les produits actifs par proximité (Jaccard)
      avec le NOM DU GROUPE, les plus proches d'abord (longue traîne conservée).
    Exclut les produits DÉJÀ associés à un groupe (cardinalité vente unique).
    """
    async with get_db() as db:
        cur = await db.execute(
            "SELECT nom FROM comparatif_groupe WHERE id = ? AND boutique_id = 1",
            (groupe_id,),
        )
        grp = await cur.fetchone()
        if not grp:
            raise HTTPException(404, "Groupe introuvable")
        nom_groupe = grp["nom"]

        # Groupe actuel de chaque produit déjà associé (pour proposer un DÉPLACEMENT).
        # On exclut seulement ceux déjà dans CE groupe ; les autres sont proposés avec mention.
        cur_d = await db.execute(
            """SELECT gv.catalogue_vente_id, gv.groupe_id, g.nom AS groupe_nom
               FROM comparatif_groupe_vente gv JOIN comparatif_groupe g ON g.id = gv.groupe_id"""
        )
        assoc = {r["catalogue_vente_id"]: (r["groupe_id"], r["groupe_nom"]) for r in await cur_d.fetchall()}

        def enrichir(p):
            ga = assoc.get(p["id"])
            p["groupe_actuel_id"] = ga[0] if ga else None
            p["groupe_actuel_nom"] = ga[1] if ga else None
            return p

        def garder(p):
            ga = assoc.get(p["id"])
            return not (ga and ga[0] == groupe_id)  # exclure uniquement ceux déjà dans CE groupe

        q = (q or "").strip()
        if q or famille or sous_famille:
            sql = "SELECT * FROM catalogue_vente WHERE boutique_id = 1 AND actif = 1"
            params = []
            if q:
                sql += " AND nom LIKE ?"; params.append(f"%{q}%")
            if famille:
                sql += " AND famille = ?"; params.append(famille)
            if sous_famille:
                sql += " AND sous_famille = ?"; params.append(sous_famille)
            sql += " ORDER BY nom LIMIT 100"
            cur_v = await db.execute(sql, params)
            items = [enrichir(dict(r)) for r in await cur_v.fetchall() if garder(dict(r))]
            return items[:100]

        # Suggestions sémantiques sur le nom du groupe.
        cur_v = await db.execute(
            "SELECT * FROM catalogue_vente WHERE boutique_id = 1 AND actif = 1 ORDER BY nom"
        )
        produits = [enrichir(dict(r)) for r in await cur_v.fetchall() if garder(dict(r))]
        for p in produits:
            p["score"] = _similarite(nom_groupe, p.get("nom") or "")
        # Plus proches d'abord ; à score égal, les LIBRES avant ceux à déplacer.
        produits.sort(key=lambda p: (-p["score"], p["groupe_actuel_id"] is not None, (p.get("nom") or "").lower()))
        return produits[:60]


@router.put("/comparatif/groupes/{groupe_id}/reference")
async def set_comparatif_reference(
    groupe_id: int, body: ComparatifReferenceUpdate, _=Depends(require_admin)
):
    """Désigne (ou retire) la ligne fournisseur de RÉFÉRENCE du groupe — l'arbitrage manuel
    de l'utilisateur (⭐). La ligne doit appartenir au groupe. null = retirer l'étoile."""
    async with get_db() as db:
        cur = await db.execute(
            "SELECT id FROM comparatif_groupe WHERE id = ? AND boutique_id = 1", (groupe_id,)
        )
        if not await cur.fetchone():
            raise HTTPException(404, "Groupe introuvable")

        ref_id = body.ligne_choisie_id
        if ref_id is not None:
            cur_m = await db.execute(
                """SELECT 1 FROM comparatif_groupe_ligne
                   WHERE groupe_id = ? AND catalogue_fournisseur_id = ?""",
                (groupe_id, ref_id),
            )
            if not await cur_m.fetchone():
                raise HTTPException(400, "Cette ligne n'appartient pas au groupe")

        await db.execute(
            "UPDATE comparatif_groupe SET ligne_choisie_id = ? WHERE id = ?",
            (ref_id, groupe_id),
        )
        await db.commit()
    return await get_comparatif(groupe_id, _)


@router.get("/comparatif/suggestions")
async def comparatif_suggestions(
    ligne_id: int = Query(..., description="Article catalogue de référence"),
    groupe_id: Optional[int] = Query(None, description="Exclure les articles déjà dans ce groupe"),
    _=Depends(require_admin),
):
    """Suggère des articles catalogue proches de `ligne_id` : même sous-famille,
    triés par similarité de désignation. Assiste la construction d'un groupe ;
    l'utilisateur garde la décision finale."""
    async with get_db() as db:
        cur = await db.execute(
            """SELECT c.* FROM catalogue_fournisseur c
               JOIN fournisseurs f ON f.id = c.fournisseur_id
               WHERE c.id = ? AND f.boutique_id = 1""",
            (ligne_id,),
        )
        ref = await cur.fetchone()
        if not ref:
            raise HTTPException(404, "Article de référence introuvable")
        ref = dict(ref)

        deja = set()
        if groupe_id:
            cur_d = await db.execute(
                "SELECT catalogue_fournisseur_id FROM comparatif_groupe_ligne WHERE groupe_id = ?",
                (groupe_id,),
            )
            deja = {r[0] for r in await cur_d.fetchall()}

        # Strict : même sous-famille uniquement (si renseignée sur la référence).
        sql = """
            SELECT c.*, f.nom AS fournisseur_nom
            FROM catalogue_fournisseur c
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            WHERE f.boutique_id = 1 AND c.actif = 1 AND c.id != ?
        """
        params = [ref["id"]]
        if ref.get("sous_famille"):
            sql += " AND c.sous_famille = ?"
            params.append(ref["sous_famille"])
        cur_c = await db.execute(sql, params)
        candidats = [dict(r) for r in await cur_c.fetchall()]

        suggestions = []
        for c in candidats:
            if c["id"] in deja:
                continue
            score = _similarite(ref["designation"], c["designation"])
            if score > 0:
                c["score"] = round(score, 3)
                c["prix_kg"] = _calc_prix_kg(
                    c.get("format_prix"), c.get("prix_achat_ht"), c.get("poids_colis_kg"), c.get("famille")
                )
                suggestions.append(c)

        suggestions.sort(key=lambda x: x["score"], reverse=True)
        return suggestions[:20]


# ---------------------------------------------------------------------------
# Auto-proposition de groupes (clustering assisté)
#
# Réglages calibrés sur un export réel (725 articles, 5 fournisseurs, juin 2026).
# La désignation est réduite à son « cœur sémantique » (on retire chiffres, unités
# et mots de conditionnement/marque/label) ; deux articles sont rapprochés s'ils
# partagent ≥ 2 mots-cœur ET au moins un mot PIVOT (le morceau/type de produit :
# cuisse, filet, quiche…), avec un indice de Jaccard ≥ SEUIL. Le pivot évite la
# grappe fourre-tout « tout le poulet ». Ces constantes sont volontairement en clair
# pour être ajustées si les libellés fournisseurs évoluent.
# ---------------------------------------------------------------------------

CLUSTER_SEUIL = 0.30          # indice de Jaccard minimal entre deux cœurs sémantiques
CLUSTER_MIN_TOKENS = 2        # nb minimal de mots-cœur communs
CLUSTER_MIN_FOURNISSEURS = 2  # une grappe n'a d'intérêt qu'avec ≥ 2 fournisseurs

# Unités de mesure (retirées du cœur).
_CLUSTER_UNITES = {"kg", "g", "gr", "l", "cl", "ml", "mg"}

# Conditionnement, marques, labels, prépositions : du bruit pour le rapprochement.
_CLUSTER_PACK = {
    "colis", "carton", "barquette", "bqt", "sachet", "scht", "sous", "vide", "satm",
    "atm", "pac", "pce", "piece", "pieces", "tranche", "tranches", "pc", "pcs", "tp",
    "plat", "col", "pv", "ofr", "sda", "vf", "aop", "hve", "bbc", "gde", "igp", "vpf",
    "bn", "pp", "sv", "vrac", "format", "demi", "lune", "entiere", "entier", "plaque",
    "poche", "seau", "bidon", "pot", "boite", "unite", "unites", "sup", "superieure",
    "superieur", "pur", "pure", "av", "peau", "pch", "bio", "certifie", "fermier",
    "fermiere", "frais", "nature", "nu",
    "de", "des", "la", "le", "les", "du", "au", "aux", "et", "en", "a",
}

# Mots « pivot » = morceau ou type de produit. Au moins un doit être commun.
_CLUSTER_PIVOTS = {
    "cuisse", "haut", "filet", "escalope", "aiguillette", "magret", "roti", "tournedos",
    "pave", "supreme", "manchon", "pilon", "aile", "gesier", "foie", "saute", "emince",
    "paupiette", "ballotine", "coffre", "grignette", "chorizo", "saucisson", "saucisse",
    "chipolata", "boudin", "andouillette", "merguez", "quiche", "pizza", "panini",
    "croissant", "parmentier", "gratin", "lasagne", "brandade", "puree", "taboule",
    "mousse", "terrine", "pate", "rillette", "cordon", "nugget", "donuts", "coquelet",
    "dinde", "canard", "poulet", "lapin", "caille", "pintade", "morue", "saumon",
    "jambon", "bacon", "boeuf", "veau", "agneau", "porc",
}


def _coeur_tokens(designation: str) -> set:
    """Cœur sémantique : tokens normalisés d'une désignation, débarrassés des
    nombres, unités et mots de conditionnement/marque/label."""
    s = unicodedata.normalize("NFKD", str(designation or "").lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    toks = "".join(c if c.isalnum() else " " for c in s).split()
    out = set()
    for t in toks:
        if any(ch.isdigit() for ch in t):
            continue
        if t in _CLUSTER_UNITES or t in _CLUSTER_PACK or len(t) < 2:
            continue
        out.add(t)
    return out


def _coeurs_proches(a: set, b: set) -> bool:
    """Deux cœurs sont rapprochés s'ils partagent ≥ CLUSTER_MIN_TOKENS mots dont
    au moins un PIVOT, avec un Jaccard ≥ CLUSTER_SEUIL."""
    inter = a & b
    if len(inter) < CLUSTER_MIN_TOKENS:
        return False
    if not (inter & _CLUSTER_PIVOTS):
        return False
    union = a | b
    return bool(union) and len(inter) / len(union) >= CLUSTER_SEUIL


def _nom_suggere(coeurs: list, designations: list) -> str:
    """Nom proposé pour une grappe : les pivots communs à tous les articles, sinon
    les mots communs, sinon la 1re désignation tronquée."""
    if not coeurs:
        return (designations[0] if designations else "Groupe").strip()
    communs = set(coeurs[0])
    for c in coeurs[1:]:
        communs &= c
    pivots = [t for t in communs if t in _CLUSTER_PIVOTS]
    mots = pivots or list(communs)
    if not mots:
        return designations[0].strip()[:60]
    # Ordre d'apparition dans la 1re désignation, pour un libellé naturel.
    ordre = [t for t in _coeur_ordre(designations[0]) if t in mots]
    ordre += [t for t in mots if t not in ordre]
    return " ".join(ordre).capitalize()


def _coeur_ordre(designation: str) -> list:
    """Comme _coeur_tokens mais en conservant l'ordre (pour nommer joliment)."""
    s = unicodedata.normalize("NFKD", str(designation or "").lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    toks = "".join(c if c.isalnum() else " " for c in s).split()
    return [t for t in toks
            if not any(ch.isdigit() for ch in t)
            and t not in _CLUSTER_UNITES and t not in _CLUSTER_PACK and len(t) >= 2]


@router.get("/comparatif/proposer")
async def proposer_groupes(_=Depends(require_admin)):
    """Balaie le catalogue et propose des grappes d'articles équivalents (même
    sous-famille, désignations proches, ≥ 2 fournisseurs) à valider d'un clic.
    Exclut les articles déjà rattachés à un groupe pour ne pas reproposer en boucle."""
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT c.*, f.nom AS fournisseur_nom
            FROM catalogue_fournisseur c
            JOIN fournisseurs f ON f.id = c.fournisseur_id
            WHERE f.boutique_id = 1 AND c.actif = 1
              AND c.id NOT IN (SELECT catalogue_fournisseur_id FROM comparatif_groupe_ligne)
            ORDER BY c.sous_famille, c.designation
            """
        )
        articles = [dict(r) for r in await cur.fetchall()]

    # Regroupement par sous-famille (les sans-sous-famille = bucket « moins fiable »).
    par_sf = {}
    for a in articles:
        a["_coeur"] = _coeur_tokens(a["designation"])
        sf = (a.get("sous_famille") or "").strip()
        par_sf.setdefault(sf, []).append(a)

    grappes = []
    mono_fournisseur = 0
    for sf, items in par_sf.items():
        pool = list(items)
        while pool:
            graine = pool.pop(0)
            grappe = [graine]
            reste = []
            for x in pool:
                if _coeurs_proches(graine["_coeur"], x["_coeur"]):
                    grappe.append(x)
                else:
                    reste.append(x)
            pool = reste
            fournisseurs = {g["fournisseur_id"] for g in grappe}
            if len(fournisseurs) < CLUSTER_MIN_FOURNISSEURS:
                if len(grappe) >= 1:
                    mono_fournisseur += 1
                continue
            lignes = []
            for g in grappe:
                lignes.append({
                    "id": g["id"],
                    "designation": g["designation"],
                    "fournisseur_nom": g["fournisseur_nom"],
                    "code_article": g["code_article"],
                    "prix_kg": _calc_prix_kg(g.get("format_prix"), g.get("prix_achat_ht"), g.get("poids_colis_kg"), g.get("famille")),
                })
            grappes.append({
                "nom_suggere": _nom_suggere([g["_coeur"] for g in grappe], [g["designation"] for g in grappe]),
                "sous_famille": sf,
                "fiable": bool(sf),  # sans sous-famille → moins fiable
                "nb_fournisseurs": len(fournisseurs),
                "lignes": lignes,
            })

    grappes.sort(key=lambda g: len(g["lignes"]), reverse=True)
    return {"grappes": grappes, "mono_fournisseur": mono_fournisseur}


@router.post("/comparatif/groupes/from-cluster", status_code=201)
async def create_groupe_from_cluster(body: ComparatifFromCluster, _=Depends(require_admin)):
    """Crée un groupe et y rattache les articles d'une grappe validée, en un appel."""
    nom = (body.nom or "").strip()
    if not nom:
        raise HTTPException(400, "Le nom du groupe est obligatoire")
    if not body.catalogue_fournisseur_ids:
        raise HTTPException(400, "Aucun article à rattacher")
    async with get_db() as db:
        cur = await db.execute(
            "INSERT INTO comparatif_groupe (boutique_id, nom) VALUES (1, ?)", (nom,)
        )
        groupe_id = cur.lastrowid
        for cat_id in body.catalogue_fournisseur_ids:
            await db.execute(
                "INSERT OR IGNORE INTO comparatif_groupe_ligne (groupe_id, catalogue_fournisseur_id) VALUES (?, ?)",
                (groupe_id, cat_id),
            )
        await db.commit()
        cur2 = await db.execute("SELECT * FROM comparatif_groupe WHERE id = ?", (groupe_id,))
        return dict(await cur2.fetchone())


# ===========================================================================
# FACTURES — rapprochement commande ↔ réception ↔ facture fournisseur
# ===========================================================================
#
# Principe : la facture ne modifie JAMAIS la réception (poids HACCP pesé à quai).
# Elle copie le poids reçu (figé), saisit le poids/prix FACTURÉ par le fournisseur,
# et croise la COMMANDE (prix négocié) pour révéler les écarts. Tout écart est
# calculé : facturé − reçu (poids), facturé − commande (prix), et sur le montant.


def _calc_ecarts_ligne(poids_recu, prix_commande, poids_facture, prix_facture):
    """Calcule (montant_facture, ecart_poids, ecart_prix, ecart_montant) pour une ligne.

    - montant facturé = poids facturé × prix facturé
    - écart poids     = poids facturé − poids reçu (positif = facturé en trop)
    - écart prix      = prix facturé − prix commande (positif = plus cher que négocié)
    - écart montant   = montant facturé − montant attendu (poids reçu × prix commande)
    Les valeurs manquantes sont traitées comme 0 pour ne pas casser le calcul.
    """
    pr = poids_recu or 0.0
    pc = prix_commande or 0.0
    pf = poids_facture or 0.0
    prix_f = prix_facture or 0.0

    montant_facture = pf * prix_f
    montant_attendu = pr * pc
    ecart_poids = pf - pr
    ecart_prix = prix_f - pc
    ecart_montant = montant_facture - montant_attendu
    return montant_facture, ecart_poids, ecart_prix, ecart_montant


async def _recalculer_ecarts_ligne(db, ligne_id: int):
    """Relit une ligne, recalcule montant + écarts, persiste."""
    cur = await db.execute("SELECT * FROM facture_lignes WHERE id = ?", (ligne_id,))
    ligne = await cur.fetchone()
    if not ligne:
        return
    montant, e_poids, e_prix, e_montant = _calc_ecarts_ligne(
        ligne["poids_recu_kg"], ligne["prix_commande_ht"],
        ligne["poids_facture_kg"], ligne["prix_facture_ht"],
    )
    await db.execute(
        """UPDATE facture_lignes
           SET montant_facture_ht = ?, ecart_poids_kg = ?, ecart_prix_ht = ?, ecart_montant_ht = ?
           WHERE id = ?""",
        (montant, e_poids, e_prix, e_montant, ligne_id),
    )


async def _recalculer_totaux_facture(db, facture_id: int):
    """Recalcule les totaux d'entête (facturé, attendu, écart) depuis les lignes."""
    cur = await db.execute(
        """SELECT
               COALESCE(SUM(montant_facture_ht), 0) AS total_facture,
               COALESCE(SUM(COALESCE(poids_recu_kg, 0) * COALESCE(prix_commande_ht, 0)), 0) AS total_attendu
           FROM facture_lignes WHERE facture_id = ?""",
        (facture_id,),
    )
    row = await cur.fetchone()
    total_facture = row["total_facture"]
    total_attendu = row["total_attendu"]
    await db.execute(
        """UPDATE factures
           SET montant_total_ht_facture = ?, montant_total_ht_attendu = ?, ecart_total_ht = ?
           WHERE id = ?""",
        (total_facture, total_attendu, total_facture - total_attendu, facture_id),
    )


@router.get("/factures/receptions-disponibles")
async def receptions_a_facturer(limit: int = Query(100)):
    """Réceptions clôturées pour la modale « Nouvelle facture ».

    Le nom du fournisseur est résolu en cascade : entête (FK ou texte libre), sinon
    le premier fournisseur trouvé sur les lignes (FK ou texte libre). Évite les
    « Fournisseur ? » quand l'info n'est portée que par les lignes. Chaque réception
    indique si elle est déjà facturée (pour griser le choix côté front).
    """
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT
                r.id,
                r.date_reception,
                COALESCE(
                    fh.nom,
                    r.fournisseur_nom,
                    (SELECT COALESCE(fl.nom, rl.fournisseur_nom)
                       FROM reception_lignes rl
                       LEFT JOIN fournisseurs fl ON fl.id = rl.fournisseur_id
                      WHERE rl.reception_id = r.id
                        AND COALESCE(fl.nom, rl.fournisseur_nom) IS NOT NULL
                      ORDER BY rl.id LIMIT 1)
                ) AS fournisseur_nom,
                (SELECT COUNT(*) FROM reception_lignes rl2 WHERE rl2.reception_id = r.id) AS nb_lignes,
                (SELECT fac.id FROM factures fac WHERE fac.reception_id = r.id LIMIT 1) AS facture_id
            FROM receptions r
            LEFT JOIN fournisseurs fh ON fh.id = r.fournisseur_principal_id
            WHERE r.statut = 'cloturee'
            ORDER BY r.created_at DESC
            LIMIT ?
            """,
            (limit,),
        )
        rows = [dict(r) for r in await cur.fetchall()]
        for row in rows:
            row["deja_facturee"] = row["facture_id"] is not None
        return rows


@router.get("/factures")
async def get_factures(
    fournisseur_id: Optional[int] = Query(None),
    statut: Optional[str] = Query(None),
    limit: int = Query(50),
):
    async with get_db() as db:
        sql = """
            SELECT fac.*, f.nom AS fournisseur_nom,
                   c.numero_commande AS numero_commande,
                   p.prenom AS personnel_prenom
            FROM factures fac
            JOIN fournisseurs f ON f.id = fac.fournisseur_id
            LEFT JOIN commandes c ON c.id = fac.commande_id
            LEFT JOIN personnel p ON p.id = fac.personnel_id
            WHERE fac.boutique_id = 1
        """
        params: list = []
        if fournisseur_id:
            sql += " AND fac.fournisseur_id = ?"
            params.append(fournisseur_id)
        if statut:
            sql += " AND fac.statut = ?"
            params.append(statut)
        sql += " ORDER BY fac.date_facture DESC, fac.id DESC LIMIT ?"
        params.append(limit)
        cur = await db.execute(sql, params)
        factures = [dict(r) for r in await cur.fetchall()]

        for fac in factures:
            cur2 = await db.execute(
                "SELECT COUNT(*) AS n, COALESCE(SUM(statut_ligne = 'litige'), 0) AS litiges "
                "FROM facture_lignes WHERE facture_id = ?",
                (fac["id"],),
            )
            r2 = await cur2.fetchone()
            fac["nb_lignes"] = r2["n"]
            fac["nb_litiges"] = r2["litiges"]

        return factures


@router.get("/factures/{facture_id}")
async def get_facture(facture_id: int):
    async with get_db() as db:
        cur = await db.execute(
            """SELECT fac.*, f.nom AS fournisseur_nom, f.email_commercial, f.telephone, f.adresse,
                      c.numero_commande AS numero_commande, c.date_commande AS date_commande,
                      p.prenom AS personnel_prenom
               FROM factures fac
               JOIN fournisseurs f ON f.id = fac.fournisseur_id
               LEFT JOIN commandes c ON c.id = fac.commande_id
               LEFT JOIN personnel p ON p.id = fac.personnel_id
               WHERE fac.id = ?""",
            (facture_id,),
        )
        facture = await cur.fetchone()
        if not facture:
            raise HTTPException(404, "Facture introuvable")
        result = dict(facture)

        cur2 = await db.execute(
            "SELECT * FROM facture_lignes WHERE facture_id = ? ORDER BY id", (facture_id,)
        )
        result["lignes"] = [dict(r) for r in await cur2.fetchall()]
        return result


async def _generer_facture_depuis_reception(
    db, reception_id: int, personnel_id: Optional[int] = None,
    prix_source: str = "bl", auto_litige_seuil_pct: Optional[float] = None,
) -> dict:
    """Crée une facture brouillon pré-remplie depuis une réception (logique partagée).

    Utilisée par l'endpoint manuel ET par le hook de clôture de réception. Ne lève
    aucune HTTPException : renvoie un statut structuré que l'appelant interprète.

    Args:
        prix_source : 'bl' → prix facturé pré-rempli avec le prix lu sur le BL (réalité
            constatée à la réception), à défaut le prix commande ; 'commande' → prix
            facturé = prix négocié (comportement d'origine, le gérant ajustera).
        auto_litige_seuil_pct : si fourni, une ligne dont l'écart prix relatif (vs prix
            commande) dépasse ce seuil est marquée 'litige' avec un commentaire.

    Returns:
        {"ok": bool, "facture_id": int|None, "raison": str|None}
        raison ∈ {"introuvable", "deja_facturee", "sans_fournisseur"} quand ok=False.
    """
    cur = await db.execute("SELECT * FROM receptions WHERE id = ?", (reception_id,))
    reception = await cur.fetchone()
    if not reception:
        return {"ok": False, "facture_id": None, "raison": "introuvable"}

    # Une réception = une facture (idempotent : le hook ne doit jamais doubler).
    cur_dup = await db.execute(
        "SELECT id FROM factures WHERE reception_id = ?", (reception_id,)
    )
    existante = await cur_dup.fetchone()
    if existante:
        return {"ok": False, "facture_id": existante["id"], "raison": "deja_facturee"}

    # Commande mappée (s'il y en a une)
    cur_map = await db.execute(
        """SELECT commande_id FROM commande_receptions_mapping
           WHERE reception_id = ? ORDER BY date_liaison DESC LIMIT 1""",
        (reception_id,),
    )
    map_row = await cur_map.fetchone()
    commande_id = map_row["commande_id"] if map_row else None

    # Fournisseur : priorité au fournisseur de la commande, sinon réception
    fournisseur_id = reception["fournisseur_principal_id"]
    if commande_id:
        cur_c = await db.execute(
            "SELECT fournisseur_id FROM commandes WHERE id = ?", (commande_id,)
        )
        c_row = await cur_c.fetchone()
        if c_row:
            fournisseur_id = c_row["fournisseur_id"]
    if not fournisseur_id:
        return {"ok": False, "facture_id": None, "raison": "sans_fournisseur"}

    # Lignes de réception (poids HACCP figé + prix BL indicatif) + désignation via COALESCE
    cur_rl = await db.execute(
        """SELECT rl.id AS reception_ligne_id, rl.catalogue_fournisseur_id,
                  rl.poids_kg AS poids_recu_kg, rl.prix_unitaire_ht AS prix_bl_ht,
                  COALESCE(p.nom, cf.designation, rl.designation_libre) AS designation,
                  cf.code_article AS code_article
           FROM reception_lignes rl
           LEFT JOIN produits p ON p.id = rl.produit_id
           LEFT JOIN catalogue_fournisseur cf ON cf.id = rl.catalogue_fournisseur_id
           WHERE rl.reception_id = ?
           ORDER BY rl.id""",
        (reception_id,),
    )
    lignes_reception = [dict(r) for r in await cur_rl.fetchall()]

    # Lignes de commande (prix négocié) pour le matching
    lignes_commande = []
    if commande_id:
        cur_cl = await db.execute(
            "SELECT * FROM commande_lignes WHERE commande_id = ?", (commande_id,)
        )
        lignes_commande = [dict(r) for r in await cur_cl.fetchall()]

    def _trouver_commande(rl):
        """Retrouve la ligne de commande (catalogue d'abord, puis désignation)."""
        if rl.get("catalogue_fournisseur_id"):
            for cl in lignes_commande:
                if cl["catalogue_fournisseur_id"] == rl["catalogue_fournisseur_id"]:
                    return cl
        desig = (rl.get("designation") or "").strip().lower()
        for cl in lignes_commande:
            if (cl["designation"] or "").strip().lower() == desig:
                return cl
        return None

    # Création de l'entête
    date_fac = date.today().isoformat()
    cur_ins = await db.execute(
        """INSERT INTO factures (boutique_id, fournisseur_id, reception_id, commande_id,
                                 date_facture, statut, personnel_id)
           VALUES (1, ?, ?, ?, ?, 'brouillon', ?)""",
        (fournisseur_id, reception_id, commande_id, date_fac, personnel_id),
    )
    facture_id = cur_ins.lastrowid

    a_un_litige = False
    for rl in lignes_reception:
        cl = _trouver_commande(rl)
        prix_commande = cl["prix_unitaire_ht"] if cl else None
        quantite_commandee = cl["quantite_commandee"] if cl else None
        unite = (cl["unite"] if cl else None) or "kg"
        poids_facture = rl["poids_recu_kg"]
        # Prix facturé pré-rempli : prix BL (réalité constatée) si dispo, sinon prix commande.
        if prix_source == "bl":
            prix_facture = rl.get("prix_bl_ht") if rl.get("prix_bl_ht") is not None else prix_commande
        else:
            prix_facture = prix_commande
        montant, e_poids, e_prix, e_montant = _calc_ecarts_ligne(
            rl["poids_recu_kg"], prix_commande, poids_facture, prix_facture
        )

        # Litige auto : écart prix relatif (vs commande) au-delà du seuil.
        statut_ligne = "ok"
        commentaire_litige = None
        if (auto_litige_seuil_pct is not None and prix_commande and prix_commande > 0
                and prix_facture is not None):
            ecart_pct = abs(prix_facture - prix_commande) / prix_commande * 100
            if ecart_pct > auto_litige_seuil_pct:
                statut_ligne = "litige"
                commentaire_litige = (
                    f"Écart prix automatique : commande {prix_commande:.2f} € / "
                    f"BL {prix_facture:.2f} € ({ecart_pct:+.1f} %)"
                )
                a_un_litige = True

        await db.execute(
            """INSERT INTO facture_lignes
               (facture_id, catalogue_fournisseur_id, reception_ligne_id, code_article,
                designation, unite, poids_recu_kg, prix_commande_ht, quantite_commandee,
                poids_facture_kg, prix_facture_ht, montant_facture_ht,
                ecart_poids_kg, ecart_prix_ht, ecart_montant_ht,
                statut_ligne, commentaire_litige)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (facture_id, rl["catalogue_fournisseur_id"], rl["reception_ligne_id"],
             rl["code_article"], rl["designation"] or "Article", unite,
             rl["poids_recu_kg"], prix_commande, quantite_commandee,
             poids_facture, prix_facture, montant, e_poids, e_prix, e_montant,
             statut_ligne, commentaire_litige),
        )

    # Si au moins une ligne est en litige, l'entête passe en 'litige' (sinon brouillon).
    if a_un_litige:
        await db.execute(
            "UPDATE factures SET statut = 'litige' WHERE id = ?", (facture_id,)
        )

    await _recalculer_totaux_facture(db, facture_id)
    await db.commit()
    return {"ok": True, "facture_id": facture_id, "raison": None}


@router.post("/factures/depuis-reception/{reception_id}", status_code=201)
async def creer_facture_depuis_reception(reception_id: int, personnel_id: Optional[int] = None):
    """Crée une facture pré-remplie à partir d'une réception (déclenchement manuel).

    Pour chaque ligne de réception : copie le poids reçu (figé) + la désignation, puis
    va chercher dans la commande mappée le prix négocié. Le prix facturé est pré-rempli
    avec le prix lu sur le BL (à défaut, le prix commande). Le gérant ajuste ensuite.
    """
    async with get_db() as db:
        res = await _generer_facture_depuis_reception(
            db, reception_id, personnel_id, prix_source="bl",
        )
    if not res["ok"]:
        if res["raison"] == "introuvable":
            raise HTTPException(404, "Réception introuvable")
        if res["raison"] == "deja_facturee":
            raise HTTPException(
                409, f"Une facture existe déjà pour cette réception (id={res['facture_id']})"
            )
        if res["raison"] == "sans_fournisseur":
            raise HTTPException(
                400,
                "Aucun fournisseur identifié pour cette réception (commande ou fournisseur principal requis)",
            )
        raise HTTPException(400, "Création de facture impossible")
    return await get_facture(res["facture_id"])


# ───────────────────────────────────────────────────────────────────────────
#  Historisation des prix d'achat + mise à jour semi-auto du catalogue
# ───────────────────────────────────────────────────────────────────────────

async def _historiser_prix_reception(db, reception_id: int, source: str = "bl") -> int:
    """Enregistre dans historique_prix_achat le prix constaté de chaque ligne de la
    réception ayant un prix + un article catalogue. N'écrase JAMAIS le prix de
    référence : se contente d'ajouter une trace (applique_au_catalogue=0).

    Le prix est normalisé en €/kg via `_calc_prix_kg`, en présumant que l'unité du
    prix réception est celle du catalogue (format_prix) — hypothèse correcte pour la
    viande (où _calc_prix_kg prend le prix tel quel) et cohérente sinon, faute d'unité
    stockée sur la ligne de réception. On mémorise aussi le €/kg de référence courant
    (prix_kg_precedent) pour afficher l'évolution.

    Returns: nombre de lignes historisées.
    """
    cur = await db.execute(
        """SELECT rl.id AS reception_ligne_id, rl.prix_unitaire_ht AS prix_bl,
                  rl.catalogue_fournisseur_id AS cat_id,
                  cf.format_prix, cf.prix_achat_ht AS prix_ref, cf.poids_colis_kg, cf.famille,
                  r.date_reception AS date_reception
           FROM reception_lignes rl
           JOIN catalogue_fournisseur cf ON cf.id = rl.catalogue_fournisseur_id
           JOIN receptions r ON r.id = rl.reception_id
           WHERE rl.reception_id = ? AND rl.prix_unitaire_ht IS NOT NULL""",
        (reception_id,),
    )
    lignes = [dict(r) for r in await cur.fetchall()]
    if not lignes:
        return 0

    # La date du constat = date de réception (quand la marchandise et son prix sont
    # arrivés), pas la date de clôture — plus juste pour la courbe en cas de saisie différée.
    n = 0
    for l in lignes:
        prix_kg = _calc_prix_kg(l["format_prix"], l["prix_bl"], l["poids_colis_kg"], l["famille"])
        prix_kg_ref = _calc_prix_kg(l["format_prix"], l["prix_ref"], l["poids_colis_kg"], l["famille"])
        date_constat = l.get("date_reception") or date.today().isoformat()
        await db.execute(
            """INSERT INTO historique_prix_achat
                   (catalogue_fournisseur_id, reception_id, reception_ligne_id,
                    prix_ht, format_prix, prix_kg, prix_kg_precedent, source,
                    applique_au_catalogue, date_constat)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?)""",
            (l["cat_id"], reception_id, l["reception_ligne_id"], l["prix_bl"],
             l["format_prix"], prix_kg, prix_kg_ref, source, date_constat),
        )
        n += 1
    await db.commit()
    return n


@router.get("/catalogue/ecarts-prix/{reception_id}")
async def ecarts_prix_reception(reception_id: int, seuil_pct: float = Query(2.0)):
    """Liste les articles de la réception dont le prix constaté s'écarte du prix de
    référence catalogue au-delà du seuil — pour le bandeau « mettre à jour le catalogue ? ».

    S'appuie sur les lignes historisées de cette réception (la plus récente par article)
    comparées au prix de référence catalogue ACTUEL. Ne renvoie que les écarts significatifs
    et encore non appliqués.
    """
    async with get_db() as db:
        cur = await db.execute(
            """
            SELECT h.catalogue_fournisseur_id AS cat_id,
                   cf.designation, cf.code_article, cf.format_prix,
                   cf.prix_achat_ht AS prix_ref_actuel,
                   h.prix_ht AS prix_constate, h.prix_kg, h.prix_kg_precedent,
                   h.id AS historique_id, h.applique_au_catalogue
            FROM historique_prix_achat h
            JOIN catalogue_fournisseur cf ON cf.id = h.catalogue_fournisseur_id
            JOIN (
                SELECT catalogue_fournisseur_id, MAX(id) AS max_id
                FROM historique_prix_achat
                WHERE reception_id = ?
                GROUP BY catalogue_fournisseur_id
            ) last ON last.max_id = h.id
            WHERE h.reception_id = ?
            ORDER BY cf.designation
            """,
            (reception_id, reception_id),
        )
        rows = [dict(r) for r in await cur.fetchall()]

    ecarts = []
    for r in rows:
        # Écart €/kg : prix constaté vs prix de référence au moment du constat
        # (prix_kg_precedent), tous deux déjà normalisés à l'historisation.
        prix_kg = r["prix_kg"]
        ref_kg = r["prix_kg_precedent"]
        if prix_kg is None or ref_kg is None or ref_kg <= 0:
            continue
        ecart_pct = (prix_kg - ref_kg) / ref_kg * 100
        if abs(ecart_pct) <= seuil_pct:
            continue
        # Si le prix de référence a déjà été aligné depuis (égal au constaté), plus d'écart.
        if r["applique_au_catalogue"]:
            continue
        ecarts.append({
            "catalogue_fournisseur_id": r["cat_id"],
            "designation": r["designation"],
            "code_article": r["code_article"],
            "prix_ref_actuel": r["prix_ref_actuel"],
            "prix_constate": r["prix_constate"],
            "prix_kg_constate": round(prix_kg, 4),
            "prix_kg_reference": round(ref_kg, 4),
            "ecart_pct": round(ecart_pct, 2),
            "deja_applique": bool(r["applique_au_catalogue"]),
        })
    return {"reception_id": reception_id, "seuil_pct": seuil_pct, "ecarts": ecarts}


class AppliquerPrixBody(BaseModel):
    nouveau_prix_ht: float                       # nouveau prix de référence (même format que le catalogue)
    reception_id: Optional[int] = None           # pour marquer la ligne d'historique correspondante


@router.post("/catalogue/{catalogue_id}/appliquer-prix")
async def appliquer_prix_catalogue(catalogue_id: int, body: AppliquerPrixBody):
    """Met à jour le PRIX DE RÉFÉRENCE d'un article catalogue (décision explicite de
    l'utilisateur depuis le bandeau d'écart). Trace la décision dans l'historique
    (applique_au_catalogue=1) pour la dernière observation de cette réception.
    """
    if body.nouveau_prix_ht < 0:
        raise HTTPException(400, "Prix négatif invalide")
    async with get_db() as db:
        cur = await db.execute(
            "SELECT id, prix_achat_ht FROM catalogue_fournisseur WHERE id = ?", (catalogue_id,)
        )
        art = await cur.fetchone()
        if not art:
            raise HTTPException(404, "Article catalogue introuvable")
        ancien = art["prix_achat_ht"]

        await db.execute(
            "UPDATE catalogue_fournisseur SET prix_achat_ht = ?, date_maj = CURRENT_TIMESTAMP WHERE id = ?",
            (body.nouveau_prix_ht, catalogue_id),
        )
        # Marquer la décision dans l'historique (dernière obs. de la réception, sinon la
        # plus récente pour cet article).
        if body.reception_id is not None:
            await db.execute(
                """UPDATE historique_prix_achat SET applique_au_catalogue = 1
                   WHERE id = (SELECT MAX(id) FROM historique_prix_achat
                               WHERE catalogue_fournisseur_id = ? AND reception_id = ?)""",
                (catalogue_id, body.reception_id),
            )
        else:
            await db.execute(
                """UPDATE historique_prix_achat SET applique_au_catalogue = 1
                   WHERE id = (SELECT MAX(id) FROM historique_prix_achat
                               WHERE catalogue_fournisseur_id = ?)""",
                (catalogue_id,),
            )
        await db.commit()
    return {"ok": True, "catalogue_fournisseur_id": catalogue_id,
            "ancien_prix_ht": ancien, "nouveau_prix_ht": body.nouveau_prix_ht}


@router.get("/catalogue/{catalogue_id}/historique-prix")
async def historique_prix_catalogue(catalogue_id: int, limit: int = Query(60, ge=1, le=500)):
    """Courbe d'évolution du prix d'achat (€/kg) d'un article, depuis historique_prix_achat.

    Renvoie les points constatés (à chaque réception), triés du plus ancien au plus récent,
    plus le prix de référence actuel du catalogue pour tracer la ligne de comparaison.
    """
    async with get_db() as db:
        cur = await db.execute(
            "SELECT designation, prix_achat_ht, format_prix, poids_colis_kg, famille "
            "FROM catalogue_fournisseur WHERE id = ?",
            (catalogue_id,),
        )
        art = await cur.fetchone()
        if not art:
            raise HTTPException(404, "Article introuvable")

        cur2 = await db.execute(
            """SELECT date_constat, prix_kg, prix_ht, source, applique_au_catalogue
               FROM historique_prix_achat
               WHERE catalogue_fournisseur_id = ?
               ORDER BY date_constat ASC, id ASC
               LIMIT ?""",
            (catalogue_id, limit),
        )
        points = [dict(r) for r in await cur2.fetchall()]

    prix_ref_kg = _calc_prix_kg(
        art["format_prix"], art["prix_achat_ht"], art["poids_colis_kg"], art["famille"]
    )
    return {
        "catalogue_fournisseur_id": catalogue_id,
        "designation": art["designation"],
        "prix_reference_kg": prix_ref_kg,
        "points": points,
    }


@router.post("/factures", status_code=201)
async def create_facture(body: FactureCreate):
    """Création manuelle d'une facture (entête + lignes optionnelles)."""
    async with get_db() as db:
        cur_f = await db.execute(
            "SELECT id FROM fournisseurs WHERE id = ? AND boutique_id = 1", (body.fournisseur_id,)
        )
        if not await cur_f.fetchone():
            raise HTTPException(404, "Fournisseur introuvable")

        date_fac = body.date_facture or date.today().isoformat()
        cur = await db.execute(
            """INSERT INTO factures (boutique_id, fournisseur_id, reception_id, commande_id,
                                     numero_facture, date_facture, statut, personnel_id, commentaire)
               VALUES (1, ?, ?, ?, ?, ?, 'brouillon', ?, ?)""",
            (body.fournisseur_id, body.reception_id, body.commande_id, body.numero_facture,
             date_fac, body.personnel_id, body.commentaire),
        )
        facture_id = cur.lastrowid

        for ligne in (body.lignes or []):
            montant, e_poids, e_prix, e_montant = _calc_ecarts_ligne(
                ligne.poids_recu_kg, ligne.prix_commande_ht,
                ligne.poids_facture_kg, ligne.prix_facture_ht,
            )
            await db.execute(
                """INSERT INTO facture_lignes
                   (facture_id, catalogue_fournisseur_id, reception_ligne_id, code_article,
                    designation, unite, poids_recu_kg, prix_commande_ht, quantite_commandee,
                    poids_facture_kg, prix_facture_ht, montant_facture_ht,
                    ecart_poids_kg, ecart_prix_ht, ecart_montant_ht)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (facture_id, ligne.catalogue_fournisseur_id, ligne.reception_ligne_id,
                 ligne.code_article, ligne.designation, ligne.unite or "kg",
                 ligne.poids_recu_kg, ligne.prix_commande_ht, ligne.quantite_commandee,
                 ligne.poids_facture_kg, ligne.prix_facture_ht, montant,
                 e_poids, e_prix, e_montant),
            )

        await _recalculer_totaux_facture(db, facture_id)
        await db.commit()
        return await get_facture(facture_id)


@router.put("/factures/{facture_id}")
async def update_facture(facture_id: int, body: FactureUpdate):
    async with get_db() as db:
        cur = await db.execute("SELECT id FROM factures WHERE id = ?", (facture_id,))
        if not await cur.fetchone():
            raise HTTPException(404, "Facture introuvable")

        fields = {k: v for k, v in body.model_dump().items() if v is not None}
        if not fields:
            raise HTTPException(400, "Aucun champ à modifier")

        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [facture_id]
        await db.execute(f"UPDATE factures SET {set_clause} WHERE id = ?", values)
        await db.commit()
        return await get_facture(facture_id)


@router.put("/factures/{facture_id}/lignes/{ligne_id}")
async def update_facture_ligne(facture_id: int, ligne_id: int, body: FactureLigneUpdate):
    async with get_db() as db:
        cur = await db.execute(
            "SELECT id FROM facture_lignes WHERE id = ? AND facture_id = ?", (ligne_id, facture_id)
        )
        if not await cur.fetchone():
            raise HTTPException(404, "Ligne introuvable")

        fields = {k: v for k, v in body.model_dump().items() if v is not None}
        if not fields:
            raise HTTPException(400, "Aucun champ à modifier")

        set_clause = ", ".join(f"{k} = ?" for k in fields)
        values = list(fields.values()) + [ligne_id]
        await db.execute(f"UPDATE facture_lignes SET {set_clause} WHERE id = ?", values)
        # Recalcule écarts de la ligne + totaux d'entête
        await _recalculer_ecarts_ligne(db, ligne_id)
        await _recalculer_totaux_facture(db, facture_id)
        await db.commit()

        cur2 = await db.execute("SELECT * FROM facture_lignes WHERE id = ?", (ligne_id,))
        return dict(await cur2.fetchone())


@router.delete("/factures/{facture_id}/lignes/{ligne_id}", status_code=204)
async def delete_facture_ligne(facture_id: int, ligne_id: int):
    async with get_db() as db:
        await db.execute(
            "DELETE FROM facture_lignes WHERE id = ? AND facture_id = ?", (ligne_id, facture_id)
        )
        await _recalculer_totaux_facture(db, facture_id)
        await db.commit()


@router.delete("/factures/{facture_id}", status_code=204)
async def delete_facture(facture_id: int):
    async with get_db() as db:
        cur = await db.execute("SELECT id FROM factures WHERE id = ?", (facture_id,))
        if not await cur.fetchone():
            raise HTTPException(404, "Facture introuvable")
        await db.execute("DELETE FROM facture_lignes WHERE facture_id = ?", (facture_id,))
        await db.execute("DELETE FROM factures WHERE id = ?", (facture_id,))
        await db.commit()


# ── Exports facture (PDF imprimable + Excel) ─────────────────────────────────
# Le récap contient TOUTES les lignes du rapprochement (pas seulement les litiges),
# avec les écarts mis en évidence. Sert de document de contrôle / demande d'avoir.

async def _charger_facture_complete(facture_id: int) -> dict:
    """Entête enrichie + lignes d'une facture (factorisé pour les exports)."""
    async with get_db() as db:
        cur = await db.execute(
            """SELECT fac.*, f.nom AS fournisseur_nom, f.email_commercial, f.telephone, f.adresse,
                      c.numero_commande AS numero_commande, c.date_commande AS date_commande
               FROM factures fac
               JOIN fournisseurs f ON f.id = fac.fournisseur_id
               LEFT JOIN commandes c ON c.id = fac.commande_id
               WHERE fac.id = ?""",
            (facture_id,),
        )
        facture = await cur.fetchone()
        if not facture:
            raise HTTPException(404, "Facture introuvable")
        result = dict(facture)
        cur2 = await db.execute(
            "SELECT * FROM facture_lignes WHERE facture_id = ? ORDER BY id", (facture_id,)
        )
        result["lignes"] = [dict(r) for r in await cur2.fetchall()]
        return result


def _fmt_eur(v) -> str:
    return f"{(v or 0):.2f} €".replace(".", ",")


def _fmt_kg(v) -> str:
    return "—" if v is None else f"{v:.3f}".replace(".", ",")


def _fmt_signe(v) -> str:
    """Écart signé avec le tiret typographique pour le négatif (cohérent avec le front)."""
    v = v or 0
    if v > 0.0001:
        return "+" + _fmt_eur(v)
    if v < -0.0001:
        return "−" + _fmt_eur(abs(v))
    return _fmt_eur(0)


@router.get("/factures/{facture_id}/export.xlsx")
async def export_facture_xlsx(facture_id: int):
    """Export Excel du rapprochement complet d'une facture."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl requis")

    fac = await _charger_facture_complete(facture_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapprochement facture"

    # Bandeau entête (méta facture)
    meta = [
        ("Fournisseur", fac.get("fournisseur_nom") or ""),
        ("N° facture", fac.get("numero_facture") or "(non saisi)"),
        ("Date facture", fac.get("date_facture") or ""),
        ("Commande", fac.get("numero_commande") or "—"),
        ("Statut", fac.get("statut") or ""),
    ]
    for i, (label, val) in enumerate(meta, 1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val)

    start = len(meta) + 2  # ligne d'en-tête du tableau

    cols = [
        ("Article", 34),
        ("Reçu (kg)", 14),
        ("Facturé (kg)", 14),
        ("Écart poids (kg)", 16),
        ("Prix cmd HT", 14),
        ("Prix fact. HT", 14),
        ("Montant facturé HT", 18),
        ("Écart HT", 14),
        ("Litige", 10),
        ("Motif litige", 30),
    ]
    header_fill = PatternFill("solid", fgColor="6B2D0F")
    thin = Side(style="thin", color="D4C5AF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=start, column=ci, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width

    rouge = Font(color="B91C1C", bold=True)
    vert = Font(color="166534", bold=True)
    for ri, l in enumerate(fac["lignes"], start + 1):
        ecart_montant = l.get("ecart_montant_ht") or 0
        ecart_poids = l.get("ecart_poids_kg") or 0
        litige = l.get("statut_ligne") == "litige"
        valeurs = [
            l.get("designation") or "",
            _fmt_kg(l.get("poids_recu_kg")),
            _fmt_kg(l.get("poids_facture_kg")),
            _fmt_kg(ecart_poids) if abs(ecart_poids) > 1e-9 else "0",
            _fmt_eur(l.get("prix_commande_ht")) if l.get("prix_commande_ht") is not None else "—",
            _fmt_eur(l.get("prix_facture_ht")) if l.get("prix_facture_ht") is not None else "—",
            _fmt_eur(l.get("montant_facture_ht")),
            _fmt_signe(ecart_montant),
            "OUI" if litige else "",
            l.get("commentaire_litige") or "",
        ]
        for ci, val in enumerate(valeurs, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            if ci in (2, 3, 4, 5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="right")
        # Couleur de l'écart HT (col 8)
        if abs(ecart_montant) > 1e-9:
            ws.cell(row=ri, column=8).font = rouge if ecart_montant > 0 else vert

    # Ligne de totaux
    tot = start + 1 + len(fac["lignes"])
    ws.cell(row=tot, column=1, value="TOTAUX").font = Font(bold=True)
    ws.cell(row=tot, column=7, value=_fmt_eur(fac.get("montant_total_ht_facture"))).font = Font(bold=True)
    cell_ecart = ws.cell(row=tot, column=8, value=_fmt_signe(fac.get("ecart_total_ht")))
    et = fac.get("ecart_total_ht") or 0
    cell_ecart.font = rouge if et > 1e-9 else (vert if et < -1e-9 else Font(bold=True))
    ws.cell(row=tot + 1, column=1, value="Attendu (cmd × reçu) HT")
    ws.cell(row=tot + 1, column=7, value=_fmt_eur(fac.get("montant_total_ht_attendu")))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    num = (fac.get("numero_facture") or f"facture-{facture_id}").replace("/", "-").replace(" ", "_")
    fname = f"rapprochement_{num}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/factures/{facture_id}/imprimer", response_class=HTMLResponse)
async def imprimer_facture(facture_id: int):
    """Page HTML imprimable du rapprochement (le navigateur fait « Enregistrer en PDF »).

    Choisi plutôt qu'un PDF serveur : WeasyPrint nécessite des libs natives absentes
    en dev Windows et fragiles sur le Pi ; l'impression navigateur donne un PDF parfait
    sans dépendance serveur.
    """
    import html as _html

    fac = await _charger_facture_complete(facture_id)
    esc = _html.escape

    def cell_ecart(v):
        v = v or 0
        if v > 0.0001:
            return f'<td class="num ecart-haut">{_html.escape(_fmt_signe(v))}</td>'
        if v < -0.0001:
            return f'<td class="num ecart-bas">{_html.escape(_fmt_signe(v))}</td>'
        return f'<td class="num ecart-nul">{_html.escape(_fmt_eur(0))}</td>'

    lignes_html = ""
    for l in fac["lignes"]:
        litige = l.get("statut_ligne") == "litige"
        ecart_poids = l.get("ecart_poids_kg") or 0
        lignes_html += f"""
        <tr class="{'litige' if litige else ''}">
          <td>{esc(l.get('designation') or '')}{f'<div class="sub">{esc(l.get("code_article"))}</div>' if l.get('code_article') else ''}</td>
          <td class="num">{esc(_fmt_kg(l.get('poids_recu_kg')))}</td>
          <td class="num">{esc(_fmt_kg(l.get('poids_facture_kg')))}</td>
          <td class="num">{esc(_fmt_kg(ecart_poids) if abs(ecart_poids) > 1e-9 else '0')}</td>
          <td class="num">{esc(_fmt_eur(l.get('prix_commande_ht')) if l.get('prix_commande_ht') is not None else '—')}</td>
          <td class="num">{esc(_fmt_eur(l.get('prix_facture_ht')) if l.get('prix_facture_ht') is not None else '—')}</td>
          <td class="num">{esc(_fmt_eur(l.get('montant_facture_ht')))}</td>
          {cell_ecart(l.get('ecart_montant_ht'))}
          <td>{'⚠ ' + esc(l.get('commentaire_litige') or 'Litige') if litige else ''}</td>
        </tr>"""

    et = fac.get("ecart_total_ht") or 0
    classe_tot = "ecart-haut" if et > 1e-9 else ("ecart-bas" if et < -1e-9 else "ecart-nul")

    html = f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="UTF-8">
<title>Rapprochement facture {esc(fac.get('numero_facture') or '')}</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; color: #1f2937; margin: 24px; font-size: 13px; }}
  h1 {{ font-size: 20px; margin: 0 0 4px; color: #6b2d0f; }}
  .meta {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 4px 24px; margin: 12px 0 18px;
           padding: 12px 16px; background: #f9f5ef; border: 1px solid #e8d9c4; border-radius: 6px; max-width: 640px; }}
  .meta b {{ color: #6b7280; font-weight: 600; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th, td {{ border: 1px solid #d4c5af; padding: 6px 8px; text-align: left; }}
  th {{ background: #6b2d0f; color: #fff; font-size: 12px; }}
  td.num, th.num {{ text-align: right; white-space: nowrap; }}
  tr.litige {{ background: #fef2f2; }}
  .sub {{ font-size: 11px; color: #9ca3af; }}
  .ecart-haut {{ color: #b91c1c; font-weight: 700; }}
  .ecart-bas  {{ color: #166534; font-weight: 700; }}
  .ecart-nul  {{ color: #6b7280; }}
  tfoot td {{ font-weight: 700; background: #f3ebdf; }}
  .imprimer {{ margin: 18px 0; }}
  .imprimer button {{ padding: 8px 18px; font-size: 14px; cursor: pointer; border: none;
                      background: #6b2d0f; color: #fff; border-radius: 6px; }}
  @media print {{ .imprimer {{ display: none; }} body {{ margin: 0; }} }}
</style></head><body>
  <h1>🧾 Rapprochement de facture</h1>
  <div class="meta">
    <div><b>Fournisseur :</b> {esc(fac.get('fournisseur_nom') or '')}</div>
    <div><b>N° facture :</b> {esc(fac.get('numero_facture') or '(non saisi)')}</div>
    <div><b>Date facture :</b> {esc(fac.get('date_facture') or '')}</div>
    <div><b>Commande :</b> {esc(fac.get('numero_commande') or '—')}</div>
    <div><b>Statut :</b> {esc(fac.get('statut') or '')}</div>
  </div>
  <div class="imprimer"><button onclick="window.print()">🖨 Imprimer / Enregistrer en PDF</button></div>
  <table>
    <thead><tr>
      <th>Article</th><th class="num">Reçu (kg)</th><th class="num">Facturé (kg)</th>
      <th class="num">Écart poids</th><th class="num">Prix cmd HT</th><th class="num">Prix fact. HT</th>
      <th class="num">Montant HT</th><th class="num">Écart HT</th><th>Litige</th>
    </tr></thead>
    <tbody>{lignes_html}</tbody>
    <tfoot>
      <tr>
        <td colspan="6">Attendu (cmd × reçu) : {esc(_fmt_eur(fac.get('montant_total_ht_attendu')))}</td>
        <td class="num">{esc(_fmt_eur(fac.get('montant_total_ht_facture')))}</td>
        <td class="num {classe_tot}">{esc(_fmt_signe(et))}</td>
        <td></td>
      </tr>
    </tfoot>
  </table>
</body></html>"""
    return HTMLResponse(content=html)


# ===========================================================================
# PILOTAGE — Chiffre d'affaires journalier
# ===========================================================================

def _panier(ttc, nb):
    """Panier moyen = CA / tickets, ou None si tickets manquant/nul."""
    if nb and ttc is not None:
        return round(float(ttc) / int(nb), 2)
    return None


def _enrichir_ca(row) -> dict:
    """Ajoute les paniers moyens calculés (global + par section)."""
    d = dict(row)
    d["panier_moyen"]       = _panier(d.get("montant_ttc"),       d.get("nb_tickets"))
    d["panier_moyen_matin"] = _panier(d.get("montant_ttc_matin"), d.get("nb_tickets_matin"))
    d["panier_moyen_soir"]  = _panier(d.get("montant_ttc_soir"),  d.get("nb_tickets_soir"))
    return d


async def _agg_periode(db, debut: str, fin: str) -> dict:
    """Agrège le CA d'une période [debut, fin] (dates ISO incluses).

    Renvoie totaux, moyennes, paniers (global/matin/soir) et répartition %.
    """
    cur = await db.execute(
        """SELECT
               COUNT(*)                             AS nb_jours,
               COALESCE(SUM(montant_ttc), 0)        AS total_ttc,
               COALESCE(SUM(nb_tickets), 0)         AS total_tickets,
               COALESCE(SUM(montant_ttc_matin), 0)  AS total_ttc_matin,
               COALESCE(SUM(nb_tickets_matin), 0)   AS total_tickets_matin,
               COALESCE(SUM(montant_ttc_soir), 0)   AS total_ttc_soir,
               COALESCE(SUM(nb_tickets_soir), 0)    AS total_tickets_soir
           FROM ca_journalier
           WHERE boutique_id = 1 AND date_ca >= ? AND date_ca <= ?""",
        (debut, fin),
    )
    r = dict(await cur.fetchone())
    nbj = r["nb_jours"] or 0
    r["date_debut"]          = debut
    r["date_fin"]            = fin
    r["ca_moyen_jour"]       = round(r["total_ttc"] / nbj, 2) if nbj else None
    r["panier_moyen"]        = _panier(r["total_ttc"],       r["total_tickets"])
    r["panier_moyen_matin"]  = _panier(r["total_ttc_matin"], r["total_tickets_matin"])
    r["panier_moyen_soir"]   = _panier(r["total_ttc_soir"],  r["total_tickets_soir"])
    tot = r["total_ttc"] or 0
    if tot > 0:
        r["part_matin"] = round(100 * r["total_ttc_matin"] / tot, 1)
        r["part_soir"]  = round(100 * r["total_ttc_soir"]  / tot, 1)
    else:
        r["part_matin"] = None
        r["part_soir"]  = None
    return r


def _evolution(courant: float, reference: float) -> dict:
    """Écart absolu et relatif (%) entre période courante et référence.

    pct = None si la référence est nulle (évolution non calculable).
    """
    delta = round((courant or 0) - (reference or 0), 2)
    pct = round(100 * delta / reference, 1) if reference else None
    return {"delta": delta, "pct": pct}


@router.get("/pilotage/ca")
async def get_ca_historique(
    date_debut: Optional[str] = Query(None, description="YYYY-MM-DD (inclus)"),
    date_fin:   Optional[str] = Query(None, description="YYYY-MM-DD (inclus)"),
    limit:      int           = Query(90, ge=1, le=1000),
):
    """Historique du CA, du plus récent au plus ancien (filtre période optionnel)."""
    clauses = ["boutique_id = 1"]
    params: list = []
    if date_debut:
        clauses.append("date_ca >= ?")
        params.append(date_debut)
    if date_fin:
        clauses.append("date_ca <= ?")
        params.append(date_fin)
    where = " AND ".join(clauses)
    async with get_db() as db:
        cur = await db.execute(
            f"""SELECT c.*, p.prenom AS personnel_prenom
                FROM ca_journalier c
                LEFT JOIN personnel p ON p.id = c.personnel_id
                WHERE {where}
                ORDER BY c.date_ca DESC
                LIMIT ?""",
            (*params, limit),
        )
        rows = await cur.fetchall()
        return [_enrichir_ca(r) for r in rows]


@router.get("/pilotage/ca/stats/resume")
async def get_ca_stats():
    """Totaux & moyennes : mois en cours et 30 derniers jours glissants."""
    today = date.today()
    fin = today.isoformat()
    async with get_db() as db:
        return {
            "mois_courant": await _agg_periode(db, today.replace(day=1).isoformat(), fin),
            "trente_jours": await _agg_periode(db, (today - timedelta(days=29)).isoformat(), fin),
        }


# Préréglages de décalage pour le comparatif "période vs période décalée".
# Chaque entrée décale la période courante d'un nombre de jours fixe (j/s)
# ou via un recul calendaire (mois/année).
_COMPARATIF_PRESETS = {"j-1", "s-1", "m-1", "a-1"}


def _decaler(d: date, preset: str, n: int) -> date:
    """Recule une date selon le préréglage (j=jour, s=semaine, m=mois, a=an)."""
    unite = preset[0]
    if unite == "j":
        return d - timedelta(days=n)
    if unite == "s":
        return d - timedelta(weeks=n)
    if unite == "m":
        mois = d.month - 1 - n
        annee = d.year + mois // 12
        mois = mois % 12 + 1
        # clamp du jour pour les mois plus courts (28/30/31)
        import calendar
        jour = min(d.day, calendar.monthrange(annee, mois)[1])
        return date(annee, mois, jour)
    if unite == "a":
        try:
            return d.replace(year=d.year - n)
        except ValueError:           # 29 février → 28
            return d.replace(year=d.year - n, day=28)
    raise ValueError(f"préréglage inconnu : {preset}")


@router.get("/pilotage/ca/stats/comparatif")
async def get_ca_comparatif(
    date_debut: str = Query(..., description="Début période courante (YYYY-MM-DD)"),
    date_fin:   str = Query(..., description="Fin période courante (YYYY-MM-DD)"),
    preset:     str = Query("j-1", description="Décalage : j-1 | s-1 | m-1 | a-1"),
    n:          int = Query(1, ge=1, le=520, description="Nombre d'unités de recul (X)"),
):
    """Compare une période courante à la même période décalée (J/S/M/A × n).

    Renvoie l'agrégat courant, l'agrégat de référence, et l'évolution (€ et %)
    sur le CA total et chaque section matin/soir.
    """
    preset = preset.lower().strip()
    if preset not in _COMPARATIF_PRESETS:
        raise HTTPException(400, f"preset invalide (attendu : {', '.join(sorted(_COMPARATIF_PRESETS))})")
    try:
        d1 = date.fromisoformat(date_debut)
        d2 = date.fromisoformat(date_fin)
    except ValueError:
        raise HTTPException(400, "dates invalides (format YYYY-MM-DD)")
    if d2 < d1:
        raise HTTPException(400, "date_fin doit être ≥ date_debut")

    ref_debut = _decaler(d1, preset, n)
    ref_fin   = _decaler(d2, preset, n)

    async with get_db() as db:
        courant   = await _agg_periode(db, d1.isoformat(), d2.isoformat())
        reference = await _agg_periode(db, ref_debut.isoformat(), ref_fin.isoformat())

    return {
        "preset":    preset,
        "n":         n,
        "courant":   courant,
        "reference": reference,
        "evolution": {
            "total_ttc":       _evolution(courant["total_ttc"],       reference["total_ttc"]),
            "total_ttc_matin": _evolution(courant["total_ttc_matin"], reference["total_ttc_matin"]),
            "total_ttc_soir":  _evolution(courant["total_ttc_soir"],  reference["total_ttc_soir"]),
        },
    }


@router.get("/pilotage/ca/stats/comparer-dates")
async def get_ca_comparer_dates(
    date_a: str = Query(..., description="Première date (YYYY-MM-DD)"),
    date_b: str = Query(..., description="Seconde date (YYYY-MM-DD)"),
):
    """Compare deux jours précis (A vs B), détail matin/soir + évolution."""
    async with get_db() as db:
        async def _jour(d: str) -> dict:
            cur = await db.execute(
                "SELECT * FROM ca_journalier WHERE boutique_id = 1 AND date_ca = ?", (d,)
            )
            row = await cur.fetchone()
            if row:
                return _enrichir_ca(row)
            # Jour non saisi → zéros (pour comparer quand même)
            return {
                "date_ca": d, "montant_ttc": 0, "nb_tickets": None,
                "montant_ttc_matin": 0, "montant_ttc_soir": 0,
                "panier_moyen": None, "panier_moyen_matin": None, "panier_moyen_soir": None,
                "saisi": False,
            }
        a = await _jour(date_a)
        b = await _jour(date_b)
    a.setdefault("saisi", True)
    b.setdefault("saisi", True)
    return {
        "a": a, "b": b,
        "evolution": {
            "total_ttc":       _evolution(a["montant_ttc"],       b["montant_ttc"]),
            "total_ttc_matin": _evolution(a["montant_ttc_matin"], b["montant_ttc_matin"]),
            "total_ttc_soir":  _evolution(a["montant_ttc_soir"],  b["montant_ttc_soir"]),
        },
    }


@router.get("/pilotage/ca/{date_ca}")
async def get_ca_jour(date_ca: str):
    """CA d'un jour précis, ou null si non saisi."""
    async with get_db() as db:
        cur = await db.execute(
            "SELECT * FROM ca_journalier WHERE boutique_id = 1 AND date_ca = ?",
            (date_ca,),
        )
        row = await cur.fetchone()
        return _enrichir_ca(row) if row else None


@router.post("/pilotage/ca", status_code=201)
async def upsert_ca_jour(body: CaJournalierUpsert):
    """Enregistre ou corrige le CA d'un jour, ventilé matin/soir.

    Le total du jour (montant_ttc / nb_tickets) est la somme des deux sections.
    """
    for nom, val in (("matin", body.montant_ttc_matin), ("soir", body.montant_ttc_soir)):
        if val < 0:
            raise HTTPException(400, f"Le montant {nom} ne peut pas être négatif")
    for nom, val in (("matin", body.nb_tickets_matin), ("soir", body.nb_tickets_soir)):
        if val is not None and val < 0:
            raise HTTPException(400, f"Le nombre de tickets {nom} ne peut pas être négatif")

    total_ttc = round((body.montant_ttc_matin or 0) + (body.montant_ttc_soir or 0), 2)
    # Total tickets : None seulement si aucune des deux sections n'a de tickets
    tm, ts = body.nb_tickets_matin, body.nb_tickets_soir
    total_tickets = (tm or 0) + (ts or 0) if (tm is not None or ts is not None) else None

    async with get_db() as db:
        await db.execute(
            """INSERT INTO ca_journalier
                   (boutique_id, date_ca, montant_ttc, nb_tickets,
                    montant_ttc_matin, nb_tickets_matin,
                    montant_ttc_soir, nb_tickets_soir,
                    commentaire, personnel_id)
               VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(boutique_id, date_ca) DO UPDATE SET
                   montant_ttc       = excluded.montant_ttc,
                   nb_tickets        = excluded.nb_tickets,
                   montant_ttc_matin = excluded.montant_ttc_matin,
                   nb_tickets_matin  = excluded.nb_tickets_matin,
                   montant_ttc_soir  = excluded.montant_ttc_soir,
                   nb_tickets_soir   = excluded.nb_tickets_soir,
                   commentaire       = excluded.commentaire,
                   personnel_id      = excluded.personnel_id,
                   updated_at        = CURRENT_TIMESTAMP""",
            (body.date_ca, total_ttc, total_tickets,
             body.montant_ttc_matin, body.nb_tickets_matin,
             body.montant_ttc_soir, body.nb_tickets_soir,
             body.commentaire, body.personnel_id),
        )
        await db.commit()
        cur = await db.execute(
            "SELECT * FROM ca_journalier WHERE boutique_id = 1 AND date_ca = ?",
            (body.date_ca,),
        )
        return _enrichir_ca(await cur.fetchone())
