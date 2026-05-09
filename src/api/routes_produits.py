"""
routes_produits.py — Catalogue produits (matières premières & produits finis)

GET    /api/produits                        → liste catalogue (filtres: type, en_stock, inclure_inactifs)
GET    /api/produits/categories             → distinct categories utilisées
GET    /api/produits/{id}                   → un produit
POST   /api/produits                        → créer produit
PUT    /api/produits/{id}                   → modifier produit
DELETE /api/produits/{id}                   → désactiver produit (soft delete)
GET    /api/produits/export/xlsx            → export Excel du catalogue
POST   /api/produits/import/xlsx            → import Excel (mode: replace | merge)
"""

import io
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from src.database import (
    get_db,
    get_produits, get_produit, create_produit, update_produit,
)

router = APIRouter(prefix="/api", tags=["produits"])

BOUTIQUE_ID = 1  # mono-boutique Phase 2

# Champs exportés/importés via Excel (ordre = ordre des colonnes dans le fichier)
COLONNES_EXPORT = (
    "id", "nom", "code_unique", "espece",
    "conditionnement", "categorie", "dlc_jours", "temperature_conservation",
    "type_produit", "actif",
)


# ---------------------------------------------------------------------------
# Schémas Pydantic
# ---------------------------------------------------------------------------

class ProduitCreate(BaseModel):
    nom: str = Field(..., min_length=1)
    categorie: str = Field(..., min_length=1)
    dlc_jours: int = 0
    temperature_conservation: str = "0°C à +4°C"
    code_unique: Optional[str] = None
    espece: Optional[str] = None
    etape: Optional[int] = None
    coupe_niveau: Optional[str] = None
    conditionnement: Optional[str] = "SOUS_VIDE"
    format_etiquette: Optional[str] = "standard_60x40"
    type_produit: Optional[str] = "brut"


class ProduitUpdate(BaseModel):
    nom: Optional[str] = None
    categorie: Optional[str] = None
    dlc_jours: Optional[int] = None
    temperature_conservation: Optional[str] = None
    code_unique: Optional[str] = None
    espece: Optional[str] = None
    etape: Optional[int] = None
    coupe_niveau: Optional[str] = None
    conditionnement: Optional[str] = None
    format_etiquette: Optional[str] = None
    type_produit: Optional[str] = None
    actif: Optional[bool] = None


# ---------------------------------------------------------------------------
# Liste / lecture
# ---------------------------------------------------------------------------

@router.get("/produits")
async def lister_produits(
    type: Optional[str] = Query(None, description="Filtrer par type : 'brut' ou 'fini'"),
    en_stock: bool = Query(False, description="Si True, retourne uniquement les produits ayant au moins un lot en stock (non périmé, non traité)"),
    inclure_inactifs: bool = Query(False, description="Inclure les produits archivés (actif=0)"),
):
    async with get_db() as db:
        if en_stock:
            # Filtre stock réel : produit doit avoir au moins une ligne de réception
            # non périmée et non traitée via le calendrier DLC.
            # On joint le lot FIFO (DLC la plus courte → date réception la plus ancienne)
            # pour exposer numero_lot, dlc et reception_ligne_id.
            _fifo_where = """
                rl2.produit_id = p.id
                AND r2.statut = 'cloturee'
                AND rl2.conforme = 1
                AND r2.livraison_refusee = 0
                AND (COALESCE(rl2.dlc, rl2.dluo) IS NULL
                     OR COALESCE(rl2.dlc, rl2.dluo) >= DATE('now'))
                AND NOT EXISTS (
                    SELECT 1 FROM dlc_devenir d
                    WHERE d.source_type = 'reception_ligne' AND d.source_id = rl2.id
                )
            """
            _fifo_order = """
                ORDER BY CASE WHEN COALESCE(rl2.dlc, rl2.dluo) IS NOT NULL THEN 0 ELSE 1 END,
                         COALESCE(rl2.dlc, rl2.dluo) ASC, r2.date_reception ASC LIMIT 1
            """
            _fifo_sub = f"""
                (SELECT rl2.numero_lot FROM reception_lignes rl2
                 JOIN receptions r2 ON r2.id = rl2.reception_id
                 WHERE {_fifo_where} {_fifo_order})
            """
            _fifo_dlc = f"""
                (SELECT rl2.dlc FROM reception_lignes rl2
                 JOIN receptions r2 ON r2.id = rl2.reception_id
                 WHERE {_fifo_where} {_fifo_order})
            """
            _fifo_dluo = f"""
                (SELECT rl2.dluo FROM reception_lignes rl2
                 JOIN receptions r2 ON r2.id = rl2.reception_id
                 WHERE {_fifo_where} {_fifo_order})
            """
            _fifo_id = f"""
                (SELECT rl2.id FROM reception_lignes rl2
                 JOIN receptions r2 ON r2.id = rl2.reception_id
                 WHERE {_fifo_where} {_fifo_order})
            """
            _exists_dispo = """
                EXISTS (
                    SELECT 1 FROM reception_lignes rl
                    JOIN receptions r ON r.id = rl.reception_id
                    WHERE rl.produit_id = p.id
                      AND r.statut = 'cloturee'
                      AND rl.conforme = 1
                      AND r.livraison_refusee = 0
                      AND (COALESCE(rl.dlc, rl.dluo) IS NULL
                           OR COALESCE(rl.dlc, rl.dluo) >= DATE('now'))
                      AND NOT EXISTS (
                          SELECT 1 FROM dlc_devenir d
                          WHERE d.source_type = 'reception_ligne' AND d.source_id = rl.id
                      )
                )
            """
            base_select = f"""
                SELECT p.*, {_fifo_sub} AS numero_lot,
                       {_fifo_dlc} AS dlc,
                       {_fifo_dluo} AS dluo,
                       {_fifo_id}  AS reception_ligne_id
                FROM produits p
                WHERE p.boutique_id = ? AND p.actif = 1
                  AND {_exists_dispo}
            """
            params: list = [BOUTIQUE_ID]
            if type:
                base_select += " AND p.type_produit = ?"
                params.append(type)
            base_select += " ORDER BY p.nom"
            cursor = await db.execute(base_select, params)
            rows = await cursor.fetchall()
            return [dict(r) for r in rows]

        return await get_produits(
            db, BOUTIQUE_ID,
            type_produit=type,
            inclure_inactifs=inclure_inactifs,
        )


@router.get("/produits/lots-disponibles")
async def lots_disponibles(
    type: Optional[str] = Query(None, description="Filtrer par type : 'brut' ou 'fini'"),
):
    """Retourne UN row par lot disponible en stock (avec infos produit jointes).

    À la différence de /produits?en_stock=true qui agrège un lot FIFO par produit,
    cet endpoint expose tous les lots éligibles, classés par DLC ascendante.
    Utilisé par la modale de substitution pour permettre le choix d'un lot précis."""
    sql = """
        SELECT
            p.id           AS id,
            p.nom          AS nom,
            p.code_unique  AS code_unique,
            p.type_produit AS type_produit,
            rl.id          AS reception_ligne_id,
            rl.numero_lot  AS numero_lot,
            rl.dlc         AS dlc,
            rl.dluo        AS dluo,
            rl.poids_kg    AS poids_kg,
            r.date_reception AS date_reception
        FROM reception_lignes rl
        JOIN receptions r ON r.id = rl.reception_id
        JOIN produits   p ON p.id = rl.produit_id
        WHERE p.boutique_id = ? AND p.actif = 1
          AND r.statut = 'cloturee'
          AND rl.conforme = 1
          AND r.livraison_refusee = 0
          AND (COALESCE(rl.dlc, rl.dluo) IS NULL
               OR COALESCE(rl.dlc, rl.dluo) >= DATE('now'))
          AND NOT EXISTS (
              SELECT 1 FROM dlc_devenir d
              WHERE d.source_type = 'reception_ligne' AND d.source_id = rl.id
          )
    """
    params: list = [BOUTIQUE_ID]
    if type:
        sql += " AND p.type_produit = ?"
        params.append(type)
    sql += """
        ORDER BY CASE WHEN COALESCE(rl.dlc, rl.dluo) IS NOT NULL THEN 0 ELSE 1 END,
                 COALESCE(rl.dlc, rl.dluo) ASC,
                 r.date_reception ASC
    """
    async with get_db() as db:
        cursor = await db.execute(sql, params)
        rows = await cursor.fetchall()
        return [dict(r) for r in rows]


@router.get("/produits/categories")
async def lister_categories(inclure_inactifs: bool = Query(False)):
    """Liste les catégories distinctes utilisées dans le catalogue."""
    where = "boutique_id = ? AND categorie IS NOT NULL"
    if not inclure_inactifs:
        where += " AND actif = 1"
    async with get_db() as db:
        cursor = await db.execute(
            f"SELECT DISTINCT categorie FROM produits WHERE {where} ORDER BY categorie",
            (BOUTIQUE_ID,),
        )
        rows = await cursor.fetchall()
    return [r[0] for r in rows]


@router.get("/produits/{produit_id}")
async def obtenir_produit(produit_id: int):
    async with get_db() as db:
        produit = await get_produit(db, produit_id)
    if not produit:
        raise HTTPException(404, "Produit non trouvé")
    return produit


# ---------------------------------------------------------------------------
# Création / modification / suppression
# ---------------------------------------------------------------------------

@router.post("/produits", status_code=201)
async def nouveau_produit(body: ProduitCreate):
    payload = {"boutique_id": BOUTIQUE_ID, **body.model_dump(exclude_none=True)}
    async with get_db() as db:
        # Vérifier unicité du code_unique si fourni
        code = payload.get("code_unique")
        if code:
            cur = await db.execute(
                "SELECT id FROM produits WHERE code_unique = ?", (code,)
            )
            if await cur.fetchone():
                raise HTTPException(409, f"Code unique '{code}' déjà utilisé")
        produit_id = await create_produit(db, payload)
        produit = await get_produit(db, produit_id)
    return produit


@router.put("/produits/{produit_id}")
async def modifier_produit(produit_id: int, body: ProduitUpdate):
    payload = body.model_dump(exclude_none=True)
    if not payload:
        raise HTTPException(400, "Aucun champ à modifier")
    async with get_db() as db:
        # Vérifier unicité du code_unique si modifié
        code = payload.get("code_unique")
        if code:
            cur = await db.execute(
                "SELECT id FROM produits WHERE code_unique = ? AND id != ?",
                (code, produit_id),
            )
            if await cur.fetchone():
                raise HTTPException(409, f"Code unique '{code}' déjà utilisé")
        ok = await update_produit(db, produit_id, payload)
        if not ok:
            raise HTTPException(404, "Produit non trouvé ou aucun champ valide")
        produit = await get_produit(db, produit_id)
    return produit


@router.delete("/produits/{produit_id}", status_code=204)
async def supprimer_produit(produit_id: int):
    """Soft delete : passe actif=0. Le produit reste lié aux historiques."""
    async with get_db() as db:
        await update_produit(db, produit_id, {"actif": False})


# ---------------------------------------------------------------------------
# Import / Export Excel
# ---------------------------------------------------------------------------

@router.get("/produits/export/xlsx")
async def exporter_xlsx():
    """Exporte tout le catalogue (incl. archivés) au format Excel."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installé sur le serveur")

    async with get_db() as db:
        cursor = await db.execute(
            "SELECT * FROM produits WHERE boutique_id = ? ORDER BY nom",
            (BOUTIQUE_ID,),
        )
        rows = [dict(r) for r in await cursor.fetchall()]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catalogue"
    ws.append(list(COLONNES_EXPORT))
    for r in rows:
        ws.append([r.get(c) for c in COLONNES_EXPORT])

    # Largeurs colonnes raisonnables
    largeurs = {
        "id": 6, "nom": 38, "code_unique": 14, "espece": 14,
        "conditionnement": 14, "categorie": 20,
        "dlc_jours": 10, "temperature_conservation": 22,
        "type_produit": 12, "actif": 8,
    }
    for idx, c in enumerate(COLONNES_EXPORT, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = largeurs.get(c, 14)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from datetime import date
    filename = f"catalogue_produits_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/produits/template/xlsx")
async def telecharger_template():
    """Retourne un fichier Excel vide avec les bonnes colonnes et une ligne exemple."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl non installé sur le serveur")

    colonnes_template = [c for c in COLONNES_EXPORT if c != "id"]
    exemple = {
        "nom":                      "Entrecôte bovin",
        "code_unique":              "VB1",
        "espece":                   "bovin",
        "conditionnement":          "SOUS_VIDE",
        "categorie":                "viande_pieces",
        "dlc_jours":                3,
        "temperature_conservation": "0°C à +4°C",
        "type_produit":             "brut",
        "actif":                    1,
    }
    largeurs = {
        "nom": 38, "code_unique": 14, "espece": 14,
        "conditionnement": 14, "categorie": 20,
        "dlc_jours": 10, "temperature_conservation": 22,
        "type_produit": 12, "actif": 8,
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modèle import"

    entete_fill = PatternFill("solid", fgColor="3D2008")
    entete_font = Font(bold=True, color="FFFFFF")
    for idx, col in enumerate(colonnes_template, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.fill = entete_fill
        cell.font = entete_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = largeurs.get(col, 14)

    ws.append([exemple.get(c, "") for c in colonnes_template])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modele_import_catalogue.xlsx"'},
    )


@router.post("/produits/import/xlsx")
async def importer_xlsx(
    fichier: UploadFile = File(...),
    mode: str = Form("merge", description="merge = upsert par code_unique; replace = vide d'abord"),
):
    """Import Excel.
    - mode=merge (défaut) : upsert par code_unique (création si nouveau, MAJ sinon).
    - mode=replace : désactive tous les produits existants puis insère les nouveaux.
    Le fichier doit avoir comme première ligne les en-têtes correspondant aux noms de colonnes.
    """
    if mode not in ("merge", "replace"):
        raise HTTPException(400, "mode doit être 'merge' ou 'replace'")
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl non installé sur le serveur")

    contenu = await fichier.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenu), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Fichier Excel invalide : {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Fichier vide")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    # Champs valides à importer (id ignoré, géré automatiquement)
    champs_valides = set(COLONNES_EXPORT) - {"id"}
    indices = {h: i for i, h in enumerate(headers) if h in champs_valides}
    if "nom" not in indices:
        raise HTTPException(400, "Colonne 'nom' obligatoire dans l'en-tête")

    crees, mis_a_jour, ignores = 0, 0, 0
    erreurs: list[str] = []

    async with get_db() as db:
        if mode == "replace":
            await db.execute(
                "UPDATE produits SET actif = 0 WHERE boutique_id = ?",
                (BOUTIQUE_ID,),
            )
            await db.commit()

        for ligne_num, row in enumerate(rows[1:], start=2):
            data = {}
            for champ, idx in indices.items():
                val = row[idx] if idx < len(row) else None
                if val is None or (isinstance(val, str) and val.strip() == ""):
                    continue
                if champ == "dlc_jours":
                    try:
                        val = int(val)
                    except (ValueError, TypeError):
                        erreurs.append(f"Ligne {ligne_num}: '{champ}' invalide ({val!r})")
                        continue
                if champ == "actif":
                    val = bool(val) if not isinstance(val, str) else val.strip().lower() in ("1", "true", "oui", "yes", "vrai")
                data[champ] = val

            if not data.get("nom"):
                ignores += 1
                continue
            data.setdefault("categorie", "matiere_premiere")

            code = data.get("code_unique")
            existant = None
            if code:
                cur = await db.execute(
                    "SELECT id FROM produits WHERE code_unique = ?", (code,)
                )
                row_existant = await cur.fetchone()
                if row_existant:
                    existant = row_existant[0]

            try:
                if existant:
                    await update_produit(db, existant, {**data, "actif": True})
                    mis_a_jour += 1
                else:
                    await create_produit(db, {"boutique_id": BOUTIQUE_ID, **data})
                    crees += 1
            except Exception as e:
                erreurs.append(f"Ligne {ligne_num}: {e}")

    return {
        "mode": mode,
        "crees": crees,
        "mis_a_jour": mis_a_jour,
        "ignores": ignores,
        "erreurs": erreurs,
    }
